import argparse
import json
import os
import re
import shutil
import sys
import time
from collections import defaultdict
from pathlib import Path

import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))
os.environ.setdefault("DJANGO_SETTINGS_MODULE", "config.settings")

import django

django.setup()

from django.db import transaction
from django.db.models import Q
from django.utils import timezone

from properties.models import GeocodeCache, Listing, Property, PropertyLocation, PropertyLocationIntelligence
from properties.scrapers.local_wordpress import _extract_wordpress_map_data
from properties.services.geocoding import Geocoder, best_address
from properties.services.normalization import (
    normalize_address,
    normalize_locality,
    normalize_neighborhood_name,
    normalize_whitespace,
)
from properties.services.territory_hierarchy import (
    infer_territory_for_point,
    territory_values_from_result,
)


INPUT_XLSX = ROOT / "tmp" / "curacion_zonas_pendientes_2026-06-18.xlsx"
OUTPUT_XLSX = ROOT / "tmp" / "curacion_zonas_pendientes_2026-06-19.xlsx"
FETCH_CACHE = ROOT / "tmp" / "curacion_map_cache_20260619.json"
DB_PATH = ROOT / "db.sqlite3"
BACKUP_PATH = ROOT / "tmp" / "db.sqlite3.backup_curacion_20260619"

MIGLIERINI_DUPLICATE_COORDS = (-34.5933423, -58.6378405)
MIGLIERINI_DELAY = 5
ODRIOZOLA_DELAY = 3


def no_zone_query():
    return (
        (Q(inferred_zone__isnull=True) | Q(inferred_zone=""))
        & (
            Q(location_intelligence__isnull=True)
            | Q(location_intelligence__zone_name__isnull=True)
            | Q(location_intelligence__zone_name="")
        )
    )


def load_manual_feedback():
    wb = load_workbook(INPUT_XLSX, read_only=True, data_only=False)
    ws = wb["Pendientes"]
    headers = [cell.value for cell in ws[1]]
    rows = []
    for values in ws.iter_rows(min_row=2, values_only=True):
        row = dict(zip(headers, values))
        decision = normalize_whitespace(str(row.get("decision_manual") or "")).upper()
        if decision == "APROBADO":
            rows.append(row)
    return rows


def safe(value):
    if value is None:
        return ""
    return str(value).strip()


def short_address(address):
    value = normalize_whitespace(address or "")
    if not value:
        return ""
    return normalize_whitespace(value.split(",", 1)[0])


def source_listing(property_obj, source_slug):
    return (
        Listing.objects.filter(property=property_obj, source__slug=source_slug, active=True)
        .order_by("-last_seen_at")
        .first()
        or Listing.objects.filter(property=property_obj, source__slug=source_slug)
        .order_by("-last_seen_at")
        .first()
    )


class Fetcher:
    def __init__(self, refresh=False):
        self.refresh = refresh
        self.session = requests.Session()
        self.session.headers.update(
            {
                "User-Agent": (
                    "Mozilla/5.0 (compatible; RadarInmobiliarioCuration/1.0; "
                    "+https://local)"
                )
            }
        )
        self.cache = {}
        self.last_request = defaultdict(float)
        if FETCH_CACHE.exists() and not refresh:
            self.cache = json.loads(FETCH_CACHE.read_text(encoding="utf-8"))

    def save(self):
        FETCH_CACHE.write_text(
            json.dumps(self.cache, ensure_ascii=False, indent=2, sort_keys=True),
            encoding="utf-8",
        )

    def map_data(self, url, source_slug, delay):
        if not url:
            return {"error": "sin_url"}
        if url in self.cache and not self.refresh:
            return self.cache[url]

        elapsed = time.monotonic() - self.last_request[source_slug]
        if elapsed < delay:
            time.sleep(delay - elapsed)
        try:
            response = self.session.get(url, timeout=35)
            self.last_request[source_slug] = time.monotonic()
            response.raise_for_status()
            soup = BeautifulSoup(response.text, "lxml")
            data = _extract_wordpress_map_data(soup)
            title = soup.get_text(" ", strip=True)[:240]
            result = {
                "latitude": data.get("latitude"),
                "longitude": data.get("longitude"),
                "address": data.get("address") or "",
                "status": response.status_code,
                "title_sample": title,
                "error": "",
            }
        except Exception as exc:
            result = {"error": repr(exc)}
        self.cache[url] = result
        self.save()
        return result


def current_counts():
    props = list(Property.objects.select_related("location_intelligence"))
    no_zone = Property.objects.filter(no_zone_query()).count()
    no_address_ids = [p.pk for p in props if not best_address(p)]
    return {
        "sin_zona_operativa": no_zone,
        "sin_direccion_util": len(no_address_ids),
        "sin_direccion_util_ids": no_address_ids,
    }


def set_manual_field(property_obj, field, value, now, changes, manual=True):
    value = normalize_whitespace(value or "")
    if not value or getattr(property_obj, field) == value:
        return
    setattr(property_obj, field, value)
    changes.add(field)
    if field == "address":
        property_obj.normalized_address = normalize_address(value)
        changes.add("normalized_address")
    if manual:
        overrides = dict(property_obj.manual_overrides or {})
        overrides[field] = now.isoformat()
        property_obj.manual_overrides = overrides
        property_obj.data_manually_corrected_at = now
        changes.update({"manual_overrides", "data_manually_corrected_at"})


def apply_property_changes(property_obj, changes):
    if changes:
        property_obj.save(update_fields=sorted(changes))


def update_source_location(property_obj, lat, lon, provider, query, apply):
    current = getattr(property_obj, "location", None)
    if current and current.manually_corrected:
        return False, "manual_location_preserved"
    if lat is None or lon is None:
        return False, "no_coordinates"
    if not apply:
        return True, "would_update"
    PropertyLocation.objects.update_or_create(
        property=property_obj,
        defaults={
            "latitude": float(lat),
            "longitude": float(lon),
            "precision": PropertyLocation.Precision.EXACT,
            "query": query or "",
            "provider": provider,
            "confidence": 1,
            "outside_target": False,
            "manually_corrected": False,
        },
    )
    return True, "updated"


def source_zone_for_inference(property_obj):
    for value in (property_obj.neighborhood, property_obj.detected_neighborhood):
        normalized = normalize_neighborhood_name(value)
        if normalized:
            return normalized
    return ""


def source_locality_for_inference(property_obj):
    for value in (property_obj.detected_locality, property_obj.locality):
        normalized = normalize_locality(value)
        if normalized:
            return normalized
    return ""


def infer_and_apply_territory(property_obj, apply, tag, allow_reviewed_polygon=False):
    location = getattr(property_obj, "location", None)
    if not location:
        return {"applied": False, "zone": "", "reason": "sin_coordenadas"}
    result = infer_territory_for_point(
        location.latitude,
        location.longitude,
        coordinate_source=location.provider,
        extra_evidence={
            "provider": location.provider,
            "precision": location.precision,
            "curacion": tag,
        },
        source_zone=source_zone_for_inference(property_obj),
        source_locality=source_locality_for_inference(property_obj),
    )
    values = territory_values_from_result(result)
    if not values["zone"]:
        return {"applied": False, "zone": "", "reason": "sin_match_poligono", "values": values}
    if values["needs_review"] and not allow_reviewed_polygon:
        return {
            "applied": False,
            "zone": values["zone"],
            "reason": "conflicto_o_revision",
            "values": values,
        }
    if not apply:
        return {"applied": True, "zone": values["zone"], "reason": "would_apply", "values": values}

    now = timezone.now()
    property_obj.inferred_partido = values["partido"]
    property_obj.inferred_locality = values["locality"]
    property_obj.inferred_zone = values["zone"]
    property_obj.territory_confidence = values["confidence"]
    property_obj.territory_source_method = values["source_method"]
    property_obj.territory_needs_review = values["needs_review"]
    property_obj.territory_evidence = values["evidence"]
    property_obj.territory_inferred_at = now
    property_obj.inferred_neighborhood = values["zone"]
    property_obj.zone_needs_review = bool(values["needs_review"])
    property_obj.zone_conflict = False
    property_obj.zone_inference_evidence = {
        **(property_obj.zone_inference_evidence or {}),
        "curacion_20260619": values["evidence"],
    }
    property_obj.zone_inferred_at = now
    property_obj.save(
        update_fields=[
            "inferred_partido",
            "inferred_locality",
            "inferred_zone",
            "territory_confidence",
            "territory_source_method",
            "territory_needs_review",
            "territory_evidence",
            "territory_inferred_at",
            "inferred_neighborhood",
            "zone_needs_review",
            "zone_conflict",
            "zone_inference_evidence",
            "zone_inferred_at",
        ]
    )
    record, _ = PropertyLocationIntelligence.objects.get_or_create(property=property_obj)
    record.partido_name = values["partido"]
    record.locality_name = values["locality"]
    record.zone_name = values["zone"]
    record.match_method = PropertyLocationIntelligence.MatchMethod.COORDINATES
    record.confidence = values["confidence"] or record.confidence
    record.evidence = {**(record.evidence or {}), "curacion_20260619": values["evidence"]}
    record.scored_at = now
    record.save(
        update_fields=[
            "partido_name",
            "locality_name",
            "zone_name",
            "match_method",
            "confidence",
            "evidence",
            "scored_at",
        ]
    )
    return {"applied": True, "zone": values["zone"], "reason": "applied", "values": values}


def infer_zone_preview(property_obj, latitude, longitude, provider, tag):
    result = infer_territory_for_point(
        latitude,
        longitude,
        coordinate_source=provider,
        extra_evidence={
            "provider": provider,
            "precision": "exact",
            "curacion": tag,
        },
        source_zone=source_zone_for_inference(property_obj),
        source_locality=source_locality_for_inference(property_obj),
    )
    values = territory_values_from_result(result)
    if not values["zone"]:
        return {"applied": False, "zone": "", "reason": "sin_match_poligono", "values": values}
    if values["needs_review"]:
        return {"applied": False, "zone": values["zone"], "reason": "conflicto_o_revision", "values": values}
    return {"applied": True, "zone": values["zone"], "reason": "would_apply", "values": values}


def apply_geocode_cache_or_api(property_obj, geocoder):
    current = getattr(property_obj, "location", None)
    if current and current.manually_corrected:
        return current, "manual"
    queries = geocoder.query_candidates(property_obj)
    if not queries:
        return None, "sin_consulta"

    for query in queries:
        cache = GeocodeCache.objects.filter(query=query).first()
        if cache:
            location = geocoder._apply(property_obj, query, cache)
            if location:
                return location, "cache"

    for query in queries:
        if GeocodeCache.objects.filter(query=query).exists():
            continue
        cache = geocoder._fetch(query, property_obj)
        location = geocoder._apply(property_obj, query, cache)
        if location:
            return location, "api"
    return None, "sin_resultado_api"


def apply_manual_feedback(rows, apply):
    now = timezone.now()
    summary = {
        "approved_ids": [],
        "address_updates": [],
        "locality_updates": [],
        "neighborhood_updates": [],
        "missing_ids": [],
    }
    for row in rows:
        property_id = int(row["id_propiedad"])
        summary["approved_ids"].append(property_id)
        try:
            property_obj = Property.objects.get(pk=property_id)
        except Property.DoesNotExist:
            summary["missing_ids"].append(property_id)
            continue

        changes = set()
        before_address = property_obj.address
        before_locality = property_obj.locality
        before_neighborhood = property_obj.neighborhood
        set_manual_field(property_obj, "address", safe(row.get("domicilio_actual")), now, changes)
        set_manual_field(property_obj, "locality", safe(row.get("localidad_actual")), now, changes)
        set_manual_field(property_obj, "neighborhood", safe(row.get("barrio_actual")), now, changes)
        if apply:
            apply_property_changes(property_obj, changes)
        if "address" in changes:
            summary["address_updates"].append(
                {"id": property_id, "from": before_address, "to": property_obj.address}
            )
        if "locality" in changes:
            summary["locality_updates"].append(
                {"id": property_id, "from": before_locality, "to": property_obj.locality}
            )
        if "neighborhood" in changes:
            summary["neighborhood_updates"].append(
                {"id": property_id, "from": before_neighborhood, "to": property_obj.neighborhood}
            )
    return summary


def process_miglierini(fetcher, apply):
    qs = (
        Property.objects.filter(listings__source__slug="miglierini")
        .distinct()
        .select_related("location", "location_intelligence")
    )
    missing_address = [p for p in qs if not safe(p.address) and not safe(p.detected_address)]
    summary = {
        "missing_address_ids": [p.pk for p in missing_address],
        "map_coords_ids": [],
        "duplicate_replaced_ids": [],
        "no_map_coords_ids": [],
        "manual_location_preserved_ids": [],
        "zone_applied_ids": [],
        "zone_conflict_ids": [],
        "no_url_ids": [],
        "errors": {},
    }
    for property_obj in missing_address:
        listing = source_listing(property_obj, "miglierini")
        if not listing:
            summary["no_url_ids"].append(property_obj.pk)
            continue
        data = fetcher.map_data(listing.url, "miglierini", MIGLIERINI_DELAY)
        lat = data.get("latitude")
        lon = data.get("longitude")
        if data.get("error"):
            summary["errors"][property_obj.pk] = data["error"]
        if lat is None or lon is None:
            summary["no_map_coords_ids"].append(property_obj.pk)
            continue
        summary["map_coords_ids"].append(property_obj.pk)
        current = getattr(property_obj, "location", None)
        duplicate = (
            current
            and not current.manually_corrected
            and abs(current.latitude - MIGLIERINI_DUPLICATE_COORDS[0]) < 0.000001
            and abs(current.longitude - MIGLIERINI_DUPLICATE_COORDS[1]) < 0.000001
        )
        changed, reason = update_source_location(
            property_obj, lat, lon, "miglierini_map", listing.url, apply
        )
        if reason == "manual_location_preserved":
            summary["manual_location_preserved_ids"].append(property_obj.pk)
            continue
        if duplicate and changed:
            summary["duplicate_replaced_ids"].append(property_obj.pk)
        if apply:
            property_obj = Property.objects.select_related("location").get(pk=property_obj.pk)
        zone_result = infer_and_apply_territory(
            property_obj,
            apply,
            "miglierini_map",
            allow_reviewed_polygon=True,
        )
        if zone_result["applied"]:
            summary["zone_applied_ids"].append(property_obj.pk)
        elif zone_result["reason"] == "conflicto_o_revision":
            summary["zone_conflict_ids"].append(property_obj.pk)
    return summary


def process_odriozola(fetcher, apply):
    qs = (
        Property.objects.filter(listings__source__slug="odriozola")
        .distinct()
        .select_related("location", "location_intelligence")
    )
    candidates = [p for p in qs if not safe(p.detected_address) or not safe(p.detected_locality)]
    now = timezone.now()
    summary = {
        "candidate_ids": [p.pk for p in candidates],
        "missing_address_ids": [p.pk for p in qs if not safe(p.address) and not safe(p.detected_address)],
        "repaired_ids": [],
        "map_coords_ids": [],
        "no_map_address_ids": [],
        "no_map_coords_ids": [],
        "zone_applied_ids": [],
        "zone_conflict_ids": [],
        "no_url_ids": [],
        "errors": {},
    }
    for property_obj in candidates:
        listing = source_listing(property_obj, "odriozola")
        if not listing:
            summary["no_url_ids"].append(property_obj.pk)
            continue
        data = fetcher.map_data(listing.url, "odriozola", ODRIOZOLA_DELAY)
        if data.get("error"):
            summary["errors"][property_obj.pk] = data["error"]
        full_address = safe(data.get("address"))
        lat = data.get("latitude")
        lon = data.get("longitude")
        if not full_address:
            summary["no_map_address_ids"].append(property_obj.pk)
        if lat is None or lon is None:
            summary["no_map_coords_ids"].append(property_obj.pk)

        changes = set()
        if full_address:
            if not safe(property_obj.address):
                set_manual_field(property_obj, "address", short_address(full_address), now, changes)
            set_manual_field(property_obj, "detected_address", short_address(full_address), now, changes, manual=False)
            if re.search(r"\bVilla\s+Tesei\b", full_address, re.I):
                set_manual_field(property_obj, "locality", "Villa Tesei", now, changes)
                set_manual_field(property_obj, "detected_locality", "Villa Tesei", now, changes, manual=False)
            elif re.search(r"\bHurlingham\b", full_address, re.I):
                set_manual_field(property_obj, "locality", "Hurlingham", now, changes)
                set_manual_field(property_obj, "detected_locality", "Hurlingham", now, changes, manual=False)
            evidence = dict(property_obj.location_evidence or {})
            evidence["odriozola_data_map_address"] = full_address
            property_obj.location_evidence = evidence
            changes.add("location_evidence")
        if apply:
            apply_property_changes(property_obj, changes)
        if full_address:
            summary["repaired_ids"].append(property_obj.pk)
        if lat is not None and lon is not None:
            summary["map_coords_ids"].append(property_obj.pk)
            update_source_location(property_obj, lat, lon, "odriozola_map", listing.url, apply)
        if apply:
            property_obj = Property.objects.select_related("location").get(pk=property_obj.pk)
        zone_result = infer_and_apply_territory(property_obj, apply, "odriozola_map")
        if zone_result["applied"]:
            summary["zone_applied_ids"].append(property_obj.pk)
        elif zone_result["reason"] == "conflicto_o_revision":
            summary["zone_conflict_ids"].append(property_obj.pk)
    return summary


def process_argenprop(apply):
    geocoder = Geocoder()
    qs = (
        Property.objects.filter(listings__source__slug="argenprop")
        .filter(no_zone_query())
        .distinct()
        .select_related("location", "location_intelligence")
    )
    candidates = [p for p in qs if best_address(p)]
    summary = {
        "candidate_ids": [p.pk for p in candidates],
        "existing_location_ids": [],
        "cache_hit_ids": [],
        "api_required_ids": [],
        "api_geocoded_ids": [],
        "geocoded_ids": [],
        "manual_location_ids": [],
        "no_location_ids": [],
        "zone_applied_ids": [],
        "zone_conflict_ids": [],
        "errors": {},
    }
    for property_obj in candidates:
        current = getattr(property_obj, "location", None)
        if current and current.manually_corrected:
            summary["manual_location_ids"].append(property_obj.pk)
        if not apply:
            if current:
                summary["existing_location_ids"].append(property_obj.pk)
                zone_result = infer_zone_preview(
                    property_obj,
                    current.latitude,
                    current.longitude,
                    current.provider,
                    "argenprop_existing_location",
                )
            else:
                queries = geocoder.query_candidates(property_obj)
                cache = (
                    GeocodeCache.objects.filter(query__in=queries)
                    .exclude(latitude__isnull=True)
                    .exclude(longitude__isnull=True)
                    .first()
                )
                if not cache:
                    summary["api_required_ids"].append(property_obj.pk)
                    summary["no_location_ids"].append(property_obj.pk)
                    continue
                summary["cache_hit_ids"].append(property_obj.pk)
                zone_result = infer_zone_preview(
                    property_obj,
                    cache.latitude,
                    cache.longitude,
                    "geocode_cache",
                    "argenprop_geocode_cache",
                )
            if zone_result["applied"]:
                summary["zone_applied_ids"].append(property_obj.pk)
            elif zone_result["reason"] == "conflicto_o_revision":
                summary["zone_conflict_ids"].append(property_obj.pk)
            continue

        try:
            location, geocode_status = apply_geocode_cache_or_api(property_obj, geocoder)
            property_obj = Property.objects.select_related("location").get(pk=property_obj.pk)
        except Exception as exc:
            summary["errors"][property_obj.pk] = repr(exc)
            continue
        if location:
            summary["geocoded_ids"].append(property_obj.pk)
            if geocode_status == "cache":
                summary["cache_hit_ids"].append(property_obj.pk)
            elif geocode_status == "api":
                summary["api_geocoded_ids"].append(property_obj.pk)
        else:
            summary["no_location_ids"].append(property_obj.pk)
            continue
        zone_result = infer_and_apply_territory(property_obj, apply, "argenprop_geocoding")
        if zone_result["applied"]:
            summary["zone_applied_ids"].append(property_obj.pk)
        elif zone_result["reason"] == "conflicto_o_revision":
            summary["zone_conflict_ids"].append(property_obj.pk)
    return summary


def first_active_listing(property_obj):
    return (
        property_obj.listings.filter(active=True).select_related("source", "agency").order_by("-last_seen_at").first()
        or property_obj.listings.select_related("source", "agency").order_by("-last_seen_at").first()
    )


def source_names(property_obj):
    names = []
    for listing in property_obj.listings.select_related("source").all():
        if listing.source.name not in names:
            names.append(listing.source.name)
    return ", ".join(names)


def agency_names(property_obj):
    names = []
    for listing in property_obj.listings.select_related("agency").all():
        if listing.agency and listing.agency.name not in names:
            names.append(listing.agency.name)
    return ", ".join(names)


def residual_reason(property_obj):
    reasons = []
    intel_zone = getattr(getattr(property_obj, "location_intelligence", None), "zone_name", "")
    if not property_obj.inferred_zone and not intel_zone:
        reasons.append("sin zona")
    if not best_address(property_obj):
        reasons.append("sin direccion util")
    return ", ".join(reasons)


def generate_residual_excel(path):
    props = list(
        Property.objects.select_related("location", "location_intelligence")
        .prefetch_related("listings__source", "listings__agency")
        .all()
    )
    pending = [
        p
        for p in props
        if residual_reason(p)
    ]
    wb = Workbook()
    ws = wb.active
    ws.title = "Pendientes"
    headers = [
        "id_propiedad",
        "titulo",
        "descripcion_resumen",
        "domicilio_actual",
        "direccion_detectada",
        "direccion_normalizada",
        "localidad_actual",
        "localidad_detectada",
        "barrio_actual",
        "barrio_detectado",
        "zona_actual",
        "zona_inteligencia",
        "latitud",
        "longitud",
        "fuente_coordenadas",
        "fuente",
        "inmobiliaria",
        "url_publicacion",
        "motivo_url",
        "estado_publicacion",
        "tipo_propiedad",
        "operacion",
        "moneda",
        "precio",
        "ambientes",
        "dormitorios",
        "banos",
        "superficie_cubierta",
        "superficie_total",
        "fecha_ultima_vista",
        "motivo_pendiente",
        "evidencia",
        "decision_manual",
        "notas_manual",
    ]
    ws.append(headers)
    for property_obj in sorted(pending, key=lambda p: (residual_reason(p), p.pk)):
        listing = first_active_listing(property_obj)
        location = getattr(property_obj, "location", None)
        intel = getattr(property_obj, "location_intelligence", None)
        description = normalize_whitespace(property_obj.description or "")
        evidence = {
            "location_precision": location.precision if location else "",
            "location_query": location.query if location else "",
            "manual_overrides": property_obj.manual_overrides or {},
            "source_locality": property_obj.locality or property_obj.detected_locality or "",
            "source_zone": property_obj.neighborhood or property_obj.detected_neighborhood or "",
            "territory_evidence": property_obj.territory_evidence or {},
        }
        ws.append(
            [
                property_obj.pk,
                property_obj.title,
                description[:500],
                property_obj.address,
                property_obj.detected_address,
                property_obj.normalized_address,
                property_obj.locality,
                property_obj.detected_locality,
                property_obj.neighborhood,
                property_obj.detected_neighborhood,
                property_obj.inferred_zone or property_obj.inferred_neighborhood,
                intel.zone_name if intel else "",
                location.latitude if location else None,
                location.longitude if location else None,
                location.provider if location else "",
                source_names(property_obj),
                agency_names(property_obj),
                listing.url if listing else "",
                "" if listing else "Sin Listing asociado en la base",
                property_obj.get_status_display(),
                property_obj.get_property_type_display(),
                property_obj.operation,
                property_obj.currency,
                str(property_obj.price) if property_obj.price is not None else "",
                property_obj.rooms,
                property_obj.bedrooms,
                str(property_obj.bathrooms) if property_obj.bathrooms is not None else "",
                str(property_obj.covered_area) if property_obj.covered_area is not None else "",
                str(property_obj.total_area) if property_obj.total_area is not None else "",
                property_obj.last_seen_at.isoformat() if property_obj.last_seen_at else "",
                residual_reason(property_obj),
                json.dumps(evidence, ensure_ascii=False, default=str),
                "",
                "",
            ]
        )

    summary = wb.create_sheet("Resumen")
    counts = current_counts()
    rows = [
        ("fecha", timezone.now().isoformat()),
        ("pendientes_total", len(pending)),
        ("sin_zona_operativa", counts["sin_zona_operativa"]),
        ("sin_direccion_util", counts["sin_direccion_util"]),
        ("archivo_origen", str(INPUT_XLSX)),
    ]
    for row in rows:
        summary.append(row)

    header_fill = PatternFill("solid", fgColor="1F4E78")
    for sheet in (ws, summary):
        for cell in sheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        sheet.freeze_panes = "A2"
    for column in range(1, ws.max_column + 1):
        letter = get_column_letter(column)
        max_len = max(len(str(ws.cell(row=row, column=column).value or "")) for row in range(1, min(ws.max_row, 80) + 1))
        ws.column_dimensions[letter].width = min(max(max_len + 2, 12), 48)
    wb.save(path)
    return {"path": str(path), "pending_rows": len(pending), "columns": headers}


def validate_excel(path):
    wb = load_workbook(path, read_only=True)
    ws = wb["Pendientes"]
    headers = [cell.value for cell in ws[1]]
    required = {"id_propiedad", "motivo_pendiente", "motivo_url", "decision_manual", "notas_manual"}
    return {
        "sheets": wb.sheetnames,
        "rows": ws.max_row - 1,
        "columns": headers,
        "missing_required_columns": sorted(required - set(headers)),
    }


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--apply", action="store_true")
    parser.add_argument("--generate-excel", action="store_true")
    parser.add_argument("--refresh-fetch-cache", action="store_true")
    args = parser.parse_args()
    apply = args.apply

    rows = load_manual_feedback()
    before = current_counts()
    fetcher = Fetcher(refresh=args.refresh_fetch_cache)

    if apply and not BACKUP_PATH.exists():
        shutil.copy2(DB_PATH, BACKUP_PATH)

    with transaction.atomic():
        feedback = apply_manual_feedback(rows, apply)
        miglierini = process_miglierini(fetcher, apply)
        odriozola = process_odriozola(fetcher, apply)
        argenprop = process_argenprop(apply)
        if not apply:
            transaction.set_rollback(True)

    excel = None
    validation = None
    if args.generate_excel:
        excel = generate_residual_excel(OUTPUT_XLSX)
        validation = validate_excel(OUTPUT_XLSX)

    after = current_counts()
    result = {
        "mode": "apply" if apply else "dry_run",
        "input_excel": str(INPUT_XLSX),
        "output_excel": str(OUTPUT_XLSX) if args.generate_excel else "",
        "backup": str(BACKUP_PATH) if apply else "",
        "manual_feedback": feedback,
        "before": before,
        "after": after,
        "miglierini": miglierini,
        "odriozola": odriozola,
        "argenprop": argenprop,
        "excel": excel,
        "excel_validation": validation,
    }
    print(json.dumps(result, ensure_ascii=False, indent=2, default=str))


if __name__ == "__main__":
    main()
