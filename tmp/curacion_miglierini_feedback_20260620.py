import argparse
import json
import os
import shutil
import sys
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo


ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))
os.environ.setdefault("DJANGO_SETTINGS_MODULE", "config.settings")

import django

django.setup()

from django.db import transaction
from django.db.models import Q
from django.utils import timezone

from properties.models import (
    GeocodeCache,
    Listing,
    Property,
    PropertyLocation,
    PropertyLocationIntelligence,
)
from properties.services.geocoding import Geocoder, best_address
from properties.services.normalization import (
    normalize_address,
    normalize_locality,
    normalize_neighborhood_name,
    normalize_whitespace,
)
from properties.services.territory_hierarchy import (
    infer_territory_for_point,
    territory_values_from_result,
)
from properties.services.zone_names import canonicalize_unified_zone_name


INPUT_XLSX = ROOT / "tmp" / "curacion_zonas_pendientes_2026-06-19_filtros.xlsx"
OUTPUT_XLSX = ROOT / "tmp" / "curacion_zonas_pendientes_2026-06-20.xlsx"
DB_PATH = ROOT / "db.sqlite3"
BACKUP_PATH = ROOT / "tmp" / "db.sqlite3.backup_curacion_miglierini_20260620"

SCRIPT_TAG = "curacion_20260620"

APPROVED_IDS = {3368, 3487, 3612, 3674, 3788, 3823, 3829}
APPROVED_BY_NOTE_IDS = {3325}
NON_CARTERO_APPROVED_IDS = {3368, 3487, 3612, 3674, 3788, 3823}
CARTERO_IDS = {3325, 3829}
CLEAR_MIAMI_IDS = {3301, 3311, 3325, 3330, 3356}
MIAMI_DEFAULT = (25.7308309, -80.444149)

PROPAGATED_ADDRESSES = {
    3606: {
        "address": "Tte. Gral. Julio Argentino Roca 1400",
        "locality": "Hurlingham",
        "reason": "patron Miglierini aprobado en 3612: Departamentos Av/Roca 1400 frente al Golf",
    }
}


def safe(value):
    if value is None:
        return ""
    return normalize_whitespace(str(value))


def no_zone_query():
    return (
        (Q(inferred_zone__isnull=True) | Q(inferred_zone=""))
        & (
            Q(location_intelligence__isnull=True)
            | Q(location_intelligence__zone_name__isnull=True)
            | Q(location_intelligence__zone_name="")
        )
    )


def first_active_listing(property_obj):
    return (
        property_obj.listings.filter(active=True)
        .select_related("source", "agency")
        .order_by("-last_seen_at")
        .first()
        or property_obj.listings.select_related("source", "agency")
        .order_by("-last_seen_at")
        .first()
    )


def source_names(property_obj):
    names = []
    for listing in property_obj.listings.select_related("source").all():
        if listing.source.name not in names:
            names.append(listing.source.name)
    return ", ".join(names)


def agency_names(property_obj):
    names = []
    for listing in property_obj.listings.select_related("agency").all():
        if listing.agency and listing.agency.name not in names:
            names.append(listing.agency.name)
    return ", ".join(names)


def load_feedback_rows():
    wb = load_workbook(INPUT_XLSX, read_only=True, data_only=False)
    ws = wb["Pendientes"]
    headers = [cell.value for cell in ws[1]]
    rows_by_id = {}
    decisions = {}
    notes = {}
    for values in ws.iter_rows(min_row=2, values_only=True):
        row = dict(zip(headers, values))
        try:
            property_id = int(row.get("id_propiedad") or 0)
        except (TypeError, ValueError):
            continue
        rows_by_id[property_id] = row
        decision = safe(row.get("decision_manual")).upper()
        if decision:
            decisions[property_id] = decision
        note = safe(row.get("notas_manual"))
        if note:
            notes[property_id] = note
    return rows_by_id, decisions, notes


def location_or_none(property_obj):
    return getattr(property_obj, "location", None)


def location_intelligence_or_none(property_obj):
    return getattr(property_obj, "location_intelligence", None)


def is_miami_default(location):
    if not location:
        return False
    return (
        abs(float(location.latitude) - MIAMI_DEFAULT[0]) < 0.000001
        and abs(float(location.longitude) - MIAMI_DEFAULT[1]) < 0.000001
    )


def residual_reason(property_obj):
    reasons = []
    intel = location_intelligence_or_none(property_obj)
    if not property_obj.inferred_zone and not (intel.zone_name if intel else ""):
        reasons.append("sin zona")
    if not best_address(property_obj):
        reasons.append("sin direccion util")
    return ", ".join(reasons)


def current_counts():
    props = list(
        Property.objects.select_related("location", "location_intelligence")
        .prefetch_related("listings__source")
        .all()
    )
    miglierini = [p for p in props if any(l.source.slug == "miglierini" for l in p.listings.all())]
    return {
        "sin_zona_operativa": Property.objects.filter(no_zone_query()).count(),
        "sin_direccion_util": sum(1 for p in props if not best_address(p)),
        "miglierini_sin_zona_operativa": sum(
            1
            for p in miglierini
            if not p.inferred_zone
            and not (location_intelligence_or_none(p).zone_name if location_intelligence_or_none(p) else "")
        ),
        "miglierini_sin_direccion_util": sum(1 for p in miglierini if not best_address(p)),
        "miglierini_pendientes": sum(1 for p in miglierini if residual_reason(p)),
    }


def set_manual_field(property_obj, field, value, now, changes, manual=True, confirm=False):
    value = safe(value)
    if not value:
        return False
    changed = getattr(property_obj, field) != value
    if changed:
        setattr(property_obj, field, value)
        changes.add(field)
    if field == "address":
        normalized = normalize_address(value)
        if property_obj.normalized_address != normalized:
            property_obj.normalized_address = normalized
            changes.add("normalized_address")
    if manual and (changed or confirm):
        overrides = dict(property_obj.manual_overrides or {})
        overrides[field] = now.isoformat()
        property_obj.manual_overrides = overrides
        property_obj.data_manually_corrected_at = now
        changes.update({"manual_overrides", "data_manually_corrected_at"})
    return changed


def save_property(property_obj, changes, apply):
    if changes and apply:
        property_obj.save(update_fields=sorted(changes))


def clear_nonmanual_location(property_obj, reason, summary, apply):
    if any(item["id"] == property_obj.pk for item in summary["locations_cleared"]):
        return False
    location = location_or_none(property_obj)
    if not location:
        return False
    if location.manually_corrected:
        summary["manual_location_preserved_ids"].append(property_obj.pk)
        return False
    summary["locations_cleared"].append(
        {
            "id": property_obj.pk,
            "provider": location.provider,
            "latitude": location.latitude,
            "longitude": location.longitude,
            "reason": reason,
        }
    )
    if apply:
        location.delete()
    return True


def source_zone_for_inference(property_obj):
    for value in (property_obj.neighborhood, property_obj.detected_neighborhood):
        normalized = normalize_neighborhood_name(value)
        if normalized:
            return normalized
    return ""


def source_locality_for_inference(property_obj):
    for value in (property_obj.detected_locality, property_obj.locality):
        normalized = normalize_locality(value)
        if normalized:
            return normalized
    return ""


def apply_zone_values(property_obj, values, evidence_key, apply):
    if not apply:
        return
    now = timezone.now()
    property_obj.inferred_partido = values.get("partido") or property_obj.inferred_partido
    property_obj.inferred_locality = values.get("locality") or property_obj.inferred_locality
    property_obj.inferred_zone = values.get("zone") or property_obj.inferred_zone
    property_obj.inferred_neighborhood = values.get("zone") or property_obj.inferred_neighborhood
    property_obj.territory_confidence = values.get("confidence") or property_obj.territory_confidence
    property_obj.territory_source_method = values.get("source_method") or property_obj.territory_source_method
    property_obj.territory_needs_review = bool(values.get("needs_review"))
    property_obj.territory_evidence = values.get("evidence") or property_obj.territory_evidence
    property_obj.territory_inferred_at = now
    property_obj.zone_needs_review = bool(values.get("needs_review"))
    property_obj.zone_conflict = False
    property_obj.zone_inference_evidence = {
        **(property_obj.zone_inference_evidence or {}),
        evidence_key: values.get("evidence") or {},
    }
    property_obj.zone_inferred_at = now
    property_obj.save(
        update_fields=[
            "inferred_partido",
            "inferred_locality",
            "inferred_zone",
            "inferred_neighborhood",
            "territory_confidence",
            "territory_source_method",
            "territory_needs_review",
            "territory_evidence",
            "territory_inferred_at",
            "zone_needs_review",
            "zone_conflict",
            "zone_inference_evidence",
            "zone_inferred_at",
        ]
    )
    record, _ = PropertyLocationIntelligence.objects.get_or_create(property=property_obj)
    record.partido_name = values.get("partido") or record.partido_name
    record.locality_name = values.get("locality") or record.locality_name
    record.zone_name = values.get("zone") or record.zone_name
    record.match_method = PropertyLocationIntelligence.MatchMethod.COORDINATES
    record.confidence = values.get("confidence") or record.confidence
    record.evidence = {**(record.evidence or {}), evidence_key: values.get("evidence") or {}}
    record.scored_at = now
    record.save(
        update_fields=[
            "partido_name",
            "locality_name",
            "zone_name",
            "match_method",
            "confidence",
            "evidence",
            "scored_at",
        ]
    )


def infer_and_apply_territory(property_obj, tag, apply, allow_reviewed_polygon=False):
    location = location_or_none(property_obj)
    if not location:
        return {"applied": False, "zone": "", "reason": "sin_coordenadas"}
    result = infer_territory_for_point(
        location.latitude,
        location.longitude,
        coordinate_source=location.provider,
        extra_evidence={
            "provider": location.provider,
            "precision": location.precision,
            "curacion": tag,
        },
        source_zone=source_zone_for_inference(property_obj),
        source_locality=source_locality_for_inference(property_obj),
    )
    values = territory_values_from_result(result)
    if not values["zone"]:
        return {"applied": False, "zone": "", "reason": "sin_match_poligono", "values": values}
    if values["needs_review"] and not allow_reviewed_polygon:
        return {
            "applied": False,
            "zone": values["zone"],
            "reason": "conflicto_o_revision",
            "values": values,
        }
    apply_zone_values(property_obj, values, SCRIPT_TAG, apply)
    return {"applied": True, "zone": values["zone"], "reason": "applied" if apply else "would_apply", "values": values}


def apply_manual_zone(property_obj, zone, reason, apply):
    zone = canonicalize_unified_zone_name(normalize_neighborhood_name(zone))
    if zone == "Barrio Cartero":
        zone = "Cartero"
    if not zone:
        return {"applied": False, "zone": "", "reason": "zona_manual_vacia"}
    evidence = {
        "curacion": SCRIPT_TAG,
        "reason": reason,
        "manual_feedback": True,
        "generic_address": property_obj.address,
    }
    values = {
        "partido": "Hurlingham",
        "locality": "Hurlingham",
        "zone": zone,
        "confidence": "manual",
        "source_method": "manual_feedback",
        "needs_review": True,
        "evidence": evidence,
    }
    if apply:
        now = timezone.now()
        overrides = dict(property_obj.manual_overrides or {})
        overrides["neighborhood"] = now.isoformat()
        overrides["inferred_zone"] = now.isoformat()
        property_obj.manual_overrides = overrides
        property_obj.neighborhood = zone
        property_obj.inferred_partido = "Hurlingham"
        property_obj.inferred_locality = "Hurlingham"
        property_obj.inferred_zone = zone
        property_obj.inferred_neighborhood = zone
        property_obj.territory_confidence = "manual"
        property_obj.territory_source_method = "manual_feedback"
        property_obj.territory_needs_review = True
        property_obj.territory_evidence = evidence
        property_obj.territory_inferred_at = now
        property_obj.zone_needs_review = True
        property_obj.zone_conflict = False
        property_obj.zone_inference_evidence = {
            **(property_obj.zone_inference_evidence or {}),
            SCRIPT_TAG: evidence,
        }
        property_obj.zone_inferred_at = now
        property_obj.data_manually_corrected_at = now
        property_obj.save(
            update_fields=[
                "manual_overrides",
                "neighborhood",
                "inferred_partido",
                "inferred_locality",
                "inferred_zone",
                "inferred_neighborhood",
                "territory_confidence",
                "territory_source_method",
                "territory_needs_review",
                "territory_evidence",
                "territory_inferred_at",
                "zone_needs_review",
                "zone_conflict",
                "zone_inference_evidence",
                "zone_inferred_at",
                "data_manually_corrected_at",
            ]
        )
        record, _ = PropertyLocationIntelligence.objects.get_or_create(property=property_obj)
        record.partido_name = "Hurlingham"
        record.locality_name = "Hurlingham"
        record.zone_name = zone
        record.match_method = PropertyLocationIntelligence.MatchMethod.ZONE
        record.confidence = "manual"
        record.evidence = {**(record.evidence or {}), SCRIPT_TAG: evidence}
        record.scored_at = now
        record.save(
            update_fields=[
                "partido_name",
                "locality_name",
                "zone_name",
                "match_method",
                "confidence",
                "evidence",
                "scored_at",
            ]
        )
    return {"applied": True, "zone": zone, "reason": "applied" if apply else "would_apply", "values": values}


def apply_geocode_cache_or_api(property_obj, apply, summary, tag):
    geocoder = Geocoder()
    current = location_or_none(property_obj)
    if current and current.manually_corrected:
        summary["manual_location_preserved_ids"].append(property_obj.pk)
        return current, "manual"
    queries = geocoder.query_candidates(property_obj)
    if not queries:
        summary["no_geocode_query_ids"].append(property_obj.pk)
        return None, "sin_consulta"

    cache = (
        GeocodeCache.objects.filter(query__in=queries)
        .exclude(latitude__isnull=True)
        .exclude(longitude__isnull=True)
        .first()
    )
    if cache:
        if apply:
            location = geocoder._apply(property_obj, cache.query, cache)
        else:
            location = PropertyLocation(
                property=property_obj,
                latitude=cache.latitude,
                longitude=cache.longitude,
                precision=PropertyLocation.Precision.EXACT,
                query=cache.query,
                provider="nominatim",
                confidence=cache.confidence,
                outside_target=False,
                manually_corrected=False,
            )
        summary["geocode_cache_ids"].append(property_obj.pk)
        return location, "cache"

    summary["geocode_api_required_ids"].append(property_obj.pk)
    if not apply:
        return None, "api_required"

    for query in queries:
        if GeocodeCache.objects.filter(query=query).exists():
            continue
        try:
            cache = geocoder._fetch(query, property_obj)
            location = geocoder._apply(property_obj, query, cache)
        except Exception as exc:
            summary["geocode_errors"].append(
                {"id": property_obj.pk, "query": query, "error": repr(exc), "tag": tag}
            )
            continue
        if location:
            summary["geocode_api_ids"].append(property_obj.pk)
            return location, "api"
    summary["no_geocode_result_ids"].append(property_obj.pk)
    return None, "sin_resultado_api"


def apply_feedback(apply):
    rows_by_id, decisions, notes = load_feedback_rows()
    now = timezone.now()
    summary = {
        "excel_approved_ids": sorted(pid for pid, decision in decisions.items() if decision == "APROBADO"),
        "approved_by_note_ids": sorted(APPROVED_BY_NOTE_IDS),
        "approved_applied_ids": [],
        "propagated_ids": sorted(PROPAGATED_ADDRESSES),
        "address_updates": [],
        "locality_updates": [],
        "neighborhood_updates": [],
        "manual_zone_ids": [],
        "manual_zone_conflicts": [],
        "locations_cleared": [],
        "manual_location_preserved_ids": [],
        "geocode_cache_ids": [],
        "geocode_api_required_ids": [],
        "geocode_api_ids": [],
        "geocode_errors": [],
        "no_geocode_query_ids": [],
        "no_geocode_result_ids": [],
        "zone_applied_ids": [],
        "zone_conflict_ids": [],
        "zone_pending_ids": [],
        "miami_clear_ids": [],
        "missing_feedback_rows": [],
        "missing_property_ids": [],
        "notes_ids": sorted(notes),
    }

    for property_id in sorted(NON_CARTERO_APPROVED_IDS):
        row = rows_by_id.get(property_id)
        if not row:
            summary["missing_feedback_rows"].append(property_id)
            continue
        try:
            property_obj = (
                Property.objects.select_related("location", "location_intelligence")
                .prefetch_related("listings__source")
                .get(pk=property_id)
            )
        except Property.DoesNotExist:
            summary["missing_property_ids"].append(property_id)
            continue
        changes = set()
        before = {
            "address": property_obj.address,
            "locality": property_obj.locality,
            "neighborhood": property_obj.neighborhood,
        }
        address_changed = set_manual_field(
            property_obj, "address", row.get("domicilio_actual"), now, changes, confirm=True
        )
        set_manual_field(property_obj, "locality", row.get("localidad_actual"), now, changes, confirm=True)
        if safe(row.get("barrio_actual")):
            set_manual_field(property_obj, "neighborhood", row.get("barrio_actual"), now, changes, confirm=True)
        if address_changed:
            clear_nonmanual_location(property_obj, "direccion_manual_aprobada", summary, apply)
        save_property(property_obj, changes, apply)
        summary["approved_applied_ids"].append(property_id)
        if "address" in changes:
            summary["address_updates"].append({"id": property_id, "from": before["address"], "to": property_obj.address})
        if "locality" in changes:
            summary["locality_updates"].append({"id": property_id, "from": before["locality"], "to": property_obj.locality})
        if "neighborhood" in changes:
            summary["neighborhood_updates"].append({"id": property_id, "from": before["neighborhood"], "to": property_obj.neighborhood})

        if apply:
            property_obj = Property.objects.select_related("location").get(pk=property_id)
        location, _status = apply_geocode_cache_or_api(property_obj, apply, summary, "approved_address")
        if not apply and location:
            property_obj.location = location
        elif apply:
            property_obj = Property.objects.select_related("location").get(pk=property_id)
        zone_result = infer_and_apply_territory(property_obj, "approved_feedback", apply)
        if zone_result["applied"]:
            summary["zone_applied_ids"].append({"id": property_id, "zone": zone_result["zone"]})
        elif zone_result["reason"] == "conflicto_o_revision":
            summary["zone_conflict_ids"].append({"id": property_id, "zone": zone_result["zone"]})
        else:
            manual_zone = safe(row.get("zona_actual")) or safe(row.get("barrio_actual"))
            if manual_zone:
                manual_result = apply_manual_zone(
                    property_obj,
                    manual_zone,
                    "zona/barrio indicado en feedback aprobado sin poligono claro",
                    apply,
                )
                summary["manual_zone_ids"].append({"id": property_id, "zone": manual_result["zone"]})
            else:
                summary["zone_pending_ids"].append({"id": property_id, "reason": zone_result["reason"]})

    for property_id in sorted(PROPAGATED_ADDRESSES):
        payload = PROPAGATED_ADDRESSES[property_id]
        try:
            property_obj = Property.objects.select_related("location").get(pk=property_id)
        except Property.DoesNotExist:
            summary["missing_property_ids"].append(property_id)
            continue
        changes = set()
        before_address = property_obj.address
        address_changed = set_manual_field(property_obj, "address", payload["address"], now, changes, confirm=True)
        set_manual_field(property_obj, "locality", payload["locality"], now, changes, confirm=True)
        if address_changed:
            clear_nonmanual_location(property_obj, "direccion_propagada_segura", summary, apply)
        save_property(property_obj, changes, apply)
        summary["address_updates"].append({"id": property_id, "from": before_address, "to": property_obj.address, "propagated": True})
        if apply:
            property_obj = Property.objects.select_related("location").get(pk=property_id)
        location, _status = apply_geocode_cache_or_api(property_obj, apply, summary, "propagated_address")
        if not apply and location:
            property_obj.location = location
        elif apply:
            property_obj = Property.objects.select_related("location").get(pk=property_id)
        zone_result = infer_and_apply_territory(property_obj, "propagated_feedback", apply)
        if zone_result["applied"]:
            summary["zone_applied_ids"].append({"id": property_id, "zone": zone_result["zone"]})
        else:
            summary["zone_pending_ids"].append({"id": property_id, "reason": zone_result["reason"]})

    for property_id in sorted(CLEAR_MIAMI_IDS):
        try:
            property_obj = Property.objects.select_related("location").get(pk=property_id)
        except Property.DoesNotExist:
            summary["missing_property_ids"].append(property_id)
            continue
        location = location_or_none(property_obj)
        if (
            location
            and is_miami_default(location)
            and location.provider == "miglierini_map"
            and not location.manually_corrected
        ):
            summary["miami_clear_ids"].append(property_id)
            clear_nonmanual_location(property_obj, "coordenada_default_miami_invalida", summary, apply)

    for property_id in sorted(CARTERO_IDS):
        try:
            property_obj = Property.objects.select_related("location").get(pk=property_id)
        except Property.DoesNotExist:
            summary["missing_property_ids"].append(property_id)
            continue
        changes = set()
        before = {
            "address": property_obj.address,
            "locality": property_obj.locality,
            "neighborhood": property_obj.neighborhood,
        }
        set_manual_field(property_obj, "address", "Barrio Cartero", now, changes, confirm=True)
        set_manual_field(property_obj, "locality", "Hurlingham", now, changes, confirm=True)
        clear_nonmanual_location(property_obj, "cartero_generico_no_geocodificar", summary, apply)
        save_property(property_obj, changes, apply)
        if "address" in changes:
            summary["address_updates"].append({"id": property_id, "from": before["address"], "to": "Barrio Cartero"})
        if "locality" in changes:
            summary["locality_updates"].append({"id": property_id, "from": before["locality"], "to": "Hurlingham"})
        manual_result = apply_manual_zone(
            property_obj,
            "Cartero",
            "correccion manual/generica por feedback; no geocodificar Barrio Cartero como direccion exacta",
            apply,
        )
        summary["manual_zone_ids"].append({"id": property_id, "zone": manual_result["zone"]})

    return summary


def excel_number(value):
    if value is None:
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def generate_residual_excel(path):
    props = list(
        Property.objects.select_related("location", "location_intelligence")
        .prefetch_related("listings__source", "listings__agency")
        .all()
    )
    pending = [property_obj for property_obj in props if residual_reason(property_obj)]

    wb = Workbook()
    ws = wb.active
    ws.title = "Pendientes"
    headers = [
        "id_propiedad",
        "titulo",
        "descripcion_resumen",
        "domicilio_actual",
        "direccion_detectada",
        "direccion_normalizada",
        "localidad_actual",
        "localidad_detectada",
        "barrio_actual",
        "barrio_detectado",
        "zona_actual",
        "zona_inteligencia",
        "latitud",
        "longitud",
        "fuente_coordenadas",
        "fuente",
        "inmobiliaria",
        "url_publicacion",
        "motivo_url",
        "estado_publicacion",
        "tipo_propiedad",
        "operacion",
        "moneda",
        "precio",
        "ambientes",
        "dormitorios",
        "banos",
        "superficie_cubierta",
        "superficie_total",
        "fecha_ultima_vista",
        "motivo_pendiente",
        "evidencia",
        "decision_manual",
        "notas_manual",
    ]
    ws.append(headers)
    for property_obj in sorted(pending, key=lambda item: (residual_reason(item), item.pk)):
        listing = first_active_listing(property_obj)
        location = location_or_none(property_obj)
        intel = location_intelligence_or_none(property_obj)
        description = normalize_whitespace(property_obj.description or "")
        evidence = {
            "location_precision": location.precision if location else "",
            "location_query": location.query if location else "",
            "manual_overrides": property_obj.manual_overrides or {},
            "source_locality": property_obj.locality or property_obj.detected_locality or "",
            "source_zone": property_obj.neighborhood or property_obj.detected_neighborhood or "",
            "territory_evidence": property_obj.territory_evidence or {},
        }
        ws.append(
            [
                property_obj.pk,
                property_obj.title,
                description[:500],
                property_obj.address,
                property_obj.detected_address,
                property_obj.normalized_address,
                property_obj.locality,
                property_obj.detected_locality,
                property_obj.neighborhood,
                property_obj.detected_neighborhood,
                property_obj.inferred_zone or property_obj.inferred_neighborhood,
                intel.zone_name if intel else "",
                location.latitude if location else None,
                location.longitude if location else None,
                location.provider if location else "",
                source_names(property_obj),
                agency_names(property_obj),
                listing.url if listing else "",
                "" if listing else "Sin Listing asociado en la base",
                property_obj.get_status_display(),
                property_obj.get_property_type_display(),
                property_obj.operation,
                property_obj.currency,
                excel_number(property_obj.price),
                property_obj.rooms,
                property_obj.bedrooms,
                excel_number(property_obj.bathrooms),
                excel_number(property_obj.covered_area),
                excel_number(property_obj.total_area),
                property_obj.last_seen_at.isoformat() if property_obj.last_seen_at else "",
                residual_reason(property_obj),
                json.dumps(evidence, ensure_ascii=False, default=str),
                "",
                "",
            ]
        )

    if ws.max_row >= 1:
        ws.auto_filter.ref = ws.dimensions
    if ws.max_row > 1:
        table = Table(displayName="PendientesTable", ref=ws.dimensions)
        style = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        table.tableStyleInfo = style
        ws.add_table(table)
    ws.freeze_panes = "A2"

    header_fill = PatternFill("solid", fgColor="1F4E78")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    price_col = headers.index("precio") + 1
    decision_col = headers.index("decision_manual") + 1
    for row in range(2, ws.max_row + 1):
        ws.cell(row=row, column=price_col).number_format = "#,##0.00"
    if ws.max_row > 1:
        decision_letter = get_column_letter(decision_col)
        validation = DataValidation(
            type="list",
            formula1='"APROBADO,RECHAZADO"',
            allow_blank=True,
            showErrorMessage=True,
            errorTitle="Valor no valido",
            error="Usa APROBADO o RECHAZADO.",
        )
        ws.add_data_validation(validation)
        validation.add(f"{decision_letter}2:{decision_letter}{ws.max_row}")
    for column in range(1, ws.max_column + 1):
        letter = get_column_letter(column)
        max_len = max(
            len(str(ws.cell(row=row, column=column).value or ""))
            for row in range(1, min(ws.max_row, 80) + 1)
        )
        ws.column_dimensions[letter].width = min(max(max_len + 2, 12), 52)

    summary = wb.create_sheet("Resumen")
    counts = current_counts()
    for row in [
        ("fecha", timezone.now().isoformat()),
        ("pendientes_total", len(pending)),
        ("sin_zona_operativa", counts["sin_zona_operativa"]),
        ("sin_direccion_util", counts["sin_direccion_util"]),
        ("miglierini_pendientes", counts["miglierini_pendientes"]),
        ("archivo_origen", str(INPUT_XLSX)),
    ]:
        summary.append(row)
    for cell in summary[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
    summary.freeze_panes = "A2"
    for column in range(1, summary.max_column + 1):
        summary.column_dimensions[get_column_letter(column)].width = 28

    wb.save(path)
    return {"path": str(path), "pending_rows": len(pending), "columns": headers}


def validate_excel(path):
    wb = load_workbook(path, read_only=False, data_only=False)
    ws = wb["Pendientes"]
    headers = [cell.value for cell in ws[1]]
    price_col = headers.index("precio") + 1
    decision_col = headers.index("decision_manual") + 1
    notes_col = headers.index("notas_manual") + 1
    price_types = {}
    nonblank_decisions = 0
    nonblank_notes = 0
    for row in range(2, ws.max_row + 1):
        price = ws.cell(row=row, column=price_col).value
        price_types[type(price).__name__] = price_types.get(type(price).__name__, 0) + 1
        if safe(ws.cell(row=row, column=decision_col).value):
            nonblank_decisions += 1
        if safe(ws.cell(row=row, column=notes_col).value):
            nonblank_notes += 1
    return {
        "sheets": wb.sheetnames,
        "rows": ws.max_row - 1,
        "auto_filter": ws.auto_filter.ref,
        "tables": list(ws.tables),
        "freeze_panes": ws.freeze_panes,
        "price_types": price_types,
        "nonblank_decisions": nonblank_decisions,
        "nonblank_notes": nonblank_notes,
        "missing_columns": sorted(
            {"motivo_url", "decision_manual", "notas_manual", "precio"} - set(headers)
        ),
    }


def sample_state(ids):
    rows = []
    for property_obj in (
        Property.objects.filter(pk__in=ids)
        .select_related("location", "location_intelligence")
        .prefetch_related("listings__source")
        .order_by("pk")
    ):
        location = location_or_none(property_obj)
        intel = location_intelligence_or_none(property_obj)
        rows.append(
            {
                "id": property_obj.pk,
                "address": property_obj.address,
                "locality": property_obj.locality,
                "neighborhood": property_obj.neighborhood,
                "inferred_zone": property_obj.inferred_zone,
                "intel_zone": intel.zone_name if intel else "",
                "location": None
                if not location
                else {
                    "lat": location.latitude,
                    "lon": location.longitude,
                    "provider": location.provider,
                    "manual": location.manually_corrected,
                },
                "manual_overrides": property_obj.manual_overrides or {},
                "data_manually_corrected_at": property_obj.data_manually_corrected_at,
            }
        )
    return rows


def write_json(path, payload):
    Path(path).write_text(
        json.dumps(payload, ensure_ascii=False, indent=2, default=str),
        encoding="utf-8",
    )


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--apply", action="store_true")
    parser.add_argument("--generate-excel", action="store_true")
    parser.add_argument("--json-out")
    args = parser.parse_args()

    before = current_counts()
    if args.apply and not BACKUP_PATH.exists():
        shutil.copy2(DB_PATH, BACKUP_PATH)

    with transaction.atomic():
        feedback = apply_feedback(args.apply)
        if not args.apply:
            transaction.set_rollback(True)

    excel = None
    excel_validation = None
    if args.generate_excel:
        excel = generate_residual_excel(OUTPUT_XLSX)
        excel_validation = validate_excel(OUTPUT_XLSX)

    after = current_counts()
    result = {
        "mode": "apply" if args.apply else "dry_run",
        "input_excel": str(INPUT_XLSX),
        "output_excel": str(OUTPUT_XLSX) if args.generate_excel else "",
        "backup": str(BACKUP_PATH) if args.apply else "",
        "before": before,
        "after": after,
        "feedback": feedback,
        "sample": sample_state(
            [
                3301,
                3311,
                3325,
                3330,
                3356,
                3368,
                3487,
                3606,
                3612,
                3674,
                3788,
                3823,
                3829,
            ]
        ),
        "excel": excel,
        "excel_validation": excel_validation,
    }
    if args.json_out:
        write_json(args.json_out, result)
    print(json.dumps(result, ensure_ascii=False, indent=2, default=str))


if __name__ == "__main__":
    main()
