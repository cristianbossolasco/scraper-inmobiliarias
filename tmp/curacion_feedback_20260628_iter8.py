import argparse
import importlib.util
import json
import os
import shutil
import sys
from collections import Counter, defaultdict
from pathlib import Path

from django.conf import settings
from django.db.models import Q
from django.utils import timezone
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))
os.environ.setdefault("DJANGO_SETTINGS_MODULE", "config.settings")

import django

django.setup()

from properties.models import GeocodeCache, Listing, Property, PropertyLocation, PropertyLocationIntelligence
from properties.services.geocoding import Geocoder, address_number, best_address, street_key
from properties.services.location_intelligence import (
    apply_location_intelligence_score,
    load_location_zones,
    score_property_location_intelligence,
)
from properties.services.normalization import (
    clean_address_for_storage,
    normalize_address,
    normalize_locality,
    normalize_whitespace,
)
from properties.services.territory_hierarchy import apply_territory_inference, infer_property_territory

INPUT_XLSX = ROOT / "tmp" / "curacion_zonas_pendientes_2026-06-28_feedback_iter7.xlsx"
OUTPUT_XLSX = ROOT / "tmp" / "curacion_zonas_pendientes_2026-06-28_feedback_iter8.xlsx"
BACKUP_PATH = ROOT / "tmp" / "db.sqlite3.backup_curacion_feedback_20260628_iter8"
DB_PATH = ROOT / "db.sqlite3"
SCRIPT_TAG = "curacion_feedback_20260628_iter8"

APPROVED_IDS = set()
REJECTED_REMOVE_IDS = {5867}
NO_ACTION_IDS = set()
MAP_ONLY_IDS = set()

REQUESTED_ZONE_IDS = {
    6123, 6135, 6136, 6137, 6138, 6139, 6140, 6141, 6142, 6143, 6144, 6145, 6146,
    6147, 6148, 6149, 6150, 6151, 6152, 6153, 6154, 6155, 6157, 6159, 6161, 6162,
    6164, 6165, 6166, 6167, 6168, 6169, 6170, 6171, 6172, 6173, 6175, 6176, 6177,
    6178, 6179, 6180, 6181, 6182, 6183, 6184, 6185, 6186, 6187, 6188, 6190, 6191,
    6192, 6193, 6194, 6195, 6196, 6198, 6199, 6200, 6201, 6204, 6205, 6206, 6207,
    6208, 6209, 6210, 6211, 6212, 6213, 6214, 6215, 6216, 6217, 6218, 6219,
}

CLEAR_ZONE_IDS = {
    6142, 6143, 6147, 6150, 6152, 6153, 6157, 6159, 6164, 6165, 6167, 6168, 6170,
    6173, 6175, 6184, 6187, 6188, 6191, 6196, 6198, 6199, 6201, 6204, 6205, 6211,
    6212, 6216,
}
EXPLICIT_REVIEW_ZONE_IDS = {
    6136, 6137, 6138, 6139, 6140, 6141, 6146, 6148, 6149, 6151, 6154, 6155, 6161,
    6162, 6166, 6169, 6172, 6176, 6177, 6178, 6179, 6180, 6181, 6182, 6183, 6185,
    6186, 6190, 6192, 6193, 6194, 6195, 6200, 6207, 6208, 6210, 6214, 6215, 6217,
    6219,
}
ALREADY_HAS_ZONE_IDS = {6171, 6206, 6209}

SUSPICIOUS_COORDINATE_FIXES = {
    6123: {"expected_zone": "Curapaytí", "reason": "coordenada Mapaprop previa fuera de zona; usar cache de Gurruchaga 2210"},
    6145: {"expected_zone": "Curapaytí", "reason": "coordenada Mapaprop previa fuera de zona; usar cache de J.Bustamante/J.Eva Perón 2200"},
}
ADDRESS_FIXES = {
    6135: {"address": "Einstein 200", "locality": "Villa Tesei", "external_geocode": True, "source": "titulo/fuente"},
    6144: {"address": "Conscripto Bernardi 1985", "locality": "Hurlingham", "external_geocode": False, "source": "titulo/fuente"},
    6218: {"address": "Conscripto Bernardi 1900", "locality": "Hurlingham", "external_geocode": False, "source": "correccion typo"},
}
VARIANT_GEOCODE_FIXES = {
    6213: {
        "variants": ["Sgto. Jose Mariano Gomez 1500", "Sargento Jose Mariano Gomez 1500"],
        "locality": "Hurlingham",
        "external_geocode": True,
    }
}

SAMPLE_IDS = sorted(
    REJECTED_REMOVE_IDS
    | {6123, 6135, 6144, 6145, 6213, 6218}
    | set(list(CLEAR_ZONE_IDS)[:6])
    | set(list(EXPLICIT_REVIEW_ZONE_IDS)[:8])
    | ALREADY_HAS_ZONE_IDS
)


def safe(value):
    return normalize_whitespace(str(value)) if value not in (None, "") else ""


def same_geocoding_target(before, after):
    return bool(before and after) and address_number(before) == address_number(after) and street_key(before) == street_key(after)


def add(summary, key, value):
    summary.setdefault(key, []).append(value)


def read_feedback():
    wb = load_workbook(INPUT_XLSX, data_only=True)
    ws = wb["Pendientes"]
    headers = [cell.value for cell in ws[1]]
    index = {header: pos + 1 for pos, header in enumerate(headers)}
    decisions = {}
    notes = {}
    counts = Counter()
    for row_number in range(2, ws.max_row + 1):
        property_id = ws.cell(row_number, index["id_propiedad"]).value
        raw_decision = ws.cell(row_number, index["decision_manual"]).value or ""
        decision = safe(raw_decision).upper()
        note = safe(ws.cell(row_number, index["notas_manual"]).value)
        counts[decision or "VACIO"] += 1
        if decision:
            decisions[int(property_id)] = decision
        if note:
            notes[int(property_id)] = note
    return {
        "rows": ws.max_row - 1,
        "decision_counts": dict(counts),
        "decisions": decisions,
        "notes": notes,
    }


def current_counts():
    queryset = Property.objects.select_related("location", "location_intelligence").filter(
        is_hidden=False
    ).exclude(status=Property.Status.REMOVED)
    no_zone = queryset.filter(
        (Q(inferred_zone__isnull=True) | Q(inferred_zone=""))
        & (
            Q(location_intelligence__isnull=True)
            | Q(location_intelligence__zone_name__isnull=True)
            | Q(location_intelligence__zone_name="")
        )
    ).count()
    no_address_queryset = Property.objects.filter(is_hidden=False).exclude(status=Property.Status.REMOVED).only(
        "address",
        "detected_address",
    )
    no_address = sum(1 for prop in no_address_queryset if not best_address(prop))
    return {"active_visible": queryset.count(), "sin_zona_operativa": no_zone, "sin_direccion_util": no_address}


def property_snapshot(property_id):
    try:
        prop = Property.objects.select_related("location", "location_intelligence").prefetch_related(
            "listings__source", "listings__agency"
        ).get(pk=property_id)
    except Property.DoesNotExist:
        return {"id": property_id, "missing": True}
    location = getattr(prop, "location", None)
    intel = getattr(prop, "location_intelligence", None)
    return {
        "id": prop.pk,
        "status": prop.status,
        "is_hidden": prop.is_hidden,
        "address": prop.address,
        "detected_address": prop.detected_address,
        "locality": prop.locality,
        "inferred_zone": prop.inferred_zone,
        "intelligence_zone": intel.zone_name if intel else "",
        "latitude": location.latitude if location else None,
        "longitude": location.longitude if location else None,
        "provider": location.provider if location else "",
        "manual_location": location.manually_corrected if location else False,
        "sources": sorted({listing.source.slug for listing in prop.listings.all() if listing.source}),
    }


def validation_state(feedback):
    decisions = feedback["decisions"]
    approved_found = sorted(pid for pid, decision in decisions.items() if decision == "APROBADO")
    rejected_found = sorted(pid for pid, decision in decisions.items() if decision == "RECHAZADO")
    return {
        "approved_found": approved_found,
        "rejected_found": rejected_found,
        "unexpected_approved": sorted(set(approved_found) - APPROVED_IDS),
        "unexpected_rejected": sorted(set(rejected_found) - REJECTED_REMOVE_IDS),
        "missing_expected_approved": sorted(APPROVED_IDS - set(approved_found)),
        "missing_expected_rejected": sorted(REJECTED_REMOVE_IDS - set(rejected_found)),
    }


def ensure_can_apply(summary):
    blockers = []
    for key in ("unexpected_approved", "unexpected_rejected", "missing_expected_approved", "missing_expected_rejected"):
        if summary.get(key):
            blockers.append({key: summary[key]})
    missing_requested = sorted(REQUESTED_ZONE_IDS - set(Property.objects.filter(pk__in=REQUESTED_ZONE_IDS).values_list("pk", flat=True)))
    if missing_requested:
        blockers.append({"missing_requested_zone_ids": missing_requested})
    if blockers:
        raise RuntimeError(f"Apply bloqueado por validacion: {blockers}")


def backup_db(summary):
    if BACKUP_PATH.exists():
        summary["backup"] = {"path": str(BACKUP_PATH), "created": False, "reason": "already_exists"}
        return
    shutil.copy2(DB_PATH, BACKUP_PATH)
    summary["backup"] = {"path": str(BACKUP_PATH), "created": True}


def normalize_storage_address(address):
    return clean_address_for_storage(address) or normalize_whitespace(address)


def apply_address(prop, address, locality, summary, apply):
    new_address = normalize_storage_address(address)
    new_locality = normalize_locality(locality or "") or normalize_whitespace(locality or "")
    old_address = prop.address or prop.detected_address or ""
    updates = {}
    if new_address and prop.address != new_address:
        updates["address"] = new_address
    if new_locality and prop.locality != new_locality:
        updates["locality"] = new_locality
    if not updates:
        return
    add(summary, "address_updates", {"id": prop.pk, **updates, "old_address": prop.address, "old_locality": prop.locality})
    if not apply:
        return
    now = timezone.now()
    overrides = dict(prop.manual_overrides or {})
    for field, value in updates.items():
        setattr(prop, field, value)
        overrides[field] = now.isoformat()
    prop.normalized_address = normalize_address(prop.address)
    prop.manual_overrides = overrides
    prop.data_manually_corrected_at = now
    prop.save(update_fields=sorted(set(updates) | {"normalized_address", "manual_overrides", "data_manually_corrected_at"}))
    location = getattr(prop, "location", None)
    new_target = prop.address or prop.detected_address or ""
    if location and not location.manually_corrected and "address" in updates and not same_geocoding_target(old_address, new_target):
        add(summary, "location_invalidated", {"id": prop.pk, "provider": location.provider, "old_lat": location.latitude, "old_lon": location.longitude})
        location.delete()


def cache_result(prop):
    geocoder = Geocoder()
    for query in geocoder.query_candidates(prop):
        cache = GeocodeCache.objects.filter(query=query).first()
        if cache and cache.latitude is not None and cache.longitude is not None:
            return query, cache
    return "", None


def geocode_property(prop, summary, apply, allow_external=False, force=True):
    location = getattr(prop, "location", None)
    if location and location.manually_corrected:
        add(summary, "geocode_skipped_manual_location", {"id": prop.pk})
        return location
    cache_query, cache = cache_result(prop)
    if cache_query:
        add(summary, "geocode_location", {"id": prop.pk, "method": "cache", "query": cache_query})
        if not apply:
            bounds = settings.HURLINGHAM_BOUNDS
            location = PropertyLocation(
                property=prop,
                latitude=cache.latitude,
                longitude=cache.longitude,
                precision=cache.precision or "exact",
                query=cache_query,
                provider="nominatim",
                confidence=cache.confidence,
                outside_target=not (
                    bounds["south"] <= cache.latitude <= bounds["north"]
                    and bounds["west"] <= cache.longitude <= bounds["east"]
                ),
                manually_corrected=False,
            )
            prop.location = location
            return None
        return Geocoder().geocode_property_from_cache(prop, force=force)
    add(summary, "geocode_cache_local_miss", {"id": prop.pk, "address": best_address(prop)})
    if not allow_external:
        return None
    add(summary, "geocode_location", {"id": prop.pk, "method": "api_allowed", "address": best_address(prop)})
    if not apply:
        return None
    return Geocoder().geocode_property(prop, force=force)


def set_curation_evidence(prop, result, reason, apply):
    evidence = dict(prop.zone_inference_evidence or {})
    evidence[SCRIPT_TAG] = {
        "reason": reason,
        "zone": result.zone,
        "needs_review": result.needs_review,
        "timestamp": timezone.now().isoformat(),
    }
    territory_evidence = dict(prop.territory_evidence or {})
    territory_evidence[SCRIPT_TAG] = evidence[SCRIPT_TAG]
    if not apply:
        return
    prop.zone_inference_evidence = evidence
    prop.territory_evidence = territory_evidence
    prop.save(update_fields=["zone_inference_evidence", "territory_evidence"])


def update_intelligence(prop, summary, apply):
    if not apply:
        return
    dataset = load_location_zones()
    if dataset["configured"]:
        score = score_property_location_intelligence(
            prop,
            zones=dataset["features"],
            source_signature=dataset["signature"],
        )
        record = apply_location_intelligence_score(prop, score)
        add(summary, "location_intelligence_updated", {"id": prop.pk, "zone_name": record.zone_name, "method": record.match_method})
        return
    record, _ = PropertyLocationIntelligence.objects.update_or_create(
        property=prop,
        defaults={
            "zone_name": prop.inferred_zone,
            "locality_name": prop.inferred_locality,
            "partido_name": prop.inferred_partido,
            "match_method": PropertyLocationIntelligence.MatchMethod.COORDINATES if hasattr(prop, "location") else PropertyLocationIntelligence.MatchMethod.ZONE,
            "confidence": prop.territory_confidence,
            "evidence": {"source": SCRIPT_TAG},
            "scored_at": timezone.now(),
        },
    )
    add(summary, "location_intelligence_updated", {"id": prop.pk, "zone_name": record.zone_name, "method": "fallback"})


def infer_and_apply_zone(prop, summary, apply, reason, allow_review=True, expected_zone=""):
    location = getattr(prop, "location", None)
    if not location:
        add(summary, "zone_skipped", {"id": prop.pk, "reason": "sin coordenadas"})
        return None
    result = infer_property_territory(prop)
    row = {
        "id": prop.pk,
        "zone": result.zone,
        "locality": result.locality,
        "needs_review": result.needs_review,
        "reason": reason,
        "lat": location.latitude,
        "lon": location.longitude,
        "provider": location.provider,
    }
    if expected_zone and result.zone != expected_zone:
        row["expected_zone"] = expected_zone
        add(summary, "zone_skipped", {**row, "reason": "zona distinta a esperada"})
        return result
    if not result.zone:
        add(summary, "zone_skipped", {**row, "reason": "poligono sin zona"})
        return result
    if result.needs_review and not allow_review:
        add(summary, "zone_skipped", {**row, "reason": "needs_review"})
        return result
    add(summary, "zone_applied", row)
    if apply:
        apply_territory_inference(prop, result)
        prop.refresh_from_db()
        set_curation_evidence(prop, result, reason, apply=True)
        update_intelligence(prop, summary, apply=True)
    return result


def remove_rejected(summary, apply):
    for property_id in sorted(REJECTED_REMOVE_IDS):
        prop = Property.objects.prefetch_related("listings").get(pk=property_id)
        add(summary, "removed", {"id": prop.pk, "title": prop.title, "reason": "feedback RECHAZADO: fuera de Palomar/Hurlingham"})
        if not apply:
            continue
        prop.status = Property.Status.REMOVED
        prop.is_hidden = True
        prop.zone_inference_evidence = {
            **(prop.zone_inference_evidence or {}),
            SCRIPT_TAG: {"action": "logical_remove", "reason": "feedback: Eliminar, es de Palomar", "timestamp": timezone.now().isoformat()},
        }
        prop.save(update_fields=["status", "is_hidden", "zone_inference_evidence"])
        prop.listings.filter(active=True).update(active=False, source_status="removed")


def process_existing_coordinate_zones(summary, apply):
    for property_id in sorted(CLEAR_ZONE_IDS | EXPLICIT_REVIEW_ZONE_IDS):
        prop = Property.objects.select_related("location", "location_intelligence").get(pk=property_id)
        allow_review = property_id in EXPLICIT_REVIEW_ZONE_IDS
        reason = "explicit_user_requested_zone_inference_iter8" if allow_review else "coordinate_polygon_clear_iter8"
        infer_and_apply_zone(prop, summary, apply, reason=reason, allow_review=allow_review)


def process_already_has_zones(summary, apply):
    for property_id in sorted(ALREADY_HAS_ZONE_IDS):
        prop = Property.objects.select_related("location", "location_intelligence").get(pk=property_id)
        result = infer_and_apply_zone(prop, summary, False, reason="already_had_zone_validation", allow_review=True)
        add(
            summary,
            "already_has_zone_validated",
            {"id": prop.pk, "current_zone": prop.inferred_zone, "dry_zone": result.zone if result else "", "needs_review": result.needs_review if result else None},
        )
        if apply and prop.inferred_zone and not getattr(prop, "location_intelligence", None):
            update_intelligence(prop, summary, apply=True)


def process_suspicious_coordinates(summary, apply):
    for property_id, config in SUSPICIOUS_COORDINATE_FIXES.items():
        prop = Property.objects.select_related("location", "location_intelligence").get(pk=property_id)
        location = getattr(prop, "location", None)
        if location and location.manually_corrected:
            add(summary, "zone_skipped", {"id": prop.pk, "reason": "ubicacion manual preservada"})
            continue
        geocode_property(prop, summary, apply, allow_external=False, force=True)
        if apply:
            prop.refresh_from_db()
        infer_and_apply_zone(
            prop,
            summary,
            apply,
            reason=config["reason"],
            allow_review=False,
            expected_zone=config["expected_zone"],
        )


def process_address_fixes(summary, apply):
    for property_id, config in ADDRESS_FIXES.items():
        prop = Property.objects.select_related("location", "location_intelligence").get(pk=property_id)
        apply_address(prop, config["address"], config["locality"], summary, apply)
        if not apply:
            prop.address = normalize_storage_address(config["address"])
            prop.locality = normalize_locality(config["locality"] or "") or normalize_whitespace(config["locality"] or "")
            prop.normalized_address = normalize_address(prop.address)
        if apply:
            prop.refresh_from_db()
        geocode_property(prop, summary, apply, allow_external=config.get("external_geocode", False), force=True)
        if apply:
            prop.refresh_from_db()
        location = getattr(prop, "location", None)
        if location and location.outside_target:
            add(summary, "zone_skipped", {"id": prop.pk, "reason": "geocoding fuera de Hurlingham", "lat": location.latitude, "lon": location.longitude})
            continue
        infer_and_apply_zone(prop, summary, apply, reason=f"address_fix_{config['source']}_iter8", allow_review=True)


def process_variant_geocode_fixes(summary, apply):
    for property_id, config in VARIANT_GEOCODE_FIXES.items():
        original = Property.objects.select_related("location", "location_intelligence").get(pk=property_id)
        if getattr(original, "location", None) and original.location.manually_corrected:
            add(summary, "geocode_skipped_manual_location", {"id": original.pk})
            continue
        resolved = False
        for variant in config["variants"]:
            prop = Property.objects.select_related("location", "location_intelligence").get(pk=property_id)
            if apply:
                apply_address(prop, variant, config["locality"], summary, apply=True)
                prop.refresh_from_db()
            else:
                prop.address = variant
                prop.locality = config["locality"]
                prop.normalized_address = normalize_address(variant)
            location = geocode_property(prop, summary, apply, allow_external=config.get("external_geocode", False), force=True)
            if not apply:
                continue
            prop.refresh_from_db()
            location = getattr(prop, "location", None)
            if location and not location.outside_target:
                result = infer_and_apply_zone(prop, summary, apply=True, reason=f"variant_geocode_{variant}_iter8", allow_review=True)
                resolved = bool(result and result.zone)
                break
            if location and location.outside_target:
                add(summary, "geocode_failed", {"id": property_id, "variant": variant, "reason": "fuera de Hurlingham", "lat": location.latitude, "lon": location.longitude})
                if apply and not location.manually_corrected:
                    location.delete()
        if not resolved:
            add(summary, "geocode_failed", {"id": property_id, "variants": config["variants"]})


def run(apply=False):
    feedback = read_feedback()
    summary = {
        "script": SCRIPT_TAG,
        "input_excel": str(INPUT_XLSX),
        "output_excel": str(OUTPUT_XLSX) if apply else "",
        "feedback": {**feedback, "decisions": {str(k): v for k, v in feedback["decisions"].items()}},
        "before": current_counts(),
    }
    summary.update(validation_state(feedback))
    ensure_can_apply(summary)
    if apply:
        backup_db(summary)
    remove_rejected(summary, apply)
    process_suspicious_coordinates(summary, apply)
    process_address_fixes(summary, apply)
    process_variant_geocode_fixes(summary, apply)
    process_existing_coordinate_zones(summary, apply)
    process_already_has_zones(summary, apply)
    summary["after"] = current_counts()
    summary["sample"] = {str(property_id): property_snapshot(property_id) for property_id in SAMPLE_IDS}
    return summary


def load_iter6_module():
    path = ROOT / "tmp" / "curacion_feedback_20260628_iter6.py"
    spec = importlib.util.spec_from_file_location("curacion_iter6_for_iter8", path)
    module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(module)
    return module


def generate_excel():
    module = load_iter6_module()
    module.OUTPUT_XLSX = OUTPUT_XLSX
    module.SCRIPT_TAG = SCRIPT_TAG
    result = module.generate_excel({})
    wb = load_workbook(OUTPUT_XLSX)
    ws = wb["Pendientes"]
    for name in list(ws.tables.keys()):
        table = ws.tables[name]
        table.name = "PendientesFeedback20260628Iter8"
        table.displayName = "PendientesFeedback20260628Iter8"
        break
    wb.save(OUTPUT_XLSX)
    return result


def validate_excel():
    wb = load_workbook(OUTPUT_XLSX, data_only=True)
    ws = wb["Pendientes"]
    headers = [cell.value for cell in ws[1]]
    price_col = headers.index("precio") + 1
    decision_col = headers.index("decision_manual") + 1
    notes_col = headers.index("notas_manual") + 1
    price_types = Counter(type(ws.cell(row, price_col).value).__name__ for row in range(2, ws.max_row + 1))
    return {
        "sheets": wb.sheetnames,
        "rows": ws.max_row - 1,
        "auto_filter": ws.auto_filter.ref,
        "tables": list(ws.tables.keys()),
        "freeze_panes": ws.freeze_panes,
        "price_types": dict(price_types),
        "nonblank_decisions": sum(1 for row in range(2, ws.max_row + 1) if ws.cell(row, decision_col).value),
        "nonblank_notes": sum(1 for row in range(2, ws.max_row + 1) if ws.cell(row, notes_col).value),
        "validations": [
            {"type": dv.type, "formula1": dv.formula1, "sqref": str(dv.sqref)}
            for dv in ws.data_validations.dataValidation
        ],
        "missing_columns": [
            column
            for column in (
                "id_propiedad",
                "titulo",
                "domicilio_actual",
                "zona_actual",
                "precio",
                "motivo_pendiente",
                "decision_manual",
                "notas_manual",
            )
            if column not in headers
        ],
    }


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--dry-run", action="store_true")
    parser.add_argument("--apply", action="store_true")
    parser.add_argument("--generate-excel", action="store_true")
    parser.add_argument("--json-out")
    args = parser.parse_args()
    if args.dry_run and args.apply:
        parser.error("--dry-run y --apply son mutuamente excluyentes")
    apply = bool(args.apply)
    summary = run(apply=apply)
    if args.generate_excel:
        summary["excel"] = generate_excel()
        summary["excel_validation"] = validate_excel()
    else:
        summary["excel"] = None
        summary["excel_validation"] = None
    output = json.dumps(summary, ensure_ascii=False, indent=2, default=str)
    print(output)
    if args.json_out:
        Path(args.json_out).write_text(output, encoding="utf-8")


if __name__ == "__main__":
    main()
