import argparse
import importlib.util
import io
import json
import os
import shutil
import sys
from collections import Counter, defaultdict
from pathlib import Path
from urllib.parse import urlparse

from django.core.management import call_command
from django.utils import timezone
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))
os.environ.setdefault("DJANGO_SETTINGS_MODULE", "config.settings")

import django

django.setup()

from properties.models import Listing, Property
from properties.services.geocoding import best_address
from properties.services.normalization import normalize_address

INPUT_XLSX = ROOT / "tmp" / "curacion_zonas_pendientes_2026-06-28_feedback_iter8.xlsx"
OUTPUT_XLSX = ROOT / "tmp" / "curacion_zonas_pendientes_2026-06-28_feedback_iter9.xlsx"
BACKUP_PATH = ROOT / "tmp" / "db.sqlite3.backup_curacion_feedback_20260628_iter9"
DB_PATH = ROOT / "db.sqlite3"
SCRIPT_TAG = "curacion_feedback_20260628_iter9"

APPROVED_IDS = {1024, 2118, 3960}
REJECTED_REMOVE_IDS = {956, 5606, 5866}
NO_ACTION_IDS = {1024}
MAP_ONLY_IDS = set()
MERGE_COMPONENTS = [(2118, 3960)]
CANONICAL_MERGE_ID = 2118

ADDRESS_FIXES = {
    2118: {
        "address": "German Argerich 1848",
        "locality": "Hurlingham",
        "reason": "feedback aprobado: Direccion curada German Argerich 1848",
    }
}

REMOVAL_REASONS = {
    956: "feedback RECHAZADO: Eliminar, esta es de Palomar",
    5606: "feedback RECHAZADO: Eliminar, esto es Villa Udaondo",
    5866: "feedback RECHAZADO: Eliminar, esta es de Martin Coronado",
}

SAMPLE_IDS = sorted(APPROVED_IDS | REJECTED_REMOVE_IDS)


def load_iter8_module():
    path = ROOT / "tmp" / "curacion_feedback_20260628_iter8.py"
    spec = importlib.util.spec_from_file_location("curacion_iter8_for_iter9", path)
    module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(module)
    module.SCRIPT_TAG = SCRIPT_TAG
    return module


BASE = load_iter8_module()


def safe(value):
    return BASE.safe(value)


def add(summary, key, value):
    summary.setdefault(key, []).append(value)


def read_feedback():
    workbook = load_workbook(INPUT_XLSX, data_only=True)
    worksheet = workbook["Pendientes"]
    headers = [cell.value for cell in worksheet[1]]
    index = {header: position + 1 for position, header in enumerate(headers)}
    rows = {}
    decisions = {}
    notes = {}
    counts = Counter()
    for row_number in range(2, worksheet.max_row + 1):
        property_id = int(worksheet.cell(row_number, index["id_propiedad"]).value)
        row = {header: worksheet.cell(row_number, index[header]).value for header in headers}
        decision = safe(row.get("decision_manual")).upper()
        note = safe(row.get("notas_manual"))
        rows[property_id] = row
        counts[decision or "VACIO"] += 1
        if decision:
            decisions[property_id] = decision
        if note:
            notes[property_id] = note
    return {
        "rows": worksheet.max_row - 1,
        "decision_counts": dict(counts),
        "rows_by_id": rows,
        "decisions": decisions,
        "notes": notes,
    }


def validation_state(feedback):
    decisions = feedback["decisions"]
    approved_found = sorted(property_id for property_id, decision in decisions.items() if decision == "APROBADO")
    rejected_found = sorted(property_id for property_id, decision in decisions.items() if decision == "RECHAZADO")
    return {
        "approved_found": approved_found,
        "rejected_found": rejected_found,
        "unexpected_approved": sorted(set(approved_found) - APPROVED_IDS),
        "unexpected_rejected": sorted(set(rejected_found) - REJECTED_REMOVE_IDS),
        "missing_expected_approved": sorted(APPROVED_IDS - set(approved_found)),
        "missing_expected_rejected": sorted(REJECTED_REMOVE_IDS - set(rejected_found)),
    }


def ensure_can_apply(summary):
    blockers = []
    for key in (
        "unexpected_approved",
        "unexpected_rejected",
        "missing_expected_approved",
        "missing_expected_rejected",
    ):
        if summary.get(key):
            blockers.append({key: summary[key]})
    missing = sorted(
        (APPROVED_IDS | REJECTED_REMOVE_IDS)
        - set(Property.objects.filter(pk__in=APPROVED_IDS | REJECTED_REMOVE_IDS).values_list("pk", flat=True))
    )
    if missing:
        blockers.append({"missing_properties": missing})
    if blockers:
        raise RuntimeError(f"Apply bloqueado por validacion: {blockers}")


def backup_db(summary):
    if BACKUP_PATH.exists():
        summary["backup"] = {"path": str(BACKUP_PATH), "created": False, "reason": "already_exists"}
        return
    shutil.copy2(DB_PATH, BACKUP_PATH)
    summary["backup"] = {"path": str(BACKUP_PATH), "created": True}


def mark_removed(summary, apply):
    for property_id in sorted(REJECTED_REMOVE_IDS):
        property_obj = Property.objects.prefetch_related("listings").get(pk=property_id)
        reason = REMOVAL_REASONS[property_id]
        active_listing_ids = list(property_obj.listings.filter(active=True).values_list("id", flat=True))
        add(
            summary,
            "removed",
            {
                "id": property_obj.pk,
                "title": property_obj.title,
                "reason": reason,
                "active_listing_ids": active_listing_ids,
            },
        )
        if not apply:
            continue
        property_obj.status = Property.Status.REMOVED
        property_obj.is_hidden = True
        evidence = dict(property_obj.zone_inference_evidence or {})
        evidence[SCRIPT_TAG] = {
            "action": "logical_remove",
            "reason": reason,
            "timestamp": timezone.now().isoformat(),
        }
        property_obj.zone_inference_evidence = evidence
        property_obj.save(update_fields=["status", "is_hidden", "zone_inference_evidence"])
        property_obj.listings.filter(active=True).update(active=False, source_status="removed")


def mark_no_action(feedback, summary, apply):
    for property_id in sorted(NO_ACTION_IDS):
        property_obj = Property.objects.get(pk=property_id)
        note = feedback["notes"].get(property_id) or "sin datos suficientes en la publicacion"
        add(summary, "reviewed_no_action", {"id": property_obj.pk, "note": note})
        if not apply:
            continue
        evidence = dict(property_obj.zone_inference_evidence or {})
        no_action = dict(evidence.get("curation_no_action") or {})
        no_action[SCRIPT_TAG] = {"note": note, "reviewed_at": timezone.now().isoformat()}
        evidence["curation_no_action"] = no_action
        property_obj.zone_inference_evidence = evidence
        property_obj.reviewed_at = timezone.now()
        property_obj.save(update_fields=["zone_inference_evidence", "reviewed_at"])


def apply_address_fix(summary, apply):
    for property_id, config in ADDRESS_FIXES.items():
        property_obj = Property.objects.select_related("location", "location_intelligence").get(pk=property_id)
        BASE.apply_address(property_obj, config["address"], config["locality"], summary, apply)
        if not apply:
            property_obj.address = config["address"]
            property_obj.locality = config["locality"]
            property_obj.normalized_address = normalize_address(config["address"])
        if apply:
            property_obj.refresh_from_db()

        location = BASE.geocode_property(property_obj, summary, apply, allow_external=True, force=True)
        if apply:
            property_obj.refresh_from_db()
            location = getattr(property_obj, "location", None)

        if not location and not apply:
            add(
                summary,
                "geocoding_required",
                {
                    "id": property_obj.pk,
                    "address": best_address(property_obj),
                    "reason": "sin coordenada cacheada en dry-run; API permitida solo en apply",
                },
            )
            add(summary, "zone_skipped", {"id": property_obj.pk, "reason": "dry-run sin coordenadas nuevas"})
            continue
        if location and getattr(location, "outside_target", False):
            add(
                summary,
                "zone_skipped",
                {
                    "id": property_obj.pk,
                    "reason": "geocoding fuera de Hurlingham",
                    "lat": location.latitude,
                    "lon": location.longitude,
                    "provider": location.provider,
                },
            )
            continue
        BASE.infer_and_apply_zone(
            property_obj,
            summary,
            apply,
            reason=config["reason"],
            allow_review=True,
        )


def run_merge(summary, apply):
    output = io.StringIO()
    component = ",".join(str(item) for item in MERGE_COMPONENTS[0])
    call_command(
        "merge_properties",
        component=[component],
        canonical_id=CANONICAL_MERGE_ID,
        dry_run=not apply,
        stdout=output,
    )
    add(
        summary,
        "merged",
        {
            "component": list(MERGE_COMPONENTS[0]),
            "canonical_id": CANONICAL_MERGE_ID,
            "dry_run": not apply,
            "output": output.getvalue().strip().splitlines(),
        },
    )


def detect_mapaprop_patagonprop_duplicates(summary):
    groups = defaultdict(list)
    listings = Listing.objects.select_related("source").filter(source__slug__in=["mapaprop", "patagonprop"])
    for listing in listings:
        tail = urlparse(listing.url or "").path.rstrip("/").rsplit("/", 1)[-1].strip().lower()
        if tail:
            groups[tail].append(listing)

    detected = []
    for tail, items in groups.items():
        property_ids = sorted({item.property_id for item in items})
        source_slugs = sorted({item.source.slug for item in items if item.source})
        if len(property_ids) > 1 and {"mapaprop", "patagonprop"}.issubset(source_slugs):
            detected.append({"tail": tail, "property_ids": property_ids, "sources": source_slugs})
    detected.sort(key=lambda item: item["property_ids"])
    target = [
        item
        for item in detected
        if set(item["property_ids"]) == set(MERGE_COMPONENTS[0])
    ]
    skipped = [item for item in detected if item not in target]
    summary["duplicate_url_tail_report"] = {
        "total_groups": len(detected),
        "target_groups": target,
        "skipped_groups_count": len(skipped),
        "skipped_groups_sample": skipped[:25],
    }


def run(apply=False):
    feedback = read_feedback()
    summary = {
        "script": SCRIPT_TAG,
        "mode": "apply" if apply else "dry_run",
        "input_excel": str(INPUT_XLSX),
        "output_excel": str(OUTPUT_XLSX) if apply else "",
        "feedback": {
            "rows": feedback["rows"],
            "decision_counts": feedback["decision_counts"],
            "decisions": {str(key): value for key, value in feedback["decisions"].items()},
            "notes": {str(key): value for key, value in feedback["notes"].items()},
        },
        "before": BASE.current_counts(),
        "decision_sets": {
            "APPROVED_IDS": sorted(APPROVED_IDS),
            "REJECTED_REMOVE_IDS": sorted(REJECTED_REMOVE_IDS),
            "NO_ACTION_IDS": sorted(NO_ACTION_IDS),
            "MAP_ONLY_IDS": sorted(MAP_ONLY_IDS),
        },
    }
    summary.update(validation_state(feedback))
    ensure_can_apply(summary)
    detect_mapaprop_patagonprop_duplicates(summary)
    if apply:
        backup_db(summary)
    mark_removed(summary, apply)
    mark_no_action(feedback, summary, apply)
    apply_address_fix(summary, apply)
    run_merge(summary, apply)
    summary["after"] = BASE.current_counts()
    summary["sample"] = {str(property_id): BASE.property_snapshot(property_id) for property_id in SAMPLE_IDS}
    return summary


def load_iter6_module():
    path = ROOT / "tmp" / "curacion_feedback_20260628_iter6.py"
    spec = importlib.util.spec_from_file_location("curacion_iter6_for_iter9", path)
    module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(module)
    return module


def generate_excel():
    module = load_iter6_module()
    module.OUTPUT_XLSX = OUTPUT_XLSX
    module.SCRIPT_TAG = SCRIPT_TAG
    result = module.generate_excel({})
    workbook = load_workbook(OUTPUT_XLSX)
    worksheet = workbook["Pendientes"]
    for name in list(worksheet.tables.keys()):
        table = worksheet.tables[name]
        table.name = "PendientesFeedback20260628Iter9"
        table.displayName = "PendientesFeedback20260628Iter9"
        break
    workbook.save(OUTPUT_XLSX)
    return result


def validate_excel():
    workbook = load_workbook(OUTPUT_XLSX, data_only=True)
    worksheet = workbook["Pendientes"]
    headers = [cell.value for cell in worksheet[1]]
    price_col = headers.index("precio") + 1
    decision_col = headers.index("decision_manual") + 1
    notes_col = headers.index("notas_manual") + 1
    price_types = Counter(
        type(worksheet.cell(row, price_col).value).__name__ for row in range(2, worksheet.max_row + 1)
    )
    return {
        "sheets": workbook.sheetnames,
        "rows": worksheet.max_row - 1,
        "auto_filter": worksheet.auto_filter.ref,
        "tables": list(worksheet.tables.keys()),
        "freeze_panes": worksheet.freeze_panes,
        "price_types": dict(price_types),
        "nonblank_decisions": sum(
            1 for row in range(2, worksheet.max_row + 1) if worksheet.cell(row, decision_col).value
        ),
        "nonblank_notes": sum(
            1 for row in range(2, worksheet.max_row + 1) if worksheet.cell(row, notes_col).value
        ),
        "validations": [
            {"type": validation.type, "formula1": validation.formula1, "sqref": str(validation.sqref)}
            for validation in worksheet.data_validations.dataValidation
        ],
        "missing_columns": [
            column
            for column in (
                "id_propiedad",
                "titulo",
                "domicilio_actual",
                "zona_actual",
                "precio",
                "motivo_pendiente",
                "decision_manual",
                "notas_manual",
            )
            if column not in headers
        ],
    }


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--dry-run", action="store_true")
    parser.add_argument("--apply", action="store_true")
    parser.add_argument("--generate-excel", action="store_true")
    parser.add_argument("--json-out")
    args = parser.parse_args()
    if args.dry_run and args.apply:
        parser.error("--dry-run y --apply son mutuamente excluyentes")
    apply = bool(args.apply)
    summary = run(apply=apply)
    if args.generate_excel:
        summary["excel"] = generate_excel()
        summary["excel_validation"] = validate_excel()
    else:
        summary["excel"] = None
        summary["excel_validation"] = None
    output = json.dumps(summary, ensure_ascii=False, indent=2, default=str)
    print(output)
    if args.json_out:
        Path(args.json_out).write_text(output, encoding="utf-8")


if __name__ == "__main__":
    main()
