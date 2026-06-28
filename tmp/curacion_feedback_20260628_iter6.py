import argparse
import json
import os
import re
import shutil
import sys
import time
from collections import Counter, defaultdict
from pathlib import Path

import requests
from bs4 import BeautifulSoup
from django.conf import settings
from django.db.models import Q
from django.utils import timezone
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))
os.environ.setdefault("DJANGO_SETTINGS_MODULE", "config.settings")

import django

django.setup()

from properties.models import GeocodeCache, Listing, Property, PropertyLocation, PropertyLocationIntelligence
from properties.scrapers.argencasas import ArgencasasScraper
from properties.scrapers.local_sites import BecerraScraper, FaellaScraper
from properties.scrapers.pending_sources import (
    FincasScraper,
    GuarnieriScraper,
    OscarDahbarScraper,
    ZonapropScraper,
    is_declared_out_of_target,
    normalize_argencasas_address,
)
from properties.scrapers.parsing import extract_map_coordinates
from properties.services.geocoding import Geocoder, best_address
from properties.services.normalization import (
    canonical_address_alias,
    classify_address_precision,
    normalize_address,
    normalize_whitespace,
)
from properties.services.territory_hierarchy import infer_territory_for_point
from properties.services.zone_inference import infer_zone_for_point


INPUT_XLSX = ROOT / "tmp" / "curacion_zonas_pendientes_2026-06-27_feedback_iter5.xlsx"
OUTPUT_XLSX = ROOT / "tmp" / "curacion_zonas_pendientes_2026-06-28_feedback_iter6.xlsx"
BACKUP_PATH = ROOT / "tmp" / "db.sqlite3.backup_curacion_feedback_20260628_iter6"
DB_PATH = ROOT / "db.sqlite3"
SCRIPT_TAG = "curacion_feedback_20260628_iter6"
ML_CHROME_RESULTS_PATH = ROOT / "tmp" / "curacion_ml_chrome_iter6.json"

APPROVED_IDS = {
    3181,
    5704,
}
REJECTED_REMOVE_IDS = set()
NO_ACTION_IDS = set()
MAP_ONLY_IDS = set()
CONFIRM_REMOVE_IDS = set()
PROPAGATE_IDS = set()
EXISTING_COORDINATE_ZONES = {
    548: "Santa Leonor",
    1065: "Luna",
    1707: "San Alberto",
    2166: "Luna",
    3329: "El Libertador",
    3449: "Nueve de Julio",
    3612: "Villa Club",
    4040: "Luna",
    4324: "San Alberto",
    5766: "Carola Lorenzini",
    5859: "Carola Lorenzini",
    5898: "San Alberto",
    6126: "Cosmopolita",
    6128: "El Libertador",
    6129: "Rimoldi",
    6130: "Villa Club",
    6131: "Mitre",
    6132: "El Porvenir",
    6133: "El Libertador",
}
COORDINATE_NO_MATCH_IDS = {
    66: "coordenada muy lejos de poligonos locales; nearest Asuncion a ~11948m",
    956: "ubicacion manual fuera de target; nearest La Juanita a ~2980m",
    1024: "nearest Dos de Abril a ~695m, supera umbral automatico",
    2118: "coordenada fuera/lejana; nearest La Juanita a ~12483m",
    3401: "nearest San Damian a ~462m, supera umbral automatico",
    3960: "coordenada fuera/lejana; nearest La Juanita a ~12483m",
    4409: "nearest San Alberto a ~106m, apenas supera umbral automatico de 100m",
    5606: "nearest Dos de Abril a ~1059m, supera umbral automatico",
    5866: "nearest Villa Alemania a ~1411m, supera umbral automatico",
    5867: "nearest Villa Alemania a ~1530m, supera umbral automatico",
    6123: "coordenada fuera/lejana; nearest Villa Alemania a ~17694m",
}
NO_ADDRESS_REPARSE_IDS = set()
INTERSECTION_IDS = {5722}
SAMPLE_IDS = sorted(
    set(APPROVED_IDS)
    | set(EXISTING_COORDINATE_ZONES)
    | set(COORDINATE_NO_MATCH_IDS)
    | {5706, 5707, 5708, 5709, 5710, 5713, 5716, 5718, 5719, 5721, 5722, 5723, 5724, 5726, 5729, 5731}
)

DIRECT_FIXES = {
    3181: {
        "address": "Virriato Unia 2412",
        "locality": "Hurlingham",
        "geocode": True,
        "external_geocode": True,
        "evidence": "feedback aprobado; cache previa sin coordenadas y fuente sin mapa embebido",
    },
    5704: {
        "address": "Julian Aguirre 2409",
        "locality": "Hurlingham",
        "geocode": True,
        "external_geocode": True,
        "evidence": "feedback aprobado + Chrome MercadoLibre",
    },
    5706: {"address": "Sgto. Jose Mariano Gomez 1520", "locality": "Hurlingham", "geocode": True, "external_geocode": True, "evidence": "Chrome MercadoLibre"},
    5707: {"address": "Sta. Ana 922", "locality": "Hurlingham", "geocode": True, "external_geocode": True, "evidence": "Chrome MercadoLibre"},
    5708: {"address": "Dip. Hector Finochietto 1527", "locality": "Hurlingham", "geocode": True, "external_geocode": True, "evidence": "Chrome MercadoLibre"},
    5709: {"address": "Juan Berduc 2175", "locality": "Hurlingham", "geocode": True, "external_geocode": True, "evidence": "Chrome MercadoLibre"},
    5710: {"address": "Altolaguirre 2435", "locality": "Hurlingham", "geocode": True, "external_geocode": True, "evidence": "Chrome MercadoLibre"},
    5713: {"address": "Dr. Delfor Diaz 1775", "locality": "Hurlingham", "geocode": True, "external_geocode": True, "evidence": "Chrome MercadoLibre"},
    5716: {"address": "Int. Mustoni 2063", "locality": "Hurlingham", "geocode": True, "external_geocode": True, "evidence": "Chrome MercadoLibre"},
    5718: {"address": "Uspallata 1940", "locality": "Hurlingham", "geocode": True, "external_geocode": True, "evidence": "Chrome MercadoLibre"},
    5719: {"address": "Felix Delatte 42", "locality": "Villa Tesei", "geocode": True, "external_geocode": True, "evidence": "Chrome MercadoLibre"},
    5721: {"address": "Potosi 565", "locality": "Hurlingham", "geocode": True, "external_geocode": True, "evidence": "Chrome MercadoLibre"},
    5722: {"address": "De La Trinidad y R. Prack", "locality": "Villa Tesei", "geocode": True, "external_geocode": True, "evidence": "Chrome MercadoLibre interseccion visible"},
    5723: {"address": "Francisco de Gurruchaga 2248", "locality": "Hurlingham", "geocode": True, "external_geocode": True, "evidence": "Chrome MercadoLibre"},
    5724: {"address": "La Patria 176", "locality": "Villa Tesei", "geocode": True, "external_geocode": True, "evidence": "Chrome MercadoLibre"},
    5726: {"address": "Coraceros 2456", "locality": "Hurlingham", "geocode": True, "external_geocode": True, "evidence": "Chrome MercadoLibre"},
    5729: {"address": "Rio Colorado 2198", "locality": "Hurlingham", "geocode": True, "external_geocode": True, "evidence": "Chrome MercadoLibre"},
    5731: {"address": "Beethoven 1283", "locality": "Hurlingham", "geocode": True, "external_geocode": True, "evidence": "Chrome MercadoLibre"},
}

HEADERS = [
    "id_propiedad",
    "titulo",
    "descripcion_resumen",
    "domicilio_actual",
    "direccion_detectada",
    "direccion_sugerida",
    "direccion_normalizada",
    "localidad_actual",
    "localidad_detectada",
    "barrio_actual",
    "barrio_detectado",
    "zona_actual",
    "zona_sugerida",
    "zona_inteligencia",
    "confianza",
    "motivo",
    "estado",
    "latitud",
    "longitud",
    "fuente_coordenadas",
    "fuente",
    "inmobiliaria",
    "url_publicacion",
    "motivo_url",
    "estado_publicacion",
    "tipo_propiedad",
    "operacion",
    "moneda",
    "precio",
    "ambientes",
    "dormitorios",
    "banos",
    "superficie_cubierta",
    "superficie_total",
    "fecha_ultima_vista",
    "motivo_pendiente",
    "evidencia",
    "decision_manual",
    "notas_manual",
]

SOURCE_DELAYS = {
    "argencasas": 1,
    "becerra": 1,
    "guarnieri": 1,
    "fincas": 1,
    "oscar-dahbar": 1,
    "remax-datawork": 1,
    "miglierini": 1,
    "zonaprop": 1,
    "faella": 1,
    "mercadolibre": 1,
    "hollmann-ariel": 1,
    "nerina-allo": 1,
    "argenprop": 1,
    "inmuebles-clarin": 1,
}
REQUEST_TIMEOUT = (3, 5)
FEEDBACK_IDS = APPROVED_IDS | REJECTED_REMOVE_IDS
SWEEP_EXTRA_LIMITS = {
    "becerra": 8,
    "zonaprop": 100,
    "faella": 40,
    "guarnieri": 5,
    "argencasas": 5,
}
FETCH_STATE = {}
BAD_OSCAR_ADDRESS_RE = re.compile(r"\bdormitorios\s+y\s+parque\b", re.I)
BAD_ADDRESS_RE = re.compile(
    r"\b(?:amplio\s+con\s+parrilla\s+hurlingham\s+ver|mensaje\s+al\s+anunciante|"
    r"dormitorios\s+y\s+patio\s+con\s+parrilla\s+hurlingham)\b",
    re.I,
)
BECERRA_RETIRED_RE = re.compile(
    r"whoops!\s+we\s+seem\s+to\s+have\s+hit\s+a\s+snag|propiedad\s+(?:retirada|no\s+disponible)",
    re.I,
)
ZONAPROP_RETIRED_RE = re.compile(
    r"este\s+aviso\s+ya\s+no\s+est[aá]\s+publicado|aviso\s+(?:finalizado|no\s+disponible)|publicaci[oó]n\s+finalizada",
    re.I,
)
ZONAPROP_OFFLINE_RE = re.compile(
    r"(?:offline-container|section-offline-disclaimer|offline-finish-message).{0,1200}?"
    r"(?:este\s+aviso\s+ya\s+no\s+est[aÃ¡]\s+publicado|fue\s+finalizado\s+por\s+el\s+anunciante)",
    re.I | re.S,
)


def safe(value):
    return normalize_whitespace(str(value)) if value not in (None, "") else ""


def no_zone_query():
    return (
        (Q(inferred_zone__isnull=True) | Q(inferred_zone=""))
        & (
            Q(location_intelligence__isnull=True)
            | Q(location_intelligence__zone_name__isnull=True)
            | Q(location_intelligence__zone_name="")
        )
    )


def location_or_none(property_obj):
    return getattr(property_obj, "location", None)


def intel_or_none(property_obj):
    return getattr(property_obj, "location_intelligence", None)


def curation_no_action(property_obj):
    evidence = property_obj.zone_inference_evidence or {}
    return bool(evidence.get("curation_no_action") or {})


def residual_reason(property_obj):
    reasons = []
    intel = intel_or_none(property_obj)
    if not property_obj.inferred_zone and not (intel.zone_name if intel else ""):
        reasons.append("sin zona")
    if (
        not best_address(property_obj)
        or is_bad_property_address(property_obj.address)
        or is_bad_property_address(property_obj.detected_address)
    ):
        reasons.append("sin direccion util")
    return ", ".join(reasons)


def is_bad_oscar_address(value):
    return bool(BAD_OSCAR_ADDRESS_RE.search(safe(value)))


def is_bad_property_address(value):
    text = safe(value)
    return bool(BAD_OSCAR_ADDRESS_RE.search(text) or BAD_ADDRESS_RE.search(text))


def current_counts():
    props = list(
        Property.objects.select_related("location", "location_intelligence").prefetch_related(
            "listings__source"
        )
    )
    by_source = defaultdict(lambda: Counter())
    for property_obj in props:
        if property_obj.is_hidden or property_obj.status == Property.Status.REMOVED or curation_no_action(property_obj):
            continue
        reason = residual_reason(property_obj)
        if not reason:
            continue
        sources = {listing.source.slug for listing in property_obj.listings.all()} or {"sin_listing"}
        for source in sources:
            by_source[source]["pendientes"] += 1
            if "sin zona" in reason:
                by_source[source]["sin_zona"] += 1
            if "sin direccion" in reason:
                by_source[source]["sin_direccion"] += 1
    no_zone = [
        prop
        for prop in props
        if not prop.is_hidden
        and prop.status != Property.Status.REMOVED
        and not curation_no_action(prop)
        and residual_reason(prop)
        and "sin zona" in residual_reason(prop)
    ]
    no_address = [
        prop
        for prop in props
        if not prop.is_hidden
        and prop.status != Property.Status.REMOVED
        and not curation_no_action(prop)
        and (not best_address(prop) or is_bad_property_address(best_address(prop)))
    ]
    return {
        "sin_zona_operativa": len(no_zone),
        "sin_direccion_util": len(no_address),
        "pendientes_por_fuente": dict(sorted((source, dict(counts)) for source, counts in by_source.items())),
    }


def load_feedback():
    wb = load_workbook(INPUT_XLSX, read_only=True, data_only=False)
    ws = wb["Pendientes"]
    headers = [cell.value for cell in ws[1]]
    rows = {}
    decisions = Counter()
    notes = {}
    for values in ws.iter_rows(min_row=2, values_only=True):
        row = dict(zip(headers, values))
        try:
            property_id = int(row.get("id_propiedad") or 0)
        except (TypeError, ValueError):
            continue
        rows[property_id] = row
        decision = safe(row.get("decision_manual")).upper()
        if decision:
            decisions[decision] += 1
        note = safe(row.get("notas_manual"))
        if note:
            notes[property_id] = note
    return rows, decisions, notes


def add_summary(summary, key, item):
    summary.setdefault(key, []).append(item)


def first_listing(property_obj):
    return property_obj.listings.select_related("source", "agency").order_by("-active", "-last_seen_at").first()


def first_listing_url(property_obj):
    listing = first_listing(property_obj)
    return listing.url if listing else ""


def mark_manual_property_field(property_obj, field, value, now, update_fields, apply):
    value = safe(value)
    if not value:
        return
    changed = getattr(property_obj, field) != value
    if changed:
        setattr(property_obj, field, value)
        update_fields.add(field)
    if field == "address":
        normalized = normalize_address(value)
        if property_obj.normalized_address != normalized:
            property_obj.normalized_address = normalized
            update_fields.add("normalized_address")
        current = location_or_none(property_obj)
        if changed and current and not current.manually_corrected and apply:
            current.delete()
    overrides = dict(property_obj.manual_overrides or {})
    overrides[field] = now.isoformat()
    property_obj.manual_overrides = overrides
    property_obj.data_manually_corrected_at = now
    update_fields.update({"manual_overrides", "data_manually_corrected_at"})


def save_property(property_obj, update_fields, apply):
    if update_fields and apply:
        property_obj.save(update_fields=sorted(update_fields))


def update_location(property_obj, latitude, longitude, provider, query, summary, apply):
    current = location_or_none(property_obj)
    if current and current.manually_corrected:
        add_summary(summary, "manual_location_preserved", property_obj.pk)
        return False
    defaults = {
        "latitude": float(latitude),
        "longitude": float(longitude),
        "precision": PropertyLocation.Precision.EXACT,
        "query": query or "",
        "provider": provider,
        "confidence": 1.0,
        "outside_target": False,
        "manually_corrected": False,
    }
    if apply:
        PropertyLocation.objects.update_or_create(property=property_obj, defaults=defaults)
        evidence = dict(property_obj.location_evidence or {})
        evidence[SCRIPT_TAG] = {
            "provider": provider,
            "latitude": float(latitude),
            "longitude": float(longitude),
            "query": query or "",
        }
        property_obj.detected_latitude = float(latitude)
        property_obj.detected_longitude = float(longitude)
        property_obj.location_source = Property.LocationSource.MAP
        property_obj.location_confidence = Property.LocationConfidence.HIGH
        property_obj.location_evidence = evidence
        property_obj.save(
            update_fields=[
                "detected_latitude",
                "detected_longitude",
                "location_source",
                "location_confidence",
                "location_evidence",
            ]
        )
    return True


def apply_zone_from_point(property_obj, latitude, longitude, provider, reason, summary, apply):
    result = infer_territory_for_point(latitude, longitude)
    zone = safe(result.zone if result else "")
    if not zone:
        add_summary(summary, "zone_skipped", {"id": property_obj.pk, "reason": "sin_match_poligono"})
        return ""
    now = timezone.now()
    evidence = dict(result.evidence if result and result.evidence else {})
    evidence.update({"provider": provider, "curacion": SCRIPT_TAG, "reason": reason})
    if apply:
        property_obj.inferred_partido = result.partido
        property_obj.inferred_locality = result.locality
        property_obj.inferred_zone = zone
        property_obj.inferred_neighborhood = zone
        property_obj.territory_confidence = result.confidence
        property_obj.territory_source_method = result.source_method
        property_obj.territory_needs_review = bool(result.needs_review)
        property_obj.territory_evidence = evidence
        property_obj.territory_inferred_at = now
        property_obj.zone_conflict = False
        property_obj.zone_needs_review = bool(result.needs_review)
        property_obj.zone_inference_evidence = evidence
        property_obj.zone_inferred_at = now
        property_obj.save(
            update_fields=[
                "inferred_partido",
                "inferred_locality",
                "inferred_zone",
                "inferred_neighborhood",
                "territory_confidence",
                "territory_source_method",
                "territory_needs_review",
                "territory_evidence",
                "territory_inferred_at",
                "zone_conflict",
                "zone_needs_review",
                "zone_inference_evidence",
                "zone_inferred_at",
            ]
        )
        PropertyLocationIntelligence.objects.update_or_create(
            property=property_obj,
            defaults={
                "partido_name": result.partido,
                "locality_name": result.locality,
                "zone_name": zone,
                "match_method": PropertyLocationIntelligence.MatchMethod.COORDINATES,
                "confidence": result.confidence,
            },
        )
    add_summary(summary, "zone_applied", {"id": property_obj.pk, "zone": zone, "provider": provider})
    return zone


def mark_removed(property_obj, reason, summary, apply):
    listing_ids = list(property_obj.listings.filter(active=True).values_list("id", flat=True))
    if apply:
        property_obj.listings.filter(active=True).update(active=False, source_status="removed", missing_runs=2)
        marker = f"{SCRIPT_TAG}: {reason}"
        notes = property_obj.personal_notes or ""
        if marker not in notes:
            property_obj.personal_notes = "\n\n".join(part for part in [notes, marker] if part)
        property_obj.status = Property.Status.REMOVED
        property_obj.is_hidden = True
        property_obj.save(update_fields=["status", "is_hidden", "personal_notes"])
    add_summary(summary, "removed", {"id": property_obj.pk, "listing_ids": listing_ids, "reason": reason})


def mark_reviewed_no_action(property_obj, note, summary, apply):
    if apply:
        evidence = dict(property_obj.zone_inference_evidence or {})
        no_action = dict(evidence.get("curation_no_action") or {})
        no_action[SCRIPT_TAG] = {"note": note or "sin datos utiles publicados", "reviewed_at": timezone.now().isoformat()}
        evidence["curation_no_action"] = no_action
        property_obj.zone_inference_evidence = evidence
        property_obj.reviewed_at = timezone.now()
        property_obj.save(update_fields=["zone_inference_evidence", "reviewed_at"])
    add_summary(summary, "reviewed_no_action", {"id": property_obj.pk, "note": note})


def fetch_url_text(url, source_slug=""):
    delay = SOURCE_DELAYS.get(source_slug, 1)
    elapsed = time.monotonic() - FETCH_STATE.get(source_slug, 0)
    if elapsed < delay:
        time.sleep(delay - elapsed)
    session = requests.Session()
    session.headers.update({"User-Agent": "Mozilla/5.0 (compatible; RadarCuration/1.0)"})
    response = session.get(url, timeout=REQUEST_TIMEOUT)
    FETCH_STATE[source_slug] = time.monotonic()
    return response.status_code, response.text or ""


def parse_with_scraper(scraper_cls, url, markup):
    scraper = scraper_cls()
    soup = BeautifulSoup(markup or "", "lxml")
    scraper.soup = lambda _url: soup
    return scraper.parse(url)


def has_removed_marker(status_code, markup):
    return status_code == 404 and re.search(
        r"propiedad\s+ha\s+sido\s+retirada\s+del\s+sistema|publicaci[oó]n\s+retirada",
        markup or "",
        re.I,
    )


def confirm_argencasas_removed(property_obj, summary, apply):
    listing = first_listing(property_obj)
    if not listing or not listing.url:
        add_summary(summary, "remove_skipped", {"id": property_obj.pk, "reason": "sin_url"})
        return
    try:
        status_code, markup = fetch_url_text(listing.url, "argencasas")
    except Exception as exc:
        add_summary(summary, "remove_skipped", {"id": property_obj.pk, "reason": repr(exc)})
        return
    if has_removed_marker(status_code, markup):
        mark_removed(property_obj, "Argencasas 404 + propiedad retirada del sistema", summary, apply)
    else:
        add_summary(summary, "remove_skipped", {"id": property_obj.pk, "http_status": status_code})


SCRAPER_BY_SOURCE = {
    "becerra": BecerraScraper,
    "fincas": FincasScraper,
    "zonaprop": ZonapropScraper,
    "guarnieri": GuarnieriScraper,
}


def parse_source_data(source_slug, url, markup, summary, property_id):
    scraper_cls = SCRAPER_BY_SOURCE.get(source_slug)
    if not scraper_cls:
        return None
    try:
        return parse_with_scraper(scraper_cls, url, markup)
    except Exception as exc:
        add_summary(summary, "source_parse_errors", {"id": property_id, "source": source_slug, "error": repr(exc)})
        return None


def source_map_coordinates(source_slug, url, markup, summary, property_id):
    parsed = parse_source_data(source_slug, url, markup, summary, property_id)
    if parsed and parsed.get("latitude") is not None and parsed.get("longitude") is not None:
        return {
            "latitude": parsed["latitude"],
            "longitude": parsed["longitude"],
            "method": f"{source_slug}_parser",
            "data": parsed,
        }
    coordinates = extract_map_coordinates(markup, require_target_bounds=True)
    if coordinates:
        coordinate = coordinates[0]
        coordinate["data"] = parsed or {}
        return coordinate
    return {"data": parsed or {}}


def apply_source_map_if_available(property_obj, reason, summary, apply, preferred_source_slug=""):
    listing = first_listing(property_obj)
    if not listing or not listing.url:
        add_summary(summary, "source_map_skipped", {"id": property_obj.pk, "reason": "sin_url"})
        return False
    source_slug = preferred_source_slug or (listing.source.slug if listing.source else "")
    try:
        status_code, markup = fetch_url_text(listing.url, source_slug)
    except Exception as exc:
        add_summary(summary, "source_fetch_errors", {"id": property_obj.pk, "source": source_slug, "error": repr(exc)})
        return False
    add_summary(summary, "source_checked", {"id": property_obj.pk, "source": source_slug, "http_status": status_code})
    if status_code >= 400:
        return False
    coordinate = source_map_coordinates(source_slug, listing.url, markup, summary, property_obj.pk)
    data = coordinate.get("data") or {}
    if data.get("address") and not best_address(property_obj) and source_slug in {"becerra", "fincas", "zonaprop", "faella"}:
        now = timezone.now()
        update_fields = set()
        mark_manual_property_field(property_obj, "address", data["address"], now, update_fields, apply)
        if data.get("locality"):
            mark_manual_property_field(property_obj, "locality", data["locality"], now, update_fields, apply)
        save_property(property_obj, update_fields, apply)
        add_summary(summary, "source_address_updates", {"id": property_obj.pk, "source": source_slug, "to": data["address"]})
    if coordinate.get("latitude") is None or coordinate.get("longitude") is None:
        add_summary(summary, "source_map_miss", {"id": property_obj.pk, "source": source_slug})
        return False
    provider = f"{source_slug}_map"
    if update_location(property_obj, coordinate["latitude"], coordinate["longitude"], provider, listing.url, summary, apply):
        add_summary(
            summary,
            "sweep_map_locations",
            {"id": property_obj.pk, "source": source_slug, "method": coordinate.get("method")},
        )
    apply_zone_from_point(
        property_obj,
        coordinate["latitude"],
        coordinate["longitude"],
        provider,
        reason,
        summary,
        apply,
    )
    return True


def apply_geocode(property_obj, reason, summary, apply):
    add_summary(summary, "geocode_required", property_obj.pk)
    if not apply:
        apply_source_map_if_available(property_obj, reason, summary, apply)
        return
    property_obj.refresh_from_db()
    current = location_or_none(property_obj)
    intel = intel_or_none(property_obj)
    if current and not current.outside_target:
        if property_obj.inferred_zone or (intel.zone_name if intel else ""):
            add_summary(summary, "geocode_skipped_already_resolved", property_obj.pk)
            return
        apply_zone_from_point(property_obj, current.latitude, current.longitude, current.provider, reason, summary, apply)
        return
    if apply_source_map_if_available(property_obj, reason, summary, apply):
        return
    geocoder = Geocoder()
    location = geocoder.geocode_property_from_cache(property_obj, force=True)
    if location and not location.outside_target:
        add_summary(
            summary,
            "geocode_location",
            {
                "id": property_obj.pk,
                "lat": location.latitude,
                "lon": location.longitude,
                "provider": location.provider,
            },
        )
        apply_zone_from_point(property_obj, location.latitude, location.longitude, location.provider, reason, summary, apply)
    elif location:
        add_summary(summary, "geocode_outside_target", {"id": property_obj.pk, "lat": location.latitude, "lon": location.longitude})
    else:
        add_summary(summary, "geocode_cache_local_miss", property_obj.pk)


def external_geocode_property(property_obj, reason, summary, apply):
    add_summary(summary, "geocode_external_required", property_obj.pk)
    if not apply:
        return
    property_obj.refresh_from_db()
    intel = intel_or_none(property_obj)
    if property_obj.inferred_zone or (intel.zone_name if intel else ""):
        add_summary(summary, "geocode_external_skipped_already_resolved", property_obj.pk)
        return
    current = location_or_none(property_obj)
    if current and current.manually_corrected:
        add_summary(summary, "manual_location_preserved", property_obj.pk)
        return
    if current and not current.outside_target:
        apply_zone_from_point(property_obj, current.latitude, current.longitude, current.provider, reason, summary, apply)
        return

    geocoder = Geocoder()
    queries = geocoder.query_candidates(property_obj)
    if not queries:
        add_summary(summary, "geocode_external_no_query", property_obj.pk)
        return
    for query in queries:
        try:
            with Geocoder._rate_lock:
                elapsed = time.monotonic() - Geocoder._last_request_at
                if elapsed < 1:
                    time.sleep(1 - elapsed)
                response = geocoder.session.get(
                    settings.NOMINATIM_URL,
                    params={
                        "q": query,
                        "format": "jsonv2",
                        "limit": 1,
                        "countrycodes": "ar",
                        "addressdetails": 1,
                    },
                    headers={"User-Agent": settings.NOMINATIM_USER_AGENT},
                    timeout=20,
                )
                Geocoder._last_request_at = time.monotonic()
            response.raise_for_status()
            results = response.json()
        except Exception as exc:
            add_summary(summary, "geocode_external_errors", {"id": property_obj.pk, "query": query, "error": repr(exc)})
            continue
        result = results[0] if results else {}
        cache, _ = GeocodeCache.objects.update_or_create(
            query=query,
            defaults={
                "latitude": float(result["lat"]) if result else None,
                "longitude": float(result["lon"]) if result else None,
                "precision": classify_address_precision(best_address(property_obj)) if result else "",
                "confidence": float(result.get("importance", 0)) if result else 0,
                "provider_payload": result,
            },
        )
        location = geocoder._apply(property_obj, query, cache)
        if location and not location.outside_target:
            add_summary(
                summary,
                "geocode_external_location",
                {"id": property_obj.pk, "lat": location.latitude, "lon": location.longitude, "query": query},
            )
            apply_zone_from_point(property_obj, location.latitude, location.longitude, location.provider, reason, summary, apply)
            return
        if location and location.outside_target:
            add_summary(summary, "geocode_external_outside_target", {"id": property_obj.pk, "lat": location.latitude, "lon": location.longitude, "query": query})
            return
    add_summary(summary, "geocode_external_no_result", property_obj.pk)


def apply_direct_fixes(summary, apply, allowed_ids=None):
    allowed_ids = set(allowed_ids or DIRECT_FIXES)
    now = timezone.now()
    for property_id, fix in DIRECT_FIXES.items():
        if property_id not in allowed_ids:
            add_summary(summary, "direct_fix_skipped", {"id": property_id, "reason": "sin_aprobacion_excel"})
            continue
        property_obj = Property.objects.select_related("location", "location_intelligence").filter(pk=property_id).first()
        if not property_obj:
            add_summary(summary, "missing_property", property_id)
            continue
        update_fields = set()
        for field in ("address", "locality", "neighborhood"):
            if fix.get(field):
                mark_manual_property_field(property_obj, field, fix[field], now, update_fields, apply)
                add_summary(summary, f"{field}_updates", {"id": property_id, "to": fix[field]})
        if fix.get("evidence"):
            evidence = dict(property_obj.zone_inference_evidence or {})
            evidence.setdefault("curation_address_evidence", {})[SCRIPT_TAG] = fix["evidence"]
            property_obj.zone_inference_evidence = evidence
            update_fields.add("zone_inference_evidence")
        if fix.get("property_type") and property_obj.property_type != fix["property_type"]:
            property_obj.property_type = fix["property_type"]
            update_fields.add("property_type")
            evidence = dict(property_obj.zone_inference_evidence or {})
            evidence[SCRIPT_TAG] = {
                **dict(evidence.get(SCRIPT_TAG) or {}),
                "property_type_evidence": fix.get("property_type_evidence") or "correccion manual de tipo",
            }
            property_obj.zone_inference_evidence = evidence
            update_fields.add("zone_inference_evidence")
            add_summary(
                summary,
                "property_type_updates",
                {"id": property_id, "to": fix["property_type"], "evidence": fix.get("property_type_evidence")},
            )
        save_property(property_obj, update_fields, apply)
        if fix.get("latitude") is not None and fix.get("longitude") is not None:
            if update_location(
                property_obj,
                fix["latitude"],
                fix["longitude"],
                fix.get("provider") or "curation_map",
                first_listing_url(property_obj),
                summary,
                apply,
            ):
                add_summary(summary, "location_updates", {"id": property_id, "provider": fix.get("provider")})
            apply_zone_from_point(
                property_obj,
                fix["latitude"],
                fix["longitude"],
                fix.get("provider") or "curation_map",
                "feedback aprobado + fuente mapa/geocoding",
                summary,
                apply,
            )
        elif fix.get("geocode"):
            reason = fix.get("evidence") or "feedback aprobado + direccion curada"
            apply_geocode(property_obj, reason, summary, apply)
            if fix.get("external_geocode"):
                external_geocode_property(property_obj, reason, summary, apply)


def apply_existing_coordinate_zones(summary, apply):
    for property_id, expected_zone in sorted(EXISTING_COORDINATE_ZONES.items()):
        property_obj = Property.objects.select_related("location", "location_intelligence").filter(pk=property_id).first()
        if not property_obj:
            add_summary(summary, "missing_property", property_id)
            continue
        location = location_or_none(property_obj)
        if not location:
            add_summary(summary, "existing_coordinate_missing_location", property_id)
            continue
        match = infer_zone_for_point(location.latitude, location.longitude, max_distance_m=100)
        inferred = safe(match.get("zone"))
        add_summary(
            summary,
            "existing_coordinate_zone_checked",
            {
                "id": property_id,
                "expected": expected_zone,
                "inferred": inferred,
                "method": match.get("method"),
                "distance_m": match.get("distance_m"),
                "provider": location.provider,
            },
        )
        if inferred != expected_zone:
            add_summary(summary, "existing_coordinate_zone_conflict", {"id": property_id, "expected": expected_zone, "inferred": inferred})
            continue
        apply_zone_from_existing_match(
            property_obj,
            location,
            match,
            "pregunta Excel: coordenada existente permite inferir zona",
            summary,
            apply,
        )


def apply_zone_from_existing_match(property_obj, location, match, reason, summary, apply):
    zone = safe(match.get("zone"))
    if not zone:
        add_summary(summary, "zone_skipped", {"id": property_obj.pk, "reason": "sin_match_poligono"})
        return ""
    now = timezone.now()
    territory = infer_territory_for_point(location.latitude, location.longitude)
    evidence = dict(match.get("evidence") or {})
    evidence.update(
        {
            "provider": location.provider,
            "curacion": SCRIPT_TAG,
            "reason": reason,
            "latitude": location.latitude,
            "longitude": location.longitude,
            "zone_method": match.get("method"),
            "zone_distance_m": match.get("distance_m"),
        }
    )
    if apply:
        property_obj.inferred_partido = safe(territory.partido if territory else property_obj.inferred_partido)
        property_obj.inferred_locality = safe(territory.locality if territory else property_obj.inferred_locality)
        property_obj.inferred_zone = zone
        property_obj.inferred_neighborhood = zone
        property_obj.territory_confidence = territory.confidence if territory else 0.85
        property_obj.territory_source_method = territory.source_method if territory else "coordinates_nearest_zone"
        property_obj.territory_needs_review = False
        property_obj.territory_evidence = evidence
        property_obj.territory_inferred_at = now
        property_obj.zone_conflict = False
        property_obj.zone_needs_review = False
        property_obj.zone_inference_evidence = evidence
        property_obj.zone_inferred_at = now
        property_obj.save(
            update_fields=[
                "inferred_partido",
                "inferred_locality",
                "inferred_zone",
                "inferred_neighborhood",
                "territory_confidence",
                "territory_source_method",
                "territory_needs_review",
                "territory_evidence",
                "territory_inferred_at",
                "zone_conflict",
                "zone_needs_review",
                "zone_inference_evidence",
                "zone_inferred_at",
            ]
        )
        PropertyLocationIntelligence.objects.update_or_create(
            property=property_obj,
            defaults={
                "partido_name": property_obj.inferred_partido,
                "locality_name": property_obj.inferred_locality,
                "zone_name": zone,
                "match_method": PropertyLocationIntelligence.MatchMethod.COORDINATES,
                "confidence": property_obj.territory_confidence or 0.85,
            },
        )
    add_summary(summary, "zone_applied", {"id": property_obj.pk, "zone": zone, "provider": location.provider})
    return zone


def record_coordinate_no_match_notes(summary, apply):
    now = timezone.now()
    for property_id, reason in sorted(COORDINATE_NO_MATCH_IDS.items()):
        property_obj = Property.objects.filter(pk=property_id).first()
        if not property_obj:
            add_summary(summary, "missing_property", property_id)
            continue
        evidence = dict(property_obj.zone_inference_evidence or {})
        evidence.setdefault("curation_coordinate_no_match", {})[SCRIPT_TAG] = {
            "reason": reason,
            "reviewed_at": now.isoformat(),
        }
        add_summary(summary, "coordinate_no_match_pending", {"id": property_id, "reason": reason})
        if apply:
            property_obj.zone_inference_evidence = evidence
            property_obj.save(update_fields=["zone_inference_evidence"])


def load_ml_chrome_results():
    if not ML_CHROME_RESULTS_PATH.exists():
        return {}
    try:
        data = json.loads(ML_CHROME_RESULTS_PATH.read_text(encoding="utf-8"))
    except Exception:
        return {}
    return {int(item.get("id")): item for item in data if item.get("id")}


def record_ml_chrome_evidence(summary, chrome_results):
    for property_id in sorted(set(DIRECT_FIXES) & set(chrome_results)):
        item = chrome_results[property_id]
        lines = item.get("addressLines") or []
        add_summary(
            summary,
            "mercadolibre_chrome_checked",
            {
                "id": property_id,
                "url": item.get("url"),
                "account_verification": bool(item.get("accountVerification")),
                "address_lines_sample": lines[:6],
                "coordinate_count": len(item.get("coordMatches") or []),
            },
        )


def canonical_propagated_address(value):
    text = safe(value)
    if not text:
        return ""
    bustamante = re.search(r"\bBustamante\s+(\d{2,5})\b", text, re.I)
    if bustamante:
        return f"Eva Peron {bustamante.group(1)}"
    return canonical_address_alias(text)


def apply_bustamante_propagation(summary, apply):
    qs = (
        Property.objects.select_related("location")
        .filter(status=Property.Status.ACTIVE, is_hidden=False)
        .filter(Q(address__icontains="Bustamante") | Q(detected_address__icontains="Bustamante"))
        .filter(location__isnull=True)
        .exclude(pk__in=DIRECT_FIXES)
        .distinct()
        .order_by("id")
    )
    now = timezone.now()
    for property_obj in qs:
        manual = property_obj.manual_overrides or {}
        if "address" in manual:
            continue
        original = property_obj.address or property_obj.detected_address
        candidate = canonical_propagated_address(original)
        if not candidate or candidate == original or not re.search(r"\d{2,5}", candidate):
            continue
        update_fields = set()
        mark_manual_property_field(property_obj, "address", candidate, now, update_fields, apply)
        save_property(property_obj, update_fields, apply)
        add_summary(summary, "propagated_address_updates", {"id": property_obj.pk, "from": original, "to": candidate, "rule": "Bustamante -> Eva Peron"})
        apply_geocode(property_obj, "propagacion segura Bustamante -> Eva Peron", summary, apply)


FINCAS_PREFIX_RE = re.compile(r"\b(?:argentina\s+)?gba\s+oeste\b", re.I)
FINCAS_INSIDE_RE = re.compile(r"\b(?:hurlingham|villa\s+tesei|william\s+c\.?\s*morris)\b", re.I)
FINCAS_OUTSIDE_RE = re.compile(r"\b(?:castelar|mor[oó]n|merlo|caseros|palomar|bella\s+vista)\b", re.I)


def fincas_raw_address(property_obj):
    for value in (property_obj.address, property_obj.detected_address):
        if value and FINCAS_PREFIX_RE.search(value):
            return value
    return ""


def fincas_prefix_classification(raw):
    if FINCAS_OUTSIDE_RE.search(raw or ""):
        return "fuera_partido"
    if FINCAS_INSIDE_RE.search(raw or ""):
        return "dentro_partido"
    return "indeterminado"


def fincas_prefix_queryset():
    return (
        Property.objects.select_related("location", "location_intelligence")
        .prefetch_related("listings__source")
        .filter(status=Property.Status.ACTIVE, is_hidden=False, listings__source__slug="fincas")
        .filter(
            Q(address__icontains="GBA Oeste")
            | Q(address__icontains="Argentina GBA")
            | Q(detected_address__icontains="GBA Oeste")
            | Q(detected_address__icontains="Argentina GBA")
        )
        .distinct()
        .order_by("id")
    )


def source_confirms_outside(markup, url):
    text = f"{url or ''} {markup or ''}"
    return bool(FINCAS_OUTSIDE_RE.search(text))


def source_confirms_becerra_retired(status_code, markup):
    return status_code in {404, 410} or bool(BECERRA_RETIRED_RE.search(markup or ""))


def source_confirms_zonaprop_retired(status_code, markup):
    if status_code in {404, 410}:
        return True
    return bool(ZONAPROP_OFFLINE_RE.search(markup or ""))


def apply_location_from_data_or_existing(property_obj, data, reason, provider_prefix, summary, apply):
    if data and data.get("latitude") is not None and data.get("longitude") is not None:
        provider = f"{provider_prefix}_map"
        if update_location(
            property_obj,
            data["latitude"],
            data["longitude"],
            provider,
            first_listing_url(property_obj),
            summary,
            apply,
        ):
            add_summary(summary, "location_updates", {"id": property_obj.pk, "provider": provider})
        apply_zone_from_point(property_obj, data["latitude"], data["longitude"], provider, reason, summary, apply)
        return
    current = location_or_none(property_obj)
    if current:
        add_summary(
            summary,
            "existing_location_used",
            {"id": property_obj.pk, "provider": current.provider, "manual": current.manually_corrected},
        )
        apply_zone_from_point(property_obj, current.latitude, current.longitude, current.provider, reason, summary, apply)
        return
    apply_geocode(property_obj, reason, summary, apply)


def is_oscar_related(property_obj):
    listing = first_listing(property_obj)
    if not listing:
        return False
    source_slug = listing.source.slug if listing.source else ""
    agency_name = listing.agency.name if listing.agency else ""
    return source_slug == "oscar-dahbar" or "oscar" in safe(agency_name).lower() or "dahbar" in safe(agency_name).lower()


def canonical_oscar_address(value):
    candidate = canonical_address_alias(value)
    return safe(candidate)


def source_confirms_oscar_outside(property_obj, markup=""):
    listing = first_listing(property_obj)
    source_text = " ".join(
        [
            listing.url if listing else "",
            property_obj.title or "",
            property_obj.address or "",
            property_obj.detected_address or "",
            property_obj.locality or "",
            property_obj.detected_locality or "",
            markup or "",
        ]
    )
    return is_declared_out_of_target(source_text)


def apply_oscar_propagated_addresses(summary, apply):
    now = timezone.now()
    for property_id in sorted(PROPAGATE_IDS):
        property_obj = Property.objects.select_related("location", "location_intelligence").filter(pk=property_id).first()
        if not property_obj:
            add_summary(summary, "missing_property", property_id)
            continue
        if not is_oscar_related(property_obj):
            add_summary(summary, "oscar_propagation_skipped", {"id": property_id, "reason": "no_es_oscar"})
            continue
        if source_confirms_oscar_outside(property_obj):
            add_summary(summary, "oscar_propagation_skipped", {"id": property_id, "reason": "fuera_de_hurlingham_en_fuente"})
            continue
        original = property_obj.address or property_obj.detected_address or ""
        candidate = canonical_oscar_address(original)
        if not candidate or not best_address(property_obj):
            add_summary(summary, "oscar_propagation_skipped", {"id": property_id, "reason": "sin_domicilio_util", "address": original})
            continue
        update_fields = set()
        mark_manual_property_field(property_obj, "address", candidate, now, update_fields, apply)
        save_property(property_obj, update_fields, apply)
        add_summary(
            summary,
            "oscar_propagated_addresses",
            {"id": property_id, "from": original, "to": candidate},
        )
        current = location_or_none(property_obj)
        if current and not current.outside_target:
            apply_zone_from_point(
                property_obj,
                current.latitude,
                current.longitude,
                current.provider,
                "Oscar Dahbar propagado con ubicacion existente",
                summary,
                apply,
            )
        else:
            apply_geocode(property_obj, "Oscar Dahbar propagado por domicilio util", summary, apply)


def apply_oscar_no_address_reparse(summary, apply):
    now = timezone.now()
    for property_id in sorted(NO_ADDRESS_REPARSE_IDS):
        property_obj = Property.objects.select_related("location", "location_intelligence").filter(pk=property_id).first()
        if not property_obj:
            add_summary(summary, "missing_property", property_id)
            continue
        if "address" not in (property_obj.manual_overrides or {}) and (
            is_bad_oscar_address(property_obj.address) or is_bad_oscar_address(property_obj.detected_address)
        ):
            add_summary(
                summary,
                "oscar_bad_address_cleared",
                {"id": property_id, "address": property_obj.address, "detected_address": property_obj.detected_address},
            )
            if apply:
                property_obj.address = ""
                property_obj.detected_address = ""
                property_obj.normalized_address = ""
                property_obj.save(update_fields=["address", "detected_address", "normalized_address"])
        listing = first_listing(property_obj)
        if not listing or not listing.url:
            add_summary(summary, "oscar_no_address_pending", {"id": property_id, "reason": "sin_url"})
            continue
        markup = ""
        parsed = None
        try:
            status_code, markup = fetch_url_text(listing.url, "oscar-dahbar")
        except Exception as exc:
            add_summary(summary, "source_fetch_errors", {"id": property_id, "source": "oscar-dahbar", "error": repr(exc)})
            status_code = None
        if status_code and status_code >= 400:
            add_summary(summary, "source_fetch_errors", {"id": property_id, "source": "oscar-dahbar", "http_status": status_code})
            continue
        if source_confirms_oscar_outside(property_obj, markup):
            mark_removed(property_obj, "Oscar Dahbar fuera del partido confirmado por fuente", summary, apply)
            continue
        if markup:
            try:
                parsed = parse_with_scraper(OscarDahbarScraper, listing.url, markup)
            except Exception as exc:
                add_summary(summary, "source_parse_errors", {"id": property_id, "source": "oscar-dahbar", "error": repr(exc)})
        if parsed and parsed.get("address") and not best_address(property_obj):
            update_fields = set()
            mark_manual_property_field(property_obj, "address", parsed["address"], now, update_fields, apply)
            save_property(property_obj, update_fields, apply)
            add_summary(summary, "oscar_source_address_updates", {"id": property_id, "to": parsed["address"]})
        coordinates = []
        if parsed and parsed.get("latitude") is not None and parsed.get("longitude") is not None:
            coordinates = [{"latitude": parsed["latitude"], "longitude": parsed["longitude"], "method": "oscar_parser"}]
        elif markup:
            coordinates = extract_map_coordinates(markup, require_target_bounds=True)
        if coordinates:
            coordinate = coordinates[0]
            if update_location(
                property_obj,
                coordinate["latitude"],
                coordinate["longitude"],
                "oscar_dahbar_map",
                listing.url,
                summary,
                apply,
            ):
                add_summary(summary, "sweep_map_locations", {"id": property_id, "source": "oscar-dahbar", "method": coordinate.get("method")})
            apply_zone_from_point(
                property_obj,
                coordinate["latitude"],
                coordinate["longitude"],
                "oscar_dahbar_map",
                "Oscar Dahbar sin domicilio con coordenada de mapa",
                summary,
                apply,
            )
            continue
        current = location_or_none(property_obj)
        if current:
            apply_zone_from_point(
                property_obj,
                current.latitude,
                current.longitude,
                current.provider,
                "Oscar Dahbar sin domicilio con coordenada existente",
                summary,
                apply,
            )
        else:
            add_summary(summary, "oscar_no_address_pending", {"id": property_id, "reason": "sin_mapa_domicilio_util"})


def confirm_oscar_remove_candidates(summary, apply):
    for property_id in sorted(CONFIRM_REMOVE_IDS):
        property_obj = Property.objects.prefetch_related("listings__source").filter(pk=property_id).first()
        if not property_obj:
            add_summary(summary, "missing_property", property_id)
            continue
        listing = first_listing(property_obj)
        markup = ""
        if listing and listing.url:
            try:
                status_code, markup = fetch_url_text(listing.url, "oscar-dahbar")
                add_summary(summary, "oscar_remove_candidate_checked", {"id": property_id, "http_status": status_code})
            except Exception as exc:
                add_summary(summary, "source_fetch_errors", {"id": property_id, "source": "oscar-dahbar", "error": repr(exc)})
        if source_confirms_oscar_outside(property_obj, markup):
            mark_removed(property_obj, "Oscar Dahbar fuera del partido confirmado por URL/titulo/fuente", summary, apply)
        else:
            add_summary(summary, "remove_skipped", {"id": property_id, "reason": "sin_confirmacion_fuera_partido"})


def apply_fincas_prefix_repairs(summary, apply):
    now = timezone.now()
    candidates = [
        prop
        for prop in fincas_prefix_queryset()
        if prop.pk in FEEDBACK_IDS or residual_reason(prop)
    ]
    add_summary(summary, "fincas_prefix_candidates", {"count": len(candidates), "ids": [prop.pk for prop in candidates]})
    for property_obj in candidates:
        raw = fincas_raw_address(property_obj)
        classification = fincas_prefix_classification(raw)
        add_summary(
            summary,
            "fincas_prefix_classified",
            {"id": property_obj.pk, "class": classification, "raw": raw},
        )
        if property_obj.pk in DIRECT_FIXES:
            add_summary(summary, "fincas_prefix_skipped", {"id": property_obj.pk, "reason": "aprobado_directo"})
            continue
        if "address" in (property_obj.manual_overrides or {}):
            add_summary(summary, "fincas_prefix_skipped", {"id": property_obj.pk, "reason": "address_manual"})
            continue
        listing = first_listing(property_obj)
        markup = ""
        parsed = None
        if listing and listing.url:
            try:
                status_code, markup = fetch_url_text(listing.url, "fincas")
            except Exception as exc:
                add_summary(summary, "source_fetch_errors", {"id": property_obj.pk, "source": "fincas", "error": repr(exc)})
                status_code = None
            if status_code and status_code >= 400:
                add_summary(summary, "source_fetch_errors", {"id": property_obj.pk, "source": "fincas", "http_status": status_code})
            elif markup:
                try:
                    parsed = parse_with_scraper(FincasScraper, listing.url, markup)
                except Exception as exc:
                    add_summary(summary, "source_parse_errors", {"id": property_obj.pk, "source": "fincas", "error": repr(exc)})
        if classification == "fuera_partido":
            if source_confirms_outside(markup, listing.url if listing else ""):
                mark_removed(property_obj, "Fincas/Haurie fuera del partido confirmado por URL o HTML", summary, apply)
            else:
                add_summary(summary, "fincas_prefix_skipped", {"id": property_obj.pk, "reason": "fuera_sin_confirmacion"})
            continue
        if classification != "dentro_partido":
            add_summary(summary, "fincas_prefix_skipped", {"id": property_obj.pk, "reason": "indeterminado"})
            continue
        candidate = safe((parsed or {}).get("address")) or normalize_argencasas_address(raw)
        if not candidate:
            add_summary(summary, "fincas_prefix_skipped", {"id": property_obj.pk, "reason": "sin_direccion_limpia"})
            continue
        update_fields = set()
        mark_manual_property_field(property_obj, "address", candidate, now, update_fields, apply)
        save_property(property_obj, update_fields, apply)
        add_summary(summary, "fincas_address_updates", {"id": property_obj.pk, "from": raw, "to": candidate})
        apply_location_from_data_or_existing(
            property_obj,
            parsed,
            "barrido Fincas/Haurie con breadcrumb territorial limpio",
            "fincas",
            summary,
            apply,
        )


def source_pending_queryset(source_slug):
    return (
        Property.objects.select_related("location", "location_intelligence")
        .prefetch_related("listings__source")
        .filter(listings__source__slug=source_slug, status=Property.Status.ACTIVE, is_hidden=False)
        .filter(no_zone_query() | Q(location__isnull=True) | Q(address="") | Q(address__isnull=True))
        .distinct()
        .order_by("id")
    )


def prioritized_source_candidates(source_slug, priority_ids=None):
    priority_ids = set(priority_ids or set())
    candidates = list(source_pending_queryset(source_slug))
    active_candidates = [prop for prop in candidates if prop.pk not in NO_ACTION_IDS and not curation_no_action(prop)]
    priority = [prop for prop in active_candidates if prop.pk in priority_ids]
    rest = [prop for prop in active_candidates if prop.pk not in priority_ids]
    limit = SWEEP_EXTRA_LIMITS.get(source_slug, 0)
    selected = priority + rest[:limit]
    skipped = rest[limit:]
    return candidates, selected, skipped


def sweep_argencasas(summary, apply):
    candidates, selected, skipped = prioritized_source_candidates("argencasas")
    add_summary(
        summary,
        "argencasas_pending_checked",
        {"count": len(candidates), "selected_ids": [prop.pk for prop in selected], "skipped_count": len(skipped)},
    )
    for property_obj in selected:
        if property_obj.pk in DIRECT_FIXES or property_obj.pk in CONFIRM_REMOVE_IDS:
            continue
        listing = first_listing(property_obj)
        if not listing or not listing.url:
            continue
        try:
            status_code, markup = fetch_url_text(listing.url, "argencasas")
        except Exception as exc:
            add_summary(summary, "source_fetch_errors", {"id": property_obj.pk, "source": "argencasas", "error": repr(exc)})
            continue
        if has_removed_marker(status_code, markup):
            mark_removed(property_obj, "Argencasas 404 + propiedad retirada del sistema en barrido", summary, apply)
            continue
        coordinates = extract_map_coordinates(markup, require_target_bounds=True)
        if coordinates:
            coordinate = coordinates[0]
            if update_location(property_obj, coordinate["latitude"], coordinate["longitude"], "argencasas_map", listing.url, summary, apply):
                add_summary(summary, "sweep_map_locations", {"id": property_obj.pk, "source": "argencasas", "method": coordinate["method"]})
            apply_zone_from_point(
                property_obj,
                coordinate["latitude"],
                coordinate["longitude"],
                "argencasas_map",
                "barrido Argencasas con coordenada fuente",
                summary,
                apply,
            )


def sweep_becerra(summary, apply):
    candidates, selected, skipped = prioritized_source_candidates("becerra", FEEDBACK_IDS)
    add_summary(
        summary,
        "becerra_pending_checked",
        {
            "count": len(candidates),
            "selected_ids": [prop.pk for prop in selected],
            "skipped_count": len(skipped),
            "skipped_ids": [prop.pk for prop in skipped],
        },
    )
    for property_obj in selected:
        if property_obj.pk in DIRECT_FIXES:
            continue
        listing = first_listing(property_obj)
        if not listing or not listing.url:
            continue
        try:
            status_code, markup = fetch_url_text(listing.url, "becerra")
        except Exception as exc:
            add_summary(summary, "source_fetch_errors", {"id": property_obj.pk, "source": "becerra", "error": repr(exc)})
            continue
        if source_confirms_becerra_retired(status_code, markup):
            mark_removed(property_obj, "Becerra pagina de error/retirada confirmada por fuente", summary, apply)
            continue
        if status_code >= 400:
            add_summary(summary, "source_fetch_errors", {"id": property_obj.pk, "source": "becerra", "http_status": status_code})
            continue
        try:
            data = parse_with_scraper(BecerraScraper, listing.url, markup)
        except Exception as exc:
            add_summary(summary, "source_parse_errors", {"id": property_obj.pk, "source": "becerra", "error": repr(exc)})
            continue
        if data.get("latitude") is not None and data.get("longitude") is not None:
            if update_location(property_obj, data["latitude"], data["longitude"], "becerra_map", listing.url, summary, apply):
                add_summary(summary, "sweep_map_locations", {"id": property_obj.pk, "source": "becerra"})
            apply_zone_from_point(
                property_obj,
                data["latitude"],
                data["longitude"],
                "becerra_map",
                "barrido Becerra con coordenada fuente",
                summary,
                apply,
            )


def sweep_zonaprop(summary, apply):
    candidates, selected, skipped = prioritized_source_candidates("zonaprop", FEEDBACK_IDS | MAP_ONLY_IDS)
    add_summary(
        summary,
        "zonaprop_pending_checked",
        {
            "count": len(candidates),
            "selected_ids": [prop.pk for prop in selected],
            "skipped_count": len(skipped),
            "skipped_ids": [prop.pk for prop in skipped],
        },
    )
    for property_obj in selected:
        if property_obj.pk in APPROVED_IDS or property_obj.pk in REJECTED_REMOVE_IDS:
            continue
        listing = first_listing(property_obj)
        if not listing or not listing.url:
            continue
        try:
            status_code, markup = fetch_url_text(listing.url, "zonaprop")
        except Exception as exc:
            add_summary(summary, "source_fetch_errors", {"id": property_obj.pk, "source": "zonaprop", "error": repr(exc)})
            continue
        if source_confirms_zonaprop_retired(status_code, markup):
            mark_removed(property_obj, "Zonaprop aviso retirado/410/404 confirmado por fuente", summary, apply)
            continue
        if status_code >= 400:
            add_summary(summary, "source_fetch_errors", {"id": property_obj.pk, "source": "zonaprop", "http_status": status_code})
            continue
        coordinate = source_map_coordinates("zonaprop", listing.url, markup, summary, property_obj.pk)
        data = coordinate.get("data") or {}
        if data.get("address"):
            add_summary(
                summary,
                "zonaprop_source_address_skipped",
                {
                    "id": property_obj.pk,
                    "address": safe(data["address"]),
                    "reason": "barrido iter5 usa mapa para zona; no pisa domicilio desde texto Zonaprop no aprobado",
                },
            )
        if coordinate.get("latitude") is not None and coordinate.get("longitude") is not None:
            if update_location(property_obj, coordinate["latitude"], coordinate["longitude"], "zonaprop_map", listing.url, summary, apply):
                add_summary(summary, "sweep_map_locations", {"id": property_obj.pk, "source": "zonaprop", "method": coordinate.get("method")})
            apply_zone_from_point(
                property_obj,
                coordinate["latitude"],
                coordinate["longitude"],
                "zonaprop_map",
                "barrido Zonaprop con coordenada fuente",
                summary,
                apply,
            )
        else:
            add_summary(summary, "zonaprop_map_miss", property_obj.pk)


def sweep_faella(summary, apply):
    candidates, selected, skipped = prioritized_source_candidates("faella", FEEDBACK_IDS)
    add_summary(
        summary,
        "faella_pending_checked",
        {
            "count": len(candidates),
            "selected_ids": [prop.pk for prop in selected],
            "skipped_count": len(skipped),
            "skipped_ids": [prop.pk for prop in skipped],
        },
    )
    for property_obj in selected:
        if property_obj.pk in DIRECT_FIXES:
            continue
        listing = first_listing(property_obj)
        if not listing or not listing.url:
            continue
        try:
            status_code, markup = fetch_url_text(listing.url, "faella")
        except Exception as exc:
            add_summary(summary, "source_fetch_errors", {"id": property_obj.pk, "source": "faella", "error": repr(exc)})
            continue
        if status_code in {401, 403, 429}:
            add_summary(summary, "mercadolibre_blocked", {"id": property_obj.pk, "http_status": status_code})
            continue
        if status_code >= 400:
            add_summary(summary, "source_fetch_errors", {"id": property_obj.pk, "source": "faella", "http_status": status_code})
            continue
        coordinates = extract_map_coordinates(markup, require_target_bounds=True)
        if not coordinates:
            add_summary(summary, "faella_map_miss", property_obj.pk)
            continue
        coordinate = coordinates[0]
        if update_location(property_obj, coordinate["latitude"], coordinate["longitude"], "faella_meli_map", listing.url, summary, apply):
            add_summary(summary, "sweep_map_locations", {"id": property_obj.pk, "source": "faella", "method": coordinate.get("method")})
        apply_zone_from_point(
            property_obj,
            coordinate["latitude"],
            coordinate["longitude"],
            "faella_meli_map",
            "barrido Faella/MercadoLibre con coordenada publica",
            summary,
            apply,
        )


def apply_map_only_feedback(summary, apply, allowed_ids):
    for property_id in sorted(MAP_ONLY_IDS & set(allowed_ids)):
        property_obj = Property.objects.select_related("location", "location_intelligence").filter(pk=property_id).first()
        if not property_obj:
            add_summary(summary, "missing_property", property_id)
            continue
        applied = apply_source_map_if_available(
            property_obj,
            "feedback aprobado: mapa fuente sin inventar domicilio",
            summary,
            apply,
        )
        if not applied:
            add_summary(summary, "map_only_pending", property_id)


def apply_coraceros_alias_similars(summary, apply):
    qs = (
        Property.objects.select_related("location", "location_intelligence")
        .prefetch_related("listings__source")
        .filter(status=Property.Status.ACTIVE, is_hidden=False)
        .filter(
            Q(address__icontains="Carocero")
            | Q(address__icontains="Coracero")
            | Q(detected_address__icontains="Carocero")
            | Q(detected_address__icontains="Coracero")
        )
        .exclude(pk__in=set(DIRECT_FIXES) | NO_ACTION_IDS | REJECTED_REMOVE_IDS)
        .distinct()
        .order_by("id")
    )
    now = timezone.now()
    candidates = []
    for property_obj in qs:
        if "address" in (property_obj.manual_overrides or {}):
            continue
        original = property_obj.address or property_obj.detected_address or ""
        candidate = canonical_address_alias(original)
        if not candidate or candidate == original or "Coraceros" not in candidate:
            continue
        candidates.append({"id": property_obj.pk, "from": original, "to": candidate})
        update_fields = set()
        mark_manual_property_field(property_obj, "address", candidate, now, update_fields, apply)
        save_property(property_obj, update_fields, apply)
        apply_geocode(property_obj, "alias seguro Caroceros/Coracero", summary, apply)
    if candidates:
        add_summary(summary, "coraceros_alias_similars", candidates)


def sweep_guarnieri(summary, apply):
    candidates, selected, skipped = prioritized_source_candidates("guarnieri")
    add_summary(
        summary,
        "guarnieri_pending_checked",
        {"count": len(candidates), "selected_ids": [prop.pk for prop in selected], "skipped_count": len(skipped)},
    )
    for property_obj in selected:
        if property_obj.pk in DIRECT_FIXES:
            continue
        listing = first_listing(property_obj)
        if not listing or not listing.url:
            continue
        try:
            status_code, markup = fetch_url_text(listing.url, "guarnieri")
        except Exception as exc:
            add_summary(summary, "source_fetch_errors", {"id": property_obj.pk, "source": "guarnieri", "error": repr(exc)})
            continue
        if status_code >= 400:
            add_summary(summary, "source_fetch_errors", {"id": property_obj.pk, "source": "guarnieri", "http_status": status_code})
            continue
        coordinates = extract_map_coordinates(markup, require_target_bounds=True)
        if coordinates:
            coordinate = coordinates[0]
            if update_location(property_obj, coordinate["latitude"], coordinate["longitude"], "guarnieri_map", listing.url, summary, apply):
                add_summary(summary, "sweep_map_locations", {"id": property_obj.pk, "source": "guarnieri", "method": coordinate["method"]})
            apply_zone_from_point(
                property_obj,
                coordinate["latitude"],
                coordinate["longitude"],
                "guarnieri_map",
                "barrido Guarnieri con coordenada fuente",
                summary,
                apply,
            )


def source_names(property_obj):
    names = []
    for listing in property_obj.listings.select_related("source").all():
        if listing.source.name not in names:
            names.append(listing.source.name)
    return ", ".join(names)


def agency_names(property_obj):
    names = []
    for listing in property_obj.listings.select_related("agency").all():
        if listing.agency and listing.agency.name not in names:
            names.append(listing.agency.name)
    return ", ".join(names)


def status_label(status):
    return dict(Property.Status.choices).get(status, status)


def residual_queryset():
    rows = (
        Property.objects.select_related("location", "location_intelligence")
        .prefetch_related("listings__source", "listings__agency")
        .filter(
            no_zone_query()
            | Q(address="")
            | Q(address__isnull=True)
            | Q(address__icontains="dormitorios y parque")
            | Q(detected_address__icontains="dormitorios y parque")
            | Q(address__icontains="Amplio con Parrilla Hurlingham")
            | Q(detected_address__icontains="Amplio con Parrilla Hurlingham")
            | Q(address__icontains="Mensaje al anunciante")
            | Q(detected_address__icontains="Mensaje al anunciante")
            | Q(address__icontains="Dormitorios Y Patio Con Parrilla")
            | Q(detected_address__icontains="Dormitorios Y Patio Con Parrilla")
        )
        .filter(is_hidden=False)
        .exclude(status=Property.Status.REMOVED)
        .distinct()
        .order_by("id")
    )
    return [property_obj for property_obj in rows if not curation_no_action(property_obj)]


def oscar_residual_reason(property_obj):
    if not is_oscar_related(property_obj):
        return residual_reason(property_obj)
    location = location_or_none(property_obj)
    address = best_address(property_obj)
    folded_address = safe(address).lower()
    if not address and not location:
        return "sin mapa/domicilio util"
    if not address:
        return "sin direccion util"
    if re.search(r"\b(?:y|esquina|entre|e/)\b", folded_address) and not location:
        return "interseccion sin coordenada"
    if not location:
        return "geocoding sin resultado"
    if location.outside_target:
        return "coordenada fuera de Hurlingham"
    if not property_obj.inferred_zone:
        return "coordenadas sin zona"
    return residual_reason(property_obj)


def property_row(property_obj, reviewed_notes):
    location = location_or_none(property_obj)
    intel = intel_or_none(property_obj)
    listing = first_listing(property_obj)
    url = listing.url if listing else ""
    evidence = {
        "location_precision": location.precision if location else "",
        "location_query": location.query if location else "",
        "manual_overrides": property_obj.manual_overrides or {},
        "source_locality": property_obj.locality,
        "source_zone": property_obj.inferred_zone,
        "territory_evidence": property_obj.territory_evidence or {},
    }
    if property_obj.pk in reviewed_notes:
        evidence["feedback_20260628"] = reviewed_notes[property_obj.pk] or "revisado manualmente"
    return [
        property_obj.pk,
        property_obj.title,
        (property_obj.description or "")[:500],
        property_obj.address or "",
        property_obj.detected_address or "",
        "",
        property_obj.normalized_address or "",
        property_obj.locality or "",
        property_obj.detected_locality or "",
        property_obj.neighborhood or "",
        property_obj.detected_neighborhood or "",
        property_obj.inferred_zone or "",
        "",
        intel.zone_name if intel else "",
        "BAJA",
        oscar_residual_reason(property_obj),
        "DUDOSO",
        location.latitude if location else None,
        location.longitude if location else None,
        location.provider if location else "",
        source_names(property_obj),
        agency_names(property_obj),
        url,
        "" if url else "Sin Listing asociado en la base",
        status_label(property_obj.status),
        property_obj.get_property_type_display(),
        property_obj.operation,
        property_obj.currency or "",
        float(property_obj.price) if property_obj.price is not None else None,
        property_obj.rooms,
        property_obj.bedrooms,
        float(property_obj.bathrooms) if property_obj.bathrooms is not None else None,
        float(property_obj.covered_area) if property_obj.covered_area is not None else None,
        float(property_obj.total_area) if property_obj.total_area is not None else None,
        property_obj.last_seen_at.isoformat() if property_obj.last_seen_at else "",
        oscar_residual_reason(property_obj),
        json.dumps(evidence, ensure_ascii=False, default=str),
        "",
        "",
    ]


def generate_excel(reviewed_notes):
    wb = Workbook()
    ws = wb.active
    ws.title = "Pendientes"
    ws.append(HEADERS)
    header_fill = PatternFill("solid", fgColor="1F4E78")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center")
    rows = list(residual_queryset())
    for property_obj in rows:
        ws.append(property_row(property_obj, reviewed_notes))

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    table = Table(displayName="PendientesFeedback20260628Iter6", ref=ws.dimensions)
    table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
    ws.add_table(table)
    decision_col = HEADERS.index("decision_manual") + 1
    validation = DataValidation(type="list", formula1='"APROBADO,RECHAZADO"', allow_blank=True)
    ws.add_data_validation(validation)
    if ws.max_row >= 2:
        validation.add(f"{get_column_letter(decision_col)}2:{get_column_letter(decision_col)}{ws.max_row}")

    widths = {
        "A": 12,
        "B": 42,
        "C": 52,
        "D": 28,
        "F": 24,
        "M": 22,
        "N": 22,
        "W": 58,
        "AJ": 24,
        "AK": 70,
        "AL": 18,
        "AM": 36,
    }
    for column in range(1, len(HEADERS) + 1):
        letter = get_column_letter(column)
        ws.column_dimensions[letter].width = widths.get(letter, 16)
    price_col = HEADERS.index("precio") + 1
    for row in range(2, ws.max_row + 1):
        ws.cell(row, price_col).number_format = "#,##0"
        ws.cell(row, HEADERS.index("descripcion_resumen") + 1).alignment = Alignment(wrap_text=True, vertical="top")
        ws.cell(row, HEADERS.index("evidencia") + 1).alignment = Alignment(wrap_text=True, vertical="top")

    summary_ws = wb.create_sheet("Resumen")
    summary_ws.append(["metrica", "valor"])
    summary_ws.append(["pendientes", len(rows)])
    summary_ws.append(["generado_por", SCRIPT_TAG])
    wb.save(OUTPUT_XLSX)
    return {"path": str(OUTPUT_XLSX), "pending_rows": len(rows), "columns": HEADERS}


def validate_excel():
    wb = load_workbook(OUTPUT_XLSX, data_only=True)
    ws = wb["Pendientes"]
    headers = [cell.value for cell in ws[1]]
    price_col = headers.index("precio") + 1
    decision_col = headers.index("decision_manual") + 1
    notes_col = headers.index("notas_manual") + 1
    price_types = Counter(type(ws.cell(row, price_col).value).__name__ for row in range(2, ws.max_row + 1))
    return {
        "sheets": wb.sheetnames,
        "rows": ws.max_row - 1,
        "auto_filter": ws.auto_filter.ref,
        "tables": list(ws.tables.keys()),
        "freeze_panes": ws.freeze_panes,
        "price_types": dict(price_types),
        "nonblank_decisions": sum(1 for row in range(2, ws.max_row + 1) if ws.cell(row, decision_col).value),
        "nonblank_notes": sum(1 for row in range(2, ws.max_row + 1) if ws.cell(row, notes_col).value),
        "validations": [
            {"type": dv.type, "formula1": dv.formula1, "sqref": str(dv.sqref)}
            for dv in ws.data_validations.dataValidation
        ],
        "missing_columns": [column for column in HEADERS if column not in headers],
    }


def cleanup_tmp(summary, apply):
    targets = []
    targets.extend(ROOT.glob("tmp/*.err.log"))
    targets.extend(ROOT.glob("tmp/*.out.log"))
    targets.append(ROOT / "tmp" / "curacion_zonas_pendientes_2026-06-19_filtros.xlsx.inspect.ndjson")
    pycache = ROOT / "tmp" / "__pycache__"
    removed = []
    for target in targets:
        if not target.exists():
            continue
        removed.append(str(target))
        if apply:
            target.unlink()
    if pycache.exists():
        removed.append(str(pycache))
        if apply:
            shutil.rmtree(pycache)
    summary["tmp_cleanup"] = removed


def sample(ids):
    out = []
    for property_obj in (
        Property.objects.select_related("location", "location_intelligence")
        .prefetch_related("listings__source")
        .filter(pk__in=ids)
        .order_by("id")
    ):
        location = location_or_none(property_obj)
        intel = intel_or_none(property_obj)
        out.append(
            {
                "id": property_obj.pk,
                "address": property_obj.address,
                "locality": property_obj.locality,
                "zone": property_obj.inferred_zone,
                "intel_zone": intel.zone_name if intel else "",
                "status": property_obj.status,
                "is_hidden": property_obj.is_hidden,
                "reviewed_at": property_obj.reviewed_at,
                "location": None
                if not location
                else {
                    "lat": location.latitude,
                    "lon": location.longitude,
                    "provider": location.provider,
                    "manual": location.manually_corrected,
                },
                "manual_overrides": property_obj.manual_overrides or {},
                "zone_evidence": property_obj.zone_inference_evidence or {},
                "corrected": bool(property_obj.data_manually_corrected_at),
                "active_listings": list(property_obj.listings.filter(active=True).values_list("id", flat=True)),
            }
        )
    return out


def apply_feedback(apply):
    summary = {
        "mode": "apply" if apply else "dry_run",
        "input_excel": str(INPUT_XLSX),
        "output_excel": str(OUTPUT_XLSX) if apply else "",
    }
    rows, decisions, notes = load_feedback()
    summary["feedback_decisions"] = dict(decisions)
    summary["approved_found"] = sorted(
        pid for pid in APPROVED_IDS if safe(rows.get(pid, {}).get("decision_manual")).upper() == "APROBADO"
    )
    summary["rejected_found"] = sorted(
        pid for pid in REJECTED_REMOVE_IDS if safe(rows.get(pid, {}).get("decision_manual")).upper() == "RECHAZADO"
    )
    summary["unexpected_approved"] = sorted(
        pid
        for pid, row in rows.items()
        if safe(row.get("decision_manual")).upper() == "APROBADO" and pid not in APPROVED_IDS
    )
    summary["unexpected_rejected"] = sorted(
        pid
        for pid, row in rows.items()
        if safe(row.get("decision_manual")).upper() == "RECHAZADO" and pid not in REJECTED_REMOVE_IDS
    )
    summary["ignored_out_of_scope_decisions"] = {
        "approved": summary["unexpected_approved"],
        "rejected": summary["unexpected_rejected"],
    }
    summary["missing_expected_approved"] = sorted(APPROVED_IDS - set(summary["approved_found"]))
    summary["missing_expected_rejected"] = sorted(REJECTED_REMOVE_IDS - set(summary["rejected_found"]))
    summary["question_notes_found"] = {
        pid: note
        for pid, note in notes.items()
        if pid in ({66} | APPROVED_IDS)
    }
    summary["before"] = current_counts()

    if args_apply_blocked := (
        apply
        and (
            summary["missing_expected_approved"]
            or summary["missing_expected_rejected"]
        )
    ):
        summary["apply_blocked"] = True
        return summary, notes

    if apply:
        if not BACKUP_PATH.exists():
            shutil.copy2(DB_PATH, BACKUP_PATH)
            summary["backup_created"] = True
        else:
            summary["backup_created"] = False
        summary["backup"] = str(BACKUP_PATH)
    else:
        summary["backup"] = ""

    chrome_results = load_ml_chrome_results()
    summary["mercadolibre_chrome_results_path"] = str(ML_CHROME_RESULTS_PATH)
    summary["mercadolibre_chrome_result_ids"] = sorted(chrome_results)
    record_ml_chrome_evidence(summary, chrome_results)

    for property_id in sorted(NO_ACTION_IDS & set(summary["approved_found"])):
        property_obj = Property.objects.filter(pk=property_id).first()
        if property_obj:
            mark_reviewed_no_action(
                property_obj,
                notes.get(property_id, "marcado como revisado sin datos suficientes"),
                summary,
                apply,
            )

    chrome_visible_ids = {property_id for property_id in chrome_results if property_id in DIRECT_FIXES}
    allowed_direct_ids = set(summary["approved_found"]) | chrome_visible_ids
    add_summary(summary, "direct_fix_allowed_ids", sorted(allowed_direct_ids))
    apply_direct_fixes(summary, apply, allowed_direct_ids)
    apply_existing_coordinate_zones(summary, apply)
    record_coordinate_no_match_notes(summary, apply)

    summary["after"] = current_counts()
    return summary, notes


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--dry-run", action="store_true")
    parser.add_argument("--apply", action="store_true")
    parser.add_argument("--generate-excel", action="store_true")
    parser.add_argument("--cleanup", action="store_true")
    parser.add_argument("--json-out")
    args = parser.parse_args()
    if args.apply and args.dry_run:
        parser.error("--apply y --dry-run son mutuamente excluyentes")

    summary, reviewed_notes = apply_feedback(args.apply)
    if args.generate_excel:
        summary["excel"] = generate_excel(reviewed_notes)
        summary["excel_validation"] = validate_excel()
    else:
        summary["excel"] = None
        summary["excel_validation"] = None
    if args.cleanup:
        cleanup_tmp(summary, args.apply)
    summary["sample"] = sample(SAMPLE_IDS)
    output = json.dumps(summary, ensure_ascii=False, indent=2, default=str)
    print(output)
    if args.json_out:
        Path(args.json_out).write_text(output, encoding="utf-8")


if __name__ == "__main__":
    main()
