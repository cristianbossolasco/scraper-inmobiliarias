import argparse
import json
import os
import re
import shutil
import sys
from collections import Counter, defaultdict
from pathlib import Path

import requests
from django.db.models import Q
from django.utils import timezone
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))
os.environ.setdefault("DJANGO_SETTINGS_MODULE", "config.settings")

import django

django.setup()

from properties.models import Listing, Property, PropertyLocation, PropertyLocationIntelligence
from properties.scrapers.parsing import clean_text, extract_map_coordinates
from properties.services.geocoding import Geocoder, best_address
from properties.services.normalization import normalize_address, normalize_whitespace
from properties.services.territory_hierarchy import infer_territory_for_point


INPUT_XLSX = ROOT / "tmp" / "curacion_zonas_pendientes_2026-06-20_mapas.xlsx"
OUTPUT_XLSX = ROOT / "tmp" / "curacion_zonas_pendientes_2026-06-26_feedback.xlsx"
BACKUP_PATH = ROOT / "tmp" / "db.sqlite3.backup_curacion_feedback_20260626"
DB_PATH = ROOT / "db.sqlite3"
SCRIPT_TAG = "curacion_feedback_20260626"

APPROVED_IDS = {3217, 3265, 3268, 3301, 3330, 3349, 3402, 3503, 5309, 5312, 5836, 5837, 5845, 6002}
REJECTED_IDS = {3578, 3333}
NO_DATA_IDS = {3301, 3330, 3349, 3402, 3503}
EXPLICIT_REMOVE_IDS = {3578, 3333, 5309, 5312}
MANUAL_FIXES = {
    3217: {
        "address": "Río Colorado 2100",
        "locality": "William C. Morris",
        "latitude": -34.588521830213,
        "longitude": -58.64304371568,
        "provider": "fincas_map",
        "reason": "feedback aprobado + coordenada embebida Argencasas/Fincas",
    },
    3265: {
        "address": "Diego de Carvajal al 800",
        "locality": "Hurlingham",
        "latitude": -34.6001003,
        "longitude": -58.6345574,
        "provider": "becerra_map",
        "force_needs_review": True,
        "reason": "feedback aprobado + mapa Becerra",
    },
    3268: {
        "address": "Nilda Figueira al 1400",
        "locality": "Hurlingham",
        "latitude": -34.5869687,
        "longitude": -58.6358693,
        "provider": "becerra_map",
        "force_needs_review": True,
        "reason": "feedback aprobado + mapa Becerra",
    },
    5836: {
        "latitude": -34.58066609762967,
        "longitude": -58.64246672233726,
        "provider": "oscar_dahbar_map",
        "reason": "feedback aprobado + mapa Oscar Dahbar",
    },
    5837: {
        "latitude": -34.59363252862368,
        "longitude": -58.645590860364834,
        "provider": "oscar_dahbar_map",
        "reason": "feedback aprobado + mapa Oscar Dahbar",
    },
    5845: {
        "address": "Concepción Arenal 2341",
        "locality": "Hurlingham",
        "provider": "feedback_geocode",
        "reason": "feedback aprobado Faella/MercadoLibre; HTML/API pública bloqueada",
        "geocode": True,
    },
    6002: {
        "address": "Altolaguirre 2435",
        "locality": "Hurlingham",
        "provider": "feedback_geocode",
        "reason": "feedback aprobado Faella/MercadoLibre; HTML/API pública bloqueada",
        "geocode": True,
    },
}

OUTSIDE_RE = re.compile(
    r"\b(?:bella\s+vista|castelar|mor[oó]n|caseros|san\s+miguel|haedo|ituzaing[oó]|ramos\s+mej[ií]a)\b",
    re.I,
)

HEADERS = [
    "id_propiedad",
    "titulo",
    "descripcion_resumen",
    "domicilio_actual",
    "direccion_detectada",
    "direccion_sugerida",
    "direccion_normalizada",
    "localidad_actual",
    "localidad_detectada",
    "barrio_actual",
    "barrio_detectado",
    "zona_actual",
    "zona_sugerida",
    "zona_inteligencia",
    "confianza",
    "motivo",
    "estado",
    "latitud",
    "longitud",
    "fuente_coordenadas",
    "fuente",
    "inmobiliaria",
    "url_publicacion",
    "motivo_url",
    "estado_publicacion",
    "tipo_propiedad",
    "operacion",
    "moneda",
    "precio",
    "ambientes",
    "dormitorios",
    "banos",
    "superficie_cubierta",
    "superficie_total",
    "fecha_ultima_vista",
    "motivo_pendiente",
    "evidencia",
    "decision_manual",
    "notas_manual",
]


def safe(value):
    return normalize_whitespace(str(value)) if value not in (None, "") else ""


def no_zone_query():
    return (
        (Q(inferred_zone__isnull=True) | Q(inferred_zone=""))
        & (
            Q(location_intelligence__isnull=True)
            | Q(location_intelligence__zone_name__isnull=True)
            | Q(location_intelligence__zone_name="")
        )
    )


def location_or_none(property_obj):
    return getattr(property_obj, "location", None)


def intel_or_none(property_obj):
    return getattr(property_obj, "location_intelligence", None)


def residual_reason(property_obj):
    reasons = []
    intel = intel_or_none(property_obj)
    if not property_obj.inferred_zone and not (intel.zone_name if intel else ""):
        reasons.append("sin zona")
    if not best_address(property_obj):
        reasons.append("sin direccion util")
    return ", ".join(reasons)


def current_counts():
    props = list(Property.objects.select_related("location", "location_intelligence").prefetch_related("listings__source"))
    by_source = defaultdict(lambda: Counter())
    for property_obj in props:
        if property_obj.is_hidden or property_obj.status == Property.Status.REMOVED:
            continue
        reason = residual_reason(property_obj)
        if not reason:
            continue
        sources = {listing.source.slug for listing in property_obj.listings.all()} or {"sin_listing"}
        for source in sources:
            by_source[source]["pendientes"] += 1
            if "sin zona" in reason:
                by_source[source]["sin_zona"] += 1
            if "sin direccion" in reason:
                by_source[source]["sin_direccion"] += 1
    return {
        "sin_zona_operativa": Property.objects.filter(no_zone_query(), is_hidden=False).exclude(status=Property.Status.REMOVED).count(),
        "sin_direccion_util": sum(1 for prop in props if not prop.is_hidden and prop.status != Property.Status.REMOVED and not best_address(prop)),
        "pendientes_por_fuente": dict(sorted((source, dict(counts)) for source, counts in by_source.items())),
    }


def load_feedback():
    wb = load_workbook(INPUT_XLSX, read_only=True, data_only=False)
    ws = wb["Pendientes"]
    headers = [cell.value for cell in ws[1]]
    rows = {}
    decisions = Counter()
    notes = {}
    for values in ws.iter_rows(min_row=2, values_only=True):
        row = dict(zip(headers, values))
        try:
            property_id = int(row.get("id_propiedad") or 0)
        except (TypeError, ValueError):
            continue
        rows[property_id] = row
        decision = safe(row.get("decision_manual")).upper()
        if decision:
            decisions[decision] += 1
        note = safe(row.get("notas_manual"))
        if note:
            notes[property_id] = note
    return rows, decisions, notes


def add_summary(summary, key, item):
    summary.setdefault(key, []).append(item)


def mark_manual_property_field(property_obj, field, value, now, update_fields, apply):
    value = safe(value)
    if not value:
        return
    if getattr(property_obj, field) != value:
        setattr(property_obj, field, value)
        update_fields.add(field)
    if field == "address":
        normalized = normalize_address(value)
        if property_obj.normalized_address != normalized:
            property_obj.normalized_address = normalized
            update_fields.add("normalized_address")
        current = location_or_none(property_obj)
        if current and not current.manually_corrected and apply:
            current.delete()
    overrides = dict(property_obj.manual_overrides or {})
    overrides[field] = now.isoformat()
    property_obj.manual_overrides = overrides
    property_obj.data_manually_corrected_at = now
    update_fields.update({"manual_overrides", "data_manually_corrected_at"})


def save_property(property_obj, update_fields, apply):
    if update_fields and apply:
        property_obj.save(update_fields=sorted(update_fields))


def update_location(property_obj, latitude, longitude, provider, query, summary, apply):
    current = location_or_none(property_obj)
    if current and current.manually_corrected:
        add_summary(summary, "manual_location_preserved", property_obj.pk)
        return False
    defaults = {
        "latitude": float(latitude),
        "longitude": float(longitude),
        "precision": PropertyLocation.Precision.EXACT,
        "query": query or "",
        "provider": provider,
        "confidence": 1.0,
        "outside_target": False,
        "manually_corrected": False,
    }
    if apply:
        PropertyLocation.objects.update_or_create(property=property_obj, defaults=defaults)
        evidence = dict(property_obj.location_evidence or {})
        evidence[SCRIPT_TAG] = {
            "provider": provider,
            "latitude": float(latitude),
            "longitude": float(longitude),
            "query": query or "",
        }
        property_obj.detected_latitude = float(latitude)
        property_obj.detected_longitude = float(longitude)
        property_obj.location_source = Property.LocationSource.MAP
        property_obj.location_confidence = Property.LocationConfidence.HIGH
        property_obj.location_evidence = evidence
        property_obj.save(
            update_fields=[
                "detected_latitude",
                "detected_longitude",
                "location_source",
                "location_confidence",
                "location_evidence",
            ]
        )
    return True


def apply_zone_from_point(property_obj, latitude, longitude, provider, reason, force_needs_review, summary, apply):
    result = infer_territory_for_point(latitude, longitude)
    zone = safe(result.zone if result else "")
    if not zone:
        add_summary(summary, "zone_skipped", {"id": property_obj.pk, "reason": "sin_match_poligono"})
        return ""
    now = timezone.now()
    evidence = dict(result.evidence if result and result.evidence else {})
    evidence.update({"provider": provider, "curacion": SCRIPT_TAG, "reason": reason})
    if apply:
        property_obj.inferred_partido = result.partido
        property_obj.inferred_locality = result.locality
        property_obj.inferred_zone = zone
        property_obj.inferred_neighborhood = zone
        property_obj.territory_confidence = result.confidence
        property_obj.territory_source_method = result.source_method
        property_obj.territory_needs_review = bool(result.needs_review or force_needs_review)
        property_obj.territory_evidence = evidence
        property_obj.territory_inferred_at = now
        property_obj.zone_conflict = False
        property_obj.zone_needs_review = bool(result.needs_review or force_needs_review)
        property_obj.zone_inference_evidence = evidence
        property_obj.zone_inferred_at = now
        property_obj.save(
            update_fields=[
                "inferred_partido",
                "inferred_locality",
                "inferred_zone",
                "inferred_neighborhood",
                "territory_confidence",
                "territory_source_method",
                "territory_needs_review",
                "territory_evidence",
                "territory_inferred_at",
                "zone_conflict",
                "zone_needs_review",
                "zone_inference_evidence",
                "zone_inferred_at",
            ]
        )
        PropertyLocationIntelligence.objects.update_or_create(
            property=property_obj,
            defaults={
                "partido_name": result.partido,
                "locality_name": result.locality,
                "zone_name": zone,
                "match_method": PropertyLocationIntelligence.MatchMethod.COORDINATES,
                "confidence": result.confidence,
            },
        )
    add_summary(summary, "zone_applied", {"id": property_obj.pk, "zone": zone, "needs_review": bool(result.needs_review or force_needs_review)})
    return zone


def apply_manual_zone(property_obj, zone, reason, summary, apply):
    now = timezone.now()
    if apply:
        evidence = dict(property_obj.zone_inference_evidence or {})
        evidence[SCRIPT_TAG] = {"reason": reason, "zone": zone}
        property_obj.inferred_zone = zone
        property_obj.inferred_neighborhood = zone
        property_obj.zone_needs_review = True
        property_obj.zone_inference_evidence = evidence
        property_obj.zone_inferred_at = now
        property_obj.save(update_fields=["inferred_zone", "inferred_neighborhood", "zone_needs_review", "zone_inference_evidence", "zone_inferred_at"])
        PropertyLocationIntelligence.objects.update_or_create(
            property=property_obj,
            defaults={
                "zone_name": zone,
                "match_method": PropertyLocationIntelligence.MatchMethod.ZONE,
                "confidence": "manual_feedback",
            },
        )
    add_summary(summary, "manual_zone_applied", {"id": property_obj.pk, "zone": zone})


def mark_removed(property_obj, reason, summary, apply):
    listing_ids = list(property_obj.listings.filter(active=True).values_list("id", flat=True))
    if apply:
        property_obj.listings.filter(active=True).update(active=False, source_status="removed", missing_runs=2)
        marker = f"{SCRIPT_TAG}: {reason}"
        notes = property_obj.personal_notes or ""
        if marker not in notes:
            property_obj.personal_notes = "\n\n".join(part for part in [notes, marker] if part)
        property_obj.status = Property.Status.REMOVED
        property_obj.is_hidden = True
        property_obj.save(update_fields=["status", "is_hidden", "personal_notes"])
    add_summary(summary, "removed", {"id": property_obj.pk, "listing_ids": listing_ids, "reason": reason})


def fetch_url_text(url):
    session = requests.Session()
    session.headers.update({"User-Agent": "Mozilla/5.0 (compatible; RadarCuration/1.0)"})
    response = session.get(url, timeout=35)
    return response.status_code, response.text or ""


def fincas_sweep(summary, apply):
    candidates = []
    qs = (
        Property.objects.prefetch_related("listings__source")
        .filter(listings__source__slug="fincas", status=Property.Status.ACTIVE, is_hidden=False)
        .distinct()
    )
    for property_obj in qs:
        if property_obj.pk in EXPLICIT_REMOVE_IDS:
            continue
        listing = property_obj.listings.filter(source__slug="fincas").order_by("-active", "-last_seen_at").first()
        text = " ".join([property_obj.title or "", property_obj.address or "", listing.url if listing else ""])
        if not OUTSIDE_RE.search(text):
            continue
        has_inside_coordinate = False
        fetch_status = None
        fetch_error = ""
        if listing and listing.url:
            try:
                fetch_status, markup = fetch_url_text(listing.url)
                has_inside_coordinate = bool(extract_map_coordinates(markup, require_target_bounds=True))
            except Exception as exc:
                fetch_error = repr(exc)
        if has_inside_coordinate:
            add_summary(summary, "fincas_outside_preserved_inside_map", {"id": property_obj.pk, "url": listing.url if listing else ""})
            continue
        item = {
            "id": property_obj.pk,
            "title": property_obj.title,
            "url": listing.url if listing else "",
            "http_status": fetch_status,
            "error": fetch_error,
        }
        candidates.append(item)
        mark_removed(property_obj, "fuera de zona detectado en Fincas/Argencasas", summary, apply)
    summary["fincas_outside_candidates"] = candidates


def apply_feedback(apply):
    summary = {
        "mode": "apply" if apply else "dry_run",
        "input_excel": str(INPUT_XLSX),
        "output_excel": str(OUTPUT_XLSX) if apply else "",
    }
    rows, decisions, notes = load_feedback()
    summary["feedback_decisions"] = dict(decisions)
    summary["approved_found"] = sorted(pid for pid in APPROVED_IDS if safe(rows.get(pid, {}).get("decision_manual")).upper() == "APROBADO")
    summary["rejected_found"] = sorted(pid for pid in REJECTED_IDS if safe(rows.get(pid, {}).get("decision_manual")).upper() == "RECHAZADO")
    summary["before"] = current_counts()
    now = timezone.now()
    geocoder = Geocoder()

    if apply:
        shutil.copy2(DB_PATH, BACKUP_PATH)
        summary["backup"] = str(BACKUP_PATH)
    else:
        summary["backup"] = ""

    for property_id, fix in MANUAL_FIXES.items():
        property_obj = Property.objects.select_related("location", "location_intelligence").filter(pk=property_id).first()
        if not property_obj:
            add_summary(summary, "missing_property", property_id)
            continue
        update_fields = set()
        for field in ("address", "locality", "neighborhood"):
            if fix.get(field):
                mark_manual_property_field(property_obj, field, fix[field], now, update_fields, apply)
                add_summary(summary, f"{field}_updates", {"id": property_id, "to": fix[field]})
        save_property(property_obj, update_fields, apply)
        if fix.get("latitude") is not None and fix.get("longitude") is not None:
            updated = update_location(
                property_obj,
                fix["latitude"],
                fix["longitude"],
                fix["provider"],
                first_listing_url(property_obj),
                summary,
                apply,
            )
            if updated:
                add_summary(summary, "location_updates", {"id": property_id, "provider": fix["provider"]})
            apply_zone_from_point(
                property_obj,
                fix["latitude"],
                fix["longitude"],
                fix["provider"],
                fix["reason"],
                bool(fix.get("force_needs_review")),
                summary,
                apply,
            )
        elif fix.get("geocode"):
            add_summary(summary, "geocode_required", property_id)
            if apply:
                property_obj.refresh_from_db()
                location = geocoder.geocode_property_from_cache(property_obj, force=True)
                if not location:
                    try:
                        location = geocoder.geocode_property(property_obj, force=True)
                    except Exception as exc:
                        add_summary(summary, "geocode_errors", {"id": property_id, "error": repr(exc)})
                        location = None
                if location and not location.outside_target:
                    add_summary(summary, "geocode_location", {"id": property_id, "lat": location.latitude, "lon": location.longitude, "provider": location.provider})
                    apply_zone_from_point(property_obj, location.latitude, location.longitude, location.provider, fix["reason"], False, summary, apply)
                elif location:
                    add_summary(summary, "geocode_outside_target", {"id": property_id, "lat": location.latitude, "lon": location.longitude})

    for property_id in EXPLICIT_REMOVE_IDS:
        property_obj = Property.objects.prefetch_related("listings").filter(pk=property_id).first()
        if property_obj:
            reason = "Zonaprop 410 Gone" if property_id in {5309, 5312} else "fuera de zona por feedback manual"
            mark_removed(property_obj, reason, summary, apply)

    fincas_sweep(summary, apply)

    for property_id in sorted(NO_DATA_IDS):
        add_summary(summary, "reviewed_no_data_left_pending", {"id": property_id, "note": notes.get(property_id, "revisado manualmente, sin datos suficientes")})

    summary["after"] = current_counts()
    return summary, notes


def first_listing(property_obj):
    return property_obj.listings.select_related("source", "agency").order_by("-active", "-last_seen_at").first()


def first_listing_url(property_obj):
    listing = first_listing(property_obj)
    return listing.url if listing else ""


def source_names(property_obj):
    names = []
    for listing in property_obj.listings.select_related("source").all():
        if listing.source.name not in names:
            names.append(listing.source.name)
    return ", ".join(names)


def agency_names(property_obj):
    names = []
    for listing in property_obj.listings.select_related("agency").all():
        if listing.agency and listing.agency.name not in names:
            names.append(listing.agency.name)
    return ", ".join(names)


def status_label(status):
    return dict(Property.Status.choices).get(status, status)


def residual_queryset():
    return (
        Property.objects.select_related("location", "location_intelligence")
        .prefetch_related("listings__source", "listings__agency")
        .filter(no_zone_query() | Q(address="") | Q(address__isnull=True))
        .filter(is_hidden=False)
        .exclude(status=Property.Status.REMOVED)
        .distinct()
        .order_by("id")
    )


def property_row(property_obj, reviewed_notes):
    location = location_or_none(property_obj)
    intel = intel_or_none(property_obj)
    listing = first_listing(property_obj)
    url = listing.url if listing else ""
    evidence = {
        "location_precision": location.precision if location else "",
        "location_query": location.query if location else "",
        "manual_overrides": property_obj.manual_overrides or {},
        "source_locality": property_obj.locality,
        "source_zone": property_obj.inferred_zone,
        "territory_evidence": property_obj.territory_evidence or {},
    }
    if property_obj.pk in reviewed_notes:
        evidence["feedback_20260626"] = reviewed_notes[property_obj.pk] or "revisado manualmente"
    return [
        property_obj.pk,
        property_obj.title,
        (property_obj.description or "")[:500],
        property_obj.address or "",
        property_obj.detected_address or "",
        "",
        property_obj.normalized_address or "",
        property_obj.locality or "",
        property_obj.detected_locality or "",
        property_obj.neighborhood or "",
        property_obj.detected_neighborhood or "",
        property_obj.inferred_zone or "",
        "",
        intel.zone_name if intel else "",
        "BAJA",
        residual_reason(property_obj),
        "DUDOSO",
        location.latitude if location else None,
        location.longitude if location else None,
        location.provider if location else "",
        source_names(property_obj),
        agency_names(property_obj),
        url,
        "" if url else "Sin Listing asociado en la base",
        status_label(property_obj.status),
        property_obj.get_property_type_display(),
        property_obj.operation,
        property_obj.currency or "",
        float(property_obj.price) if property_obj.price is not None else None,
        property_obj.rooms,
        property_obj.bedrooms,
        float(property_obj.bathrooms) if property_obj.bathrooms is not None else None,
        float(property_obj.covered_area) if property_obj.covered_area is not None else None,
        float(property_obj.total_area) if property_obj.total_area is not None else None,
        property_obj.last_seen_at.isoformat() if property_obj.last_seen_at else "",
        residual_reason(property_obj),
        json.dumps(evidence, ensure_ascii=False, default=str),
        "",
        "",
    ]


def generate_excel(reviewed_notes):
    wb = Workbook()
    ws = wb.active
    ws.title = "Pendientes"
    ws.append(HEADERS)
    header_fill = PatternFill("solid", fgColor="1F4E78")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center")
    rows = list(residual_queryset())
    for property_obj in rows:
        ws.append(property_row(property_obj, reviewed_notes))

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    table = Table(displayName="PendientesFeedbackTable", ref=ws.dimensions)
    table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
    ws.add_table(table)
    decision_col = HEADERS.index("decision_manual") + 1
    validation = DataValidation(type="list", formula1='"APROBADO,RECHAZADO"', allow_blank=True)
    ws.add_data_validation(validation)
    if ws.max_row >= 2:
        validation.add(f"{get_column_letter(decision_col)}2:{get_column_letter(decision_col)}{ws.max_row}")

    widths = {
        "A": 12,
        "B": 42,
        "C": 52,
        "D": 28,
        "H": 18,
        "K": 22,
        "L": 22,
        "R": 58,
        "AE": 24,
        "AF": 70,
        "AG": 18,
        "AH": 36,
    }
    for column in range(1, len(HEADERS) + 1):
        letter = get_column_letter(column)
        ws.column_dimensions[letter].width = widths.get(letter, 16)
    price_col = HEADERS.index("precio") + 1
    for row in range(2, ws.max_row + 1):
        ws.cell(row, price_col).number_format = "#,##0"
        ws.cell(row, HEADERS.index("descripcion_resumen") + 1).alignment = Alignment(wrap_text=True, vertical="top")
        ws.cell(row, HEADERS.index("evidencia") + 1).alignment = Alignment(wrap_text=True, vertical="top")

    summary_ws = wb.create_sheet("Resumen")
    summary_ws.append(["metrica", "valor"])
    summary_ws.append(["pendientes", len(rows)])
    summary_ws.append(["generado_por", SCRIPT_TAG])
    wb.save(OUTPUT_XLSX)
    return {"path": str(OUTPUT_XLSX), "pending_rows": len(rows), "columns": HEADERS}


def validate_excel():
    wb = load_workbook(OUTPUT_XLSX, data_only=True)
    ws = wb["Pendientes"]
    headers = [cell.value for cell in ws[1]]
    price_col = headers.index("precio") + 1
    decision_col = headers.index("decision_manual") + 1
    notes_col = headers.index("notas_manual") + 1
    price_types = Counter(type(ws.cell(row, price_col).value).__name__ for row in range(2, ws.max_row + 1))
    return {
        "sheets": wb.sheetnames,
        "rows": ws.max_row - 1,
        "auto_filter": ws.auto_filter.ref,
        "tables": list(ws.tables.keys()),
        "freeze_panes": ws.freeze_panes,
        "price_types": dict(price_types),
        "nonblank_decisions": sum(1 for row in range(2, ws.max_row + 1) if ws.cell(row, decision_col).value),
        "nonblank_notes": sum(1 for row in range(2, ws.max_row + 1) if ws.cell(row, notes_col).value),
        "validations": [
            {"type": dv.type, "formula1": dv.formula1, "sqref": str(dv.sqref)}
            for dv in ws.data_validations.dataValidation
        ],
        "missing_columns": [column for column in HEADERS if column not in headers],
    }


def sample(ids):
    out = []
    for property_obj in (
        Property.objects.select_related("location", "location_intelligence")
        .prefetch_related("listings__source")
        .filter(pk__in=ids)
        .order_by("id")
    ):
        location = location_or_none(property_obj)
        intel = intel_or_none(property_obj)
        out.append(
            {
                "id": property_obj.pk,
                "address": property_obj.address,
                "locality": property_obj.locality,
                "zone": property_obj.inferred_zone,
                "intel_zone": intel.zone_name if intel else "",
                "status": property_obj.status,
                "is_hidden": property_obj.is_hidden,
                "location": None
                if not location
                else {
                    "lat": location.latitude,
                    "lon": location.longitude,
                    "provider": location.provider,
                    "manual": location.manually_corrected,
                },
                "manual_overrides": property_obj.manual_overrides or {},
                "corrected": bool(property_obj.data_manually_corrected_at),
                "active_listings": list(property_obj.listings.filter(active=True).values_list("id", flat=True)),
            }
        )
    return out


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--apply", action="store_true")
    parser.add_argument("--generate-excel", action="store_true")
    parser.add_argument("--json-out")
    args = parser.parse_args()

    summary, reviewed_notes = apply_feedback(args.apply)
    if args.generate_excel:
        summary["excel"] = generate_excel(reviewed_notes)
        summary["excel_validation"] = validate_excel()
    else:
        summary["excel"] = None
        summary["excel_validation"] = None
    summary["sample"] = sample(sorted(set(MANUAL_FIXES) | EXPLICIT_REMOVE_IDS))
    output = json.dumps(summary, ensure_ascii=False, indent=2, default=str)
    print(output)
    if args.json_out:
        Path(args.json_out).write_text(output, encoding="utf-8")


if __name__ == "__main__":
    main()
