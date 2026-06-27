import argparse
import html
import json
import os
import re
import shutil
import sys
import time
from collections import Counter, defaultdict
from pathlib import Path
from urllib.parse import urlparse

import requests
from bs4 import BeautifulSoup
from django.conf import settings
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo


ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))
os.environ.setdefault("DJANGO_SETTINGS_MODULE", "config.settings")

import django

django.setup()

from django.db import transaction
from django.db.models import Q
from django.utils import timezone

from properties.models import (
    GeocodeCache,
    Listing,
    Property,
    PropertyLocation,
    PropertyLocationIntelligence,
)
from properties.services.geocoding import Geocoder, best_address
from properties.services.normalization import (
    normalize_address,
    normalize_locality,
    normalize_neighborhood_name,
    normalize_whitespace,
)
from properties.services.territory_hierarchy import infer_territory_for_point, territory_values_from_result
from properties.services.zone_names import canonicalize_unified_zone_name


INPUT_XLSX = ROOT / "tmp" / "curacion_zonas_pendientes_2026-06-20.xlsx"
OUTPUT_XLSX = ROOT / "tmp" / "curacion_zonas_pendientes_2026-06-20_mapas.xlsx"
FETCH_CACHE = ROOT / "tmp" / "curacion_mapas_fetch_cache_20260620.json"
DB_PATH = ROOT / "db.sqlite3"
BACKUP_PATH = ROOT / "tmp" / "db.sqlite3.backup_curacion_mapas_20260620"
SCRIPT_TAG = "curacion_mapas_20260620"

MIAMI_DEFAULT = (25.7308309, -80.444149)
APPROVED_IDS_EXPECTED = {
    3165,
    3278,
    5717,
    5727,
    87,
    293,
    394,
    413,
    436,
    446,
    786,
    987,
    1013,
    1014,
    1021,
    1028,
}
MAP_SWEEP_SOURCES = {
    "miglierini",
    "marcelo-russo",
    "riquelme",
    "paula-fossati",
    "argenprop",
    "inmuebles-clarin",
}
APPROVED_NETWORK_SOURCES = MAP_SWEEP_SOURCES | {"zonaprop", "analia-fernandez", "faella", "mercadolibre"}
SOURCE_DELAYS = {
    "miglierini": 5,
    "marcelo-russo": 3,
    "riquelme": 3,
    "paula-fossati": 4,
    "argenprop": 3,
    "inmuebles-clarin": 4,
    "zonaprop": 6,
    "analia-fernandez": 3,
    "faella": 4,
    "mercadolibre": 6,
}
PROVIDER_BY_SOURCE = {
    "miglierini": "miglierini_map",
    "marcelo-russo": "marcelo_russo_map",
    "riquelme": "riquelme_map",
    "paula-fossati": "paula_fossati_map",
    "argenprop": "argenprop_map",
    "inmuebles-clarin": "inmuebles_clarin_map",
    "zonaprop": "zonaprop_map",
    "analia-fernandez": "analia_fernandez_map",
    "faella": "faella_map",
    "mercadolibre": "mercadolibre_map",
}


def safe(value):
    if value is None:
        return ""
    return normalize_whitespace(str(value))


def _float(value):
    if value is None:
        return None
    try:
        return float(str(value).replace(",", "."))
    except (TypeError, ValueError):
        return None


def in_target_bounds(latitude, longitude):
    if latitude is None or longitude is None:
        return False
    bounds = settings.HURLINGHAM_BOUNDS
    return (
        bounds["south"] <= latitude <= bounds["north"]
        and bounds["west"] <= longitude <= bounds["east"]
    )


def is_default_miami(latitude, longitude):
    if latitude is None or longitude is None:
        return False
    return abs(latitude - MIAMI_DEFAULT[0]) < 0.000001 and abs(longitude - MIAMI_DEFAULT[1]) < 0.000001


def normalized_zone(value):
    zone = canonicalize_unified_zone_name(normalize_neighborhood_name(value))
    return "Cartero" if zone == "Barrio Cartero" else zone


def no_zone_query():
    return (
        (Q(inferred_zone__isnull=True) | Q(inferred_zone=""))
        & (
            Q(location_intelligence__isnull=True)
            | Q(location_intelligence__zone_name__isnull=True)
            | Q(location_intelligence__zone_name="")
        )
    )


def location_or_none(property_obj):
    return getattr(property_obj, "location", None)


def intel_or_none(property_obj):
    return getattr(property_obj, "location_intelligence", None)


def residual_reason(property_obj):
    reasons = []
    intel = intel_or_none(property_obj)
    if not property_obj.inferred_zone and not (intel.zone_name if intel else ""):
        reasons.append("sin zona")
    if not best_address(property_obj):
        reasons.append("sin direccion util")
    return ", ".join(reasons)


def current_counts():
    props = list(
        Property.objects.select_related("location", "location_intelligence")
        .prefetch_related("listings__source")
        .all()
    )
    by_source = defaultdict(lambda: Counter())
    for property_obj in props:
        reason = residual_reason(property_obj)
        if not reason:
            continue
        sources = {listing.source.slug for listing in property_obj.listings.all()} or {"sin_listing"}
        for source in sources:
            by_source[source]["pendientes"] += 1
            if "sin zona" in reason:
                by_source[source]["sin_zona"] += 1
            if "sin direccion" in reason:
                by_source[source]["sin_direccion"] += 1
    miglierini = by_source.get("miglierini", Counter())
    return {
        "sin_zona_operativa": Property.objects.filter(no_zone_query()).count(),
        "sin_direccion_util": sum(1 for prop in props if not best_address(prop)),
        "miglierini_pendientes": miglierini.get("pendientes", 0),
        "pendientes_por_fuente": dict(sorted((source, dict(counts)) for source, counts in by_source.items())),
    }


def load_feedback_rows():
    wb = load_workbook(INPUT_XLSX, read_only=True, data_only=False)
    ws = wb["Pendientes"]
    headers = [cell.value for cell in ws[1]]
    rows_by_id = {}
    decisions = {}
    notes = {}
    for values in ws.iter_rows(min_row=2, values_only=True):
        row = dict(zip(headers, values))
        try:
            property_id = int(row.get("id_propiedad") or 0)
        except (TypeError, ValueError):
            continue
        rows_by_id[property_id] = row
        decision = safe(row.get("decision_manual")).upper()
        if decision:
            decisions[property_id] = decision
        note = safe(row.get("notas_manual"))
        if note:
            notes[property_id] = note
    approved_ids = sorted(pid for pid, decision in decisions.items() if decision == "APROBADO")
    return rows_by_id, decisions, notes, approved_ids


def first_active_listing(property_obj, preferred_sources=None):
    preferred_sources = set(preferred_sources or [])
    qs = property_obj.listings.select_related("source", "agency").order_by("-active", "-last_seen_at")
    if preferred_sources:
        for listing in qs:
            if listing.source.slug in preferred_sources:
                return listing
    return qs.first()


def source_names(property_obj):
    names = []
    for listing in property_obj.listings.select_related("source").all():
        if listing.source.name not in names:
            names.append(listing.source.name)
    return ", ".join(names)


def agency_names(property_obj):
    names = []
    for listing in property_obj.listings.select_related("agency").all():
        if listing.agency and listing.agency.name not in names:
            names.append(listing.agency.name)
    return ", ".join(names)


def source_slug_from_url(url, fallback=""):
    host = urlparse(url or "").netloc.lower()
    if "miglieriniprop" in host:
        return "miglierini"
    if "riquelmepropiedades" in host:
        return "riquelme"
    if "marcelorussoprop" in host:
        return "marcelo-russo"
    if "paulafossati" in host:
        return "paula-fossati"
    if "argenprop" in host:
        return "argenprop"
    if "inmuebles.clarin" in host:
        return "inmuebles-clarin"
    if "zonaprop" in host:
        return "zonaprop"
    if "fernandezpropiedades" in host:
        return "analia-fernandez"
    if "mercadolibre" in host:
        return fallback or "mercadolibre"
    return fallback


def add_coordinate(candidates, method, latitude, longitude, address="", confidence="high"):
    lat = _float(latitude)
    lon = _float(longitude)
    if lat is None or lon is None:
        return
    if is_default_miami(lat, lon) or not in_target_bounds(lat, lon):
        return
    candidates.append(
        {
            "method": method,
            "latitude": lat,
            "longitude": lon,
            "address": safe(address),
            "confidence": confidence,
        }
    )


def extract_map_coordinates(markup):
    soup = BeautifulSoup(markup or "", "html.parser")
    candidates = []

    for tag in soup.select("[data-latitude][data-longitude]"):
        add_coordinate(
            candidates,
            "data-latitude",
            tag.get("data-latitude"),
            tag.get("data-longitude"),
        )

    for tag in soup.select("[data-map]"):
        raw = html.unescape(tag.get("data-map") or "")
        try:
            payload = json.loads(raw)
        except json.JSONDecodeError:
            continue
        if isinstance(payload, list):
            payload = payload[0] if payload else {}
        if not isinstance(payload, dict):
            continue
        add_coordinate(
            candidates,
            "data-map",
            payload.get("latitude") or payload.get("lat"),
            payload.get("longitude") or payload.get("lng") or payload.get("lang") or payload.get("lon"),
            payload.get("address") or "",
        )

    for match in re.finditer(r"propertyMapData\s*=\s*(\{.*?\})\s*;", markup or "", re.S):
        try:
            payload = json.loads(match.group(1))
        except json.JSONDecodeError:
            continue
        add_coordinate(
            candidates,
            "propertyMapData",
            payload.get("latitude") or payload.get("lat"),
            payload.get("longitude") or payload.get("lng") or payload.get("lang") or payload.get("lon"),
            payload.get("address") or "",
        )

    for tag in soup.select('script[type="application/ld+json"]'):
        try:
            payload = json.loads(tag.get_text() or "{}")
        except json.JSONDecodeError:
            continue
        stack = [payload]
        while stack:
            item = stack.pop()
            if isinstance(item, list):
                stack.extend(item)
            elif isinstance(item, dict):
                geo = item.get("geo")
                if isinstance(geo, dict):
                    add_coordinate(
                        candidates,
                        "jsonld_geo",
                        geo.get("latitude") or geo.get("lat"),
                        geo.get("longitude") or geo.get("lng") or geo.get("lon"),
                    )
                stack.extend(value for value in item.values() if isinstance(value, (list, dict)))

    patterns = (
        r'"latitude"\s*:\s*"?(?P<lat>-?\d+[\.,]\d+)"?.{0,180}?"longitude"\s*:\s*"?(?P<lon>-?\d+[\.,]\d+)"?',
        r'"lat"\s*:\s*"?(?P<lat>-?\d+[\.,]\d+)"?.{0,180}?"lng"\s*:\s*"?(?P<lon>-?\d+[\.,]\d+)"?',
        r"data-latitude\s*=\s*['\"](?P<lat>-?\d+[\.,]\d+)['\"].{0,180}?data-longitude\s*=\s*['\"](?P<lon>-?\d+[\.,]\d+)['\"]",
    )
    for pattern in patterns:
        for match in re.finditer(pattern, markup or "", re.I | re.S):
            add_coordinate(candidates, "regex_pair", match.group("lat"), match.group("lon"))

    latitudes = re.findall(r"-34[\.,]\d{4,}", markup or "")
    longitudes = re.findall(r"-58[\.,]\d{4,}", markup or "")
    if latitudes and longitudes:
        add_coordinate(candidates, "fallback_latlon", latitudes[0], longitudes[0], confidence="medium")

    output = []
    seen = set()
    preference = {
        "data-latitude": 0,
        "propertyMapData": 1,
        "data-map": 2,
        "jsonld_geo": 3,
        "regex_pair": 4,
        "fallback_latlon": 5,
    }
    for item in sorted(candidates, key=lambda candidate: preference.get(candidate["method"], 99)):
        key = (round(item["latitude"], 7), round(item["longitude"], 7))
        if key in seen:
            continue
        seen.add(key)
        output.append(item)
    return output


def choose_coordinate(candidates, allow_fallback):
    for candidate in candidates:
        if candidate["method"] != "fallback_latlon" or allow_fallback:
            return candidate
    return None


class Fetcher:
    def __init__(self, refresh=False):
        self.refresh = refresh
        self.session = requests.Session()
        self.session.headers.update(
            {
                "User-Agent": (
                    "Mozilla/5.0 (compatible; RadarInmobiliarioCuration/1.0; "
                    "+https://local)"
                )
            }
        )
        self.cache = {}
        self.last_request = defaultdict(float)
        if FETCH_CACHE.exists() and not refresh:
            try:
                self.cache = json.loads(FETCH_CACHE.read_text(encoding="utf-8"))
            except json.JSONDecodeError:
                self.cache = {}

    def save(self):
        serialized = json.dumps(self.cache, ensure_ascii=False, indent=2, sort_keys=True, default=str)
        tmp_path = FETCH_CACHE.with_name(f"{FETCH_CACHE.stem}.{os.getpid()}.tmp")
        try:
            tmp_path.write_text(serialized, encoding="utf-8")
            for attempt in range(5):
                try:
                    tmp_path.replace(FETCH_CACHE)
                    return
                except PermissionError:
                    time.sleep(0.2 * (attempt + 1))
            FETCH_CACHE.write_text(serialized, encoding="utf-8")
        except OSError:
            pass
        finally:
            try:
                if tmp_path.exists():
                    tmp_path.unlink()
            except OSError:
                pass

    def fetch_map(self, url, source_slug):
        if not url:
            return {"status": "no_url", "coordinates": [], "blocked": False, "error": "sin_url"}
        if url in self.cache and not self.refresh:
            return self.cache[url]
        delay = SOURCE_DELAYS.get(source_slug, 3)
        elapsed = time.monotonic() - self.last_request[source_slug]
        if elapsed < delay:
            time.sleep(delay - elapsed)
        try:
            response = self.session.get(url, timeout=35)
            self.last_request[source_slug] = time.monotonic()
            text = response.text or ""
            blocked = bool(
                re.search(
                    r"just a moment|enable javascript and cookies|captcha|cf_chl|cloudflare",
                    text,
                    re.I,
                )
            )
            coordinates = extract_map_coordinates(text)
            result = {
                "status": response.status_code,
                "bytes": len(text),
                "blocked": blocked,
                "coordinates": coordinates,
                "error": "",
            }
        except Exception as exc:
            result = {"status": "error", "bytes": 0, "blocked": False, "coordinates": [], "error": repr(exc)}
        self.cache[url] = result
        self.save()
        return result


def set_manual_field(property_obj, field, value, now, changes, manual=True):
    value = safe(value)
    if not value:
        return False
    changed = getattr(property_obj, field) != value
    if changed:
        setattr(property_obj, field, value)
        changes.add(field)
    if field == "address":
        normalized = normalize_address(value)
        if property_obj.normalized_address != normalized:
            property_obj.normalized_address = normalized
            changes.add("normalized_address")
    if manual:
        overrides = dict(property_obj.manual_overrides or {})
        overrides[field] = now.isoformat()
        property_obj.manual_overrides = overrides
        property_obj.data_manually_corrected_at = now
        changes.update({"manual_overrides", "data_manually_corrected_at"})
    return changed


def save_property(property_obj, changes, apply):
    if changes and apply:
        property_obj.save(update_fields=sorted(changes))


def update_source_location(property_obj, coordinate, source_slug, url, summary, apply):
    current = location_or_none(property_obj)
    if current and current.manually_corrected:
        summary["manual_location_preserved_ids"].append(property_obj.pk)
        return False, "manual_location_preserved"
    provider = PROVIDER_BY_SOURCE.get(source_slug, f"{source_slug}_map")
    data = {
        "latitude": coordinate["latitude"],
        "longitude": coordinate["longitude"],
        "precision": PropertyLocation.Precision.EXACT,
        "query": url or "",
        "provider": provider,
        "confidence": 1.0 if coordinate.get("confidence") == "high" else 0.8,
        "outside_target": False,
        "manually_corrected": False,
    }
    if apply:
        PropertyLocation.objects.update_or_create(property=property_obj, defaults=data)
        evidence = dict(property_obj.location_evidence or {})
        evidence[SCRIPT_TAG] = {
            "source": source_slug,
            "url": url,
            "method": coordinate["method"],
            "latitude": coordinate["latitude"],
            "longitude": coordinate["longitude"],
        }
        property_obj.detected_latitude = coordinate["latitude"]
        property_obj.detected_longitude = coordinate["longitude"]
        property_obj.location_source = Property.LocationSource.MAP
        property_obj.location_confidence = Property.LocationConfidence.HIGH
        property_obj.location_evidence = evidence
        property_obj.save(
            update_fields=[
                "detected_latitude",
                "detected_longitude",
                "location_source",
                "location_confidence",
                "location_evidence",
            ]
        )
    return True, "updated"


def create_detected_location_if_needed(property_obj, summary, apply):
    if location_or_none(property_obj):
        return False
    lat = property_obj.detected_latitude
    lon = property_obj.detected_longitude
    if lat is None or lon is None or not in_target_bounds(lat, lon) or is_default_miami(lat, lon):
        return False
    if apply:
        PropertyLocation.objects.update_or_create(
            property=property_obj,
            defaults={
                "latitude": lat,
                "longitude": lon,
                "precision": PropertyLocation.Precision.EXACT,
                "query": "detected_coordinates",
                "provider": "detected_coordinates",
                "confidence": 0.9,
                "outside_target": False,
                "manually_corrected": False,
            },
        )
    summary["detected_location_created_ids"].append(property_obj.pk)
    return True


def source_zone_for_inference(property_obj, row=None):
    values = []
    if row:
        values.extend([row.get("zona_actual"), row.get("barrio_actual")])
    values.extend([property_obj.inferred_zone, property_obj.neighborhood, property_obj.detected_neighborhood])
    for value in values:
        zone = normalized_zone(value)
        if zone:
            return zone
    return ""


def source_locality_for_inference(property_obj, row=None):
    values = []
    if row:
        values.append(row.get("localidad_actual"))
    values.extend([property_obj.locality, property_obj.detected_locality])
    for value in values:
        locality = normalize_locality(value)
        if locality:
            return locality
    return ""


def apply_zone_from_location(property_obj, *, row=None, tag, coordinate_evidence=None, apply=True):
    location = location_or_none(property_obj)
    if not location:
        if property_obj.detected_latitude is not None and property_obj.detected_longitude is not None:
            coordinate_source = "detected_coordinates"
            latitude = property_obj.detected_latitude
            longitude = property_obj.detected_longitude
        else:
            return {"applied": False, "zone": "", "reason": "sin_coordenadas"}
    else:
        if location.outside_target or is_default_miami(location.latitude, location.longitude):
            return {"applied": False, "zone": "", "reason": "coordenada_fuera_o_invalida"}
        coordinate_source = location.provider
        latitude = location.latitude
        longitude = location.longitude
    result = infer_territory_for_point(
        latitude,
        longitude,
        coordinate_source=coordinate_source,
        extra_evidence={
            "curacion": tag,
            "latitude": latitude,
            "longitude": longitude,
            **(coordinate_evidence or {}),
        },
        source_zone=source_zone_for_inference(property_obj, row),
        source_locality=source_locality_for_inference(property_obj, row),
    )
    values = territory_values_from_result(result)
    if not values["zone"]:
        return {"applied": False, "zone": "", "reason": "sin_match_poligono", "values": values}
    approved_zone = normalized_zone(row.get("zona_actual") if row else "") or normalized_zone(row.get("barrio_actual") if row else "")
    source_zone = source_zone_for_inference(property_obj, row)
    conflict = bool(
        (approved_zone and approved_zone != values["zone"])
        or (source_zone and source_zone != values["zone"])
        or values["needs_review"]
    )
    evidence = {
        **values["evidence"],
        "curacion": tag,
        "map_wins_conflict": bool(conflict),
        "approved_zone": approved_zone,
        "source_zone_used": source_zone,
        "coordinate_evidence": coordinate_evidence or {},
    }
    if apply:
        now = timezone.now()
        property_obj.inferred_partido = values["partido"]
        property_obj.inferred_locality = values["locality"]
        property_obj.inferred_zone = values["zone"]
        property_obj.inferred_neighborhood = values["zone"]
        property_obj.territory_confidence = values["confidence"]
        property_obj.territory_source_method = values["source_method"]
        property_obj.territory_needs_review = conflict
        property_obj.territory_evidence = evidence
        property_obj.territory_inferred_at = now
        property_obj.zone_needs_review = conflict
        property_obj.zone_conflict = bool(conflict and source_zone and source_zone != values["zone"])
        property_obj.zone_inference_evidence = {
            **(property_obj.zone_inference_evidence or {}),
            SCRIPT_TAG: evidence,
        }
        property_obj.zone_inferred_at = now
        property_obj.save(
            update_fields=[
                "inferred_partido",
                "inferred_locality",
                "inferred_zone",
                "inferred_neighborhood",
                "territory_confidence",
                "territory_source_method",
                "territory_needs_review",
                "territory_evidence",
                "territory_inferred_at",
                "zone_needs_review",
                "zone_conflict",
                "zone_inference_evidence",
                "zone_inferred_at",
            ]
        )
        record, _ = PropertyLocationIntelligence.objects.get_or_create(property=property_obj)
        record.partido_name = values["partido"]
        record.locality_name = values["locality"]
        record.zone_name = values["zone"]
        record.match_method = PropertyLocationIntelligence.MatchMethod.COORDINATES
        record.confidence = values["confidence"] or record.confidence
        record.evidence = {**(record.evidence or {}), SCRIPT_TAG: evidence}
        record.scored_at = now
        record.save(
            update_fields=[
                "partido_name",
                "locality_name",
                "zone_name",
                "match_method",
                "confidence",
                "evidence",
                "scored_at",
            ]
        )
    return {
        "applied": True,
        "zone": values["zone"],
        "reason": "applied" if apply else "would_apply",
        "conflict": conflict,
        "values": values,
    }


def apply_manual_zone_if_needed(property_obj, row, summary, apply):
    manual_zone = normalized_zone(row.get("zona_actual")) or normalized_zone(row.get("barrio_actual"))
    if not manual_zone:
        return False
    if apply:
        now = timezone.now()
        evidence = {
            "curacion": SCRIPT_TAG,
            "reason": "zona/barrio aprobado sin coordenada util",
            "manual_feedback": True,
        }
        property_obj.inferred_partido = "Hurlingham"
        property_obj.inferred_locality = normalize_locality(row.get("localidad_actual")) or property_obj.locality or ""
        property_obj.inferred_zone = manual_zone
        property_obj.inferred_neighborhood = manual_zone
        property_obj.territory_confidence = "manual"
        property_obj.territory_source_method = "manual_feedback"
        property_obj.territory_needs_review = True
        property_obj.territory_evidence = evidence
        property_obj.territory_inferred_at = now
        property_obj.zone_needs_review = True
        property_obj.zone_conflict = False
        property_obj.zone_inference_evidence = {
            **(property_obj.zone_inference_evidence or {}),
            SCRIPT_TAG: evidence,
        }
        property_obj.zone_inferred_at = now
        property_obj.save(
            update_fields=[
                "inferred_partido",
                "inferred_locality",
                "inferred_zone",
                "inferred_neighborhood",
                "territory_confidence",
                "territory_source_method",
                "territory_needs_review",
                "territory_evidence",
                "territory_inferred_at",
                "zone_needs_review",
                "zone_conflict",
                "zone_inference_evidence",
                "zone_inferred_at",
            ]
        )
        record, _ = PropertyLocationIntelligence.objects.get_or_create(property=property_obj)
        record.partido_name = "Hurlingham"
        record.locality_name = property_obj.inferred_locality
        record.zone_name = manual_zone
        record.match_method = PropertyLocationIntelligence.MatchMethod.ZONE
        record.confidence = "manual"
        record.evidence = {**(record.evidence or {}), SCRIPT_TAG: evidence}
        record.scored_at = now
        record.save(update_fields=["partido_name", "locality_name", "zone_name", "match_method", "confidence", "evidence", "scored_at"])
    summary["manual_zone_applied_ids"].append({"id": property_obj.pk, "zone": manual_zone})
    return True


def geocode_approved_address(property_obj, row, summary, apply):
    if not best_address(property_obj):
        summary["geocode_no_address_ids"].append(property_obj.pk)
        return False
    geocoder = Geocoder()
    queries = geocoder.query_candidates(property_obj)
    if not queries:
        summary["geocode_no_query_ids"].append(property_obj.pk)
        return False
    cache = (
        GeocodeCache.objects.filter(query__in=queries)
        .exclude(latitude__isnull=True)
        .exclude(longitude__isnull=True)
        .first()
    )
    if cache:
        if apply:
            location = geocoder._apply(property_obj, cache.query, cache)
            property_obj.location = location
        else:
            property_obj.location = PropertyLocation(
                property=property_obj,
                latitude=cache.latitude,
                longitude=cache.longitude,
                precision=PropertyLocation.Precision.EXACT,
                query=cache.query,
                provider="nominatim",
                confidence=cache.confidence,
                outside_target=not in_target_bounds(cache.latitude, cache.longitude),
                manually_corrected=False,
            )
        summary["geocode_cache_ids"].append(property_obj.pk)
        return True
    summary["geocode_api_required_ids"].append(property_obj.pk)
    if not apply:
        return False
    for query in queries:
        if GeocodeCache.objects.filter(query=query).exists():
            continue
        try:
            cache = geocoder._fetch(query, property_obj)
            location = geocoder._apply(property_obj, query, cache)
        except Exception as exc:
            summary["geocode_errors"].append({"id": property_obj.pk, "query": query, "error": repr(exc)})
            continue
        if location:
            property_obj.location = location
            summary["geocode_api_ids"].append(property_obj.pk)
            return True
    summary["geocode_no_result_ids"].append(property_obj.pk)
    return False


def apply_approved_feedback(rows_by_id, approved_ids, fetcher, apply):
    now = timezone.now()
    summary = {
        "approved_ids": approved_ids,
        "address_updates": [],
        "locality_updates": [],
        "neighborhood_updates": [],
        "map_coordinates_ids": [],
        "map_no_coordinates_ids": [],
        "map_fetch_errors": [],
        "location_updates": [],
        "zone_applied_ids": [],
        "zone_conflict_ids": [],
        "zone_pending_ids": [],
        "manual_zone_applied_ids": [],
        "geocode_cache_ids": [],
        "geocode_api_required_ids": [],
        "geocode_api_ids": [],
        "geocode_no_result_ids": [],
        "geocode_no_query_ids": [],
        "geocode_no_address_ids": [],
        "geocode_errors": [],
        "manual_location_preserved_ids": [],
        "missing_property_ids": [],
    }
    for property_id in approved_ids:
        row = rows_by_id[property_id]
        try:
            property_obj = (
                Property.objects.select_related("location", "location_intelligence")
                .prefetch_related("listings__source", "listings__agency")
                .get(pk=property_id)
            )
        except Property.DoesNotExist:
            summary["missing_property_ids"].append(property_id)
            continue
        changes = set()
        before = {"address": property_obj.address, "locality": property_obj.locality, "neighborhood": property_obj.neighborhood}
        set_manual_field(property_obj, "address", row.get("domicilio_actual"), now, changes)
        set_manual_field(property_obj, "locality", row.get("localidad_actual"), now, changes)
        set_manual_field(property_obj, "neighborhood", row.get("barrio_actual"), now, changes)
        save_property(property_obj, changes, apply)
        if "address" in changes:
            summary["address_updates"].append({"id": property_id, "from": before["address"], "to": property_obj.address})
        if "locality" in changes:
            summary["locality_updates"].append({"id": property_id, "from": before["locality"], "to": property_obj.locality})
        if "neighborhood" in changes:
            summary["neighborhood_updates"].append({"id": property_id, "from": before["neighborhood"], "to": property_obj.neighborhood})

        listing = first_active_listing(property_obj)
        url = safe(row.get("url_publicacion")) or (listing.url if listing else "")
        source_slug = source_slug_from_url(url, listing.source.slug if listing else "")
        coordinate = None
        if url and source_slug in APPROVED_NETWORK_SOURCES:
            fetch = fetcher.fetch_map(url, source_slug)
            if fetch.get("error"):
                summary["map_fetch_errors"].append({"id": property_id, "source": source_slug, "error": fetch["error"]})
            allow_fallback = source_slug in {"riquelme", "paula-fossati"} or property_id in approved_ids
            coordinate = choose_coordinate(fetch.get("coordinates") or [], allow_fallback=allow_fallback)
            if coordinate:
                summary["map_coordinates_ids"].append({"id": property_id, "source": source_slug, "method": coordinate["method"], "lat": coordinate["latitude"], "lon": coordinate["longitude"]})
                if apply:
                    property_obj = Property.objects.select_related("location").get(pk=property_id)
                updated, reason = update_source_location(property_obj, coordinate, source_slug, url, summary, apply)
                if updated:
                    summary["location_updates"].append({"id": property_id, "source": source_slug, "reason": reason})
                    if apply:
                        property_obj = Property.objects.select_related("location").get(pk=property_id)
                    else:
                        property_obj.location = PropertyLocation(
                            property=property_obj,
                            latitude=coordinate["latitude"],
                            longitude=coordinate["longitude"],
                            precision=PropertyLocation.Precision.EXACT,
                            query=url,
                            provider=PROVIDER_BY_SOURCE.get(source_slug, f"{source_slug}_map"),
                            confidence=1.0,
                            outside_target=False,
                            manually_corrected=False,
                        )
            else:
                summary["map_no_coordinates_ids"].append({"id": property_id, "source": source_slug})

        if not coordinate and not location_or_none(property_obj):
            geocode_approved_address(property_obj, row, summary, apply)
            if apply:
                property_obj = Property.objects.select_related("location").get(pk=property_id)

        zone_result = apply_zone_from_location(
            property_obj,
            row=row,
            tag="approved_feedback_mapas",
            coordinate_evidence={"source": source_slug, "url": url, "coordinate": coordinate or {}},
            apply=apply,
        )
        if zone_result["applied"]:
            item = {"id": property_id, "zone": zone_result["zone"], "conflict": bool(zone_result.get("conflict"))}
            summary["zone_applied_ids"].append(item)
            if zone_result.get("conflict"):
                summary["zone_conflict_ids"].append(item)
        else:
            if not apply_manual_zone_if_needed(property_obj, row, summary, apply):
                summary["zone_pending_ids"].append({"id": property_id, "reason": zone_result["reason"]})
    return summary


def pending_properties():
    return (
        Property.objects.select_related("location", "location_intelligence")
        .prefetch_related("listings__source", "listings__agency")
        .filter(no_zone_query() | Q(address="") | Q(address__isnull=True))
        .distinct()
    )


def apply_existing_coordinate_sweep(apply):
    summary = {
        "candidate_ids": [],
        "detected_location_created_ids": [],
        "zone_applied_ids": [],
        "zone_conflict_ids": [],
        "zone_skipped_ids": [],
    }
    for property_obj in pending_properties():
        if not (
            location_or_none(property_obj)
            or (property_obj.detected_latitude is not None and property_obj.detected_longitude is not None)
        ):
            continue
        if not residual_reason(property_obj) or not (
            not property_obj.inferred_zone
            and not (intel_or_none(property_obj).zone_name if intel_or_none(property_obj) else "")
        ):
            continue
        summary["candidate_ids"].append(property_obj.pk)
        create_detected_location_if_needed(property_obj, summary, apply)
        if apply:
            property_obj = Property.objects.select_related("location", "location_intelligence").get(pk=property_obj.pk)
        zone_result = apply_zone_from_location(property_obj, tag="existing_coordinates_sweep", apply=apply)
        if zone_result["applied"]:
            item = {"id": property_obj.pk, "zone": zone_result["zone"], "conflict": bool(zone_result.get("conflict"))}
            summary["zone_applied_ids"].append(item)
            if zone_result.get("conflict"):
                summary["zone_conflict_ids"].append(item)
        else:
            summary["zone_skipped_ids"].append({"id": property_obj.pk, "reason": zone_result["reason"]})
    return summary


def apply_map_source_sweep(fetcher, approved_ids, apply):
    summary = {
        "candidate_ids_by_source": defaultdict(list),
        "map_coordinates_ids": [],
        "map_no_coordinates_ids": [],
        "map_fetch_errors": [],
        "location_updates": [],
        "manual_location_preserved_ids": [],
        "zone_applied_ids": [],
        "zone_conflict_ids": [],
        "zone_skipped_ids": [],
    }
    for property_obj in pending_properties():
        if property_obj.pk in approved_ids:
            continue
        sources = {listing.source.slug for listing in property_obj.listings.all()}
        target_sources = sorted(sources & MAP_SWEEP_SOURCES)
        if not target_sources:
            continue
        source_slug = target_sources[0]
        listing = first_active_listing(property_obj, preferred_sources=target_sources)
        if not listing:
            continue
        summary["candidate_ids_by_source"][source_slug].append(property_obj.pk)
        fetch = fetcher.fetch_map(listing.url, source_slug)
        if fetch.get("error"):
            summary["map_fetch_errors"].append({"id": property_obj.pk, "source": source_slug, "error": fetch["error"]})
        allow_fallback = source_slug in {"riquelme", "paula-fossati"}
        coordinate = choose_coordinate(fetch.get("coordinates") or [], allow_fallback=allow_fallback)
        if not coordinate:
            summary["map_no_coordinates_ids"].append({"id": property_obj.pk, "source": source_slug})
            continue
        summary["map_coordinates_ids"].append({"id": property_obj.pk, "source": source_slug, "method": coordinate["method"], "lat": coordinate["latitude"], "lon": coordinate["longitude"]})
        updated, reason = update_source_location(property_obj, coordinate, source_slug, listing.url, summary, apply)
        if updated:
            summary["location_updates"].append({"id": property_obj.pk, "source": source_slug, "reason": reason})
        if apply:
            property_obj = Property.objects.select_related("location", "location_intelligence").get(pk=property_obj.pk)
        else:
            property_obj.location = PropertyLocation(
                property=property_obj,
                latitude=coordinate["latitude"],
                longitude=coordinate["longitude"],
                precision=PropertyLocation.Precision.EXACT,
                query=listing.url,
                provider=PROVIDER_BY_SOURCE.get(source_slug, f"{source_slug}_map"),
                confidence=1.0,
                outside_target=False,
                manually_corrected=False,
            )
        zone_result = apply_zone_from_location(
            property_obj,
            tag="map_source_sweep",
            coordinate_evidence={"source": source_slug, "url": listing.url, "coordinate": coordinate},
            apply=apply,
        )
        if zone_result["applied"]:
            item = {"id": property_obj.pk, "zone": zone_result["zone"], "source": source_slug, "conflict": bool(zone_result.get("conflict"))}
            summary["zone_applied_ids"].append(item)
            if zone_result.get("conflict"):
                summary["zone_conflict_ids"].append(item)
        else:
            summary["zone_skipped_ids"].append({"id": property_obj.pk, "source": source_slug, "reason": zone_result["reason"]})
    summary["candidate_ids_by_source"] = dict(summary["candidate_ids_by_source"])
    return summary


def excel_number(value):
    if value is None:
        return None
    try:
        numeric = float(value)
    except (TypeError, ValueError):
        return None
    return int(numeric) if numeric.is_integer() else numeric


def generate_residual_excel(path):
    props = list(
        Property.objects.select_related("location", "location_intelligence")
        .prefetch_related("listings__source", "listings__agency")
        .all()
    )
    pending = [property_obj for property_obj in props if residual_reason(property_obj)]
    wb = Workbook()
    ws = wb.active
    ws.title = "Pendientes"
    headers = [
        "id_propiedad",
        "titulo",
        "descripcion_resumen",
        "domicilio_actual",
        "direccion_detectada",
        "direccion_normalizada",
        "localidad_actual",
        "localidad_detectada",
        "barrio_actual",
        "barrio_detectado",
        "zona_actual",
        "zona_inteligencia",
        "latitud",
        "longitud",
        "fuente_coordenadas",
        "fuente",
        "inmobiliaria",
        "url_publicacion",
        "motivo_url",
        "estado_publicacion",
        "tipo_propiedad",
        "operacion",
        "moneda",
        "precio",
        "ambientes",
        "dormitorios",
        "banos",
        "superficie_cubierta",
        "superficie_total",
        "fecha_ultima_vista",
        "motivo_pendiente",
        "evidencia",
        "decision_manual",
        "notas_manual",
    ]
    ws.append(headers)
    for property_obj in sorted(pending, key=lambda prop: (residual_reason(prop), prop.pk)):
        listing = first_active_listing(property_obj)
        location = location_or_none(property_obj)
        intel = intel_or_none(property_obj)
        evidence = {
            "location_precision": location.precision if location else "",
            "location_query": location.query if location else "",
            "manual_overrides": property_obj.manual_overrides or {},
            "source_locality": property_obj.locality or property_obj.detected_locality or "",
            "source_zone": property_obj.neighborhood or property_obj.detected_neighborhood or "",
            "territory_evidence": property_obj.territory_evidence or {},
            "location_evidence": property_obj.location_evidence or {},
        }
        ws.append(
            [
                property_obj.pk,
                property_obj.title,
                normalize_whitespace(property_obj.description or "")[:500],
                property_obj.address,
                property_obj.detected_address,
                property_obj.normalized_address,
                property_obj.locality,
                property_obj.detected_locality,
                property_obj.neighborhood,
                property_obj.detected_neighborhood,
                property_obj.inferred_zone or property_obj.inferred_neighborhood,
                intel.zone_name if intel else "",
                location.latitude if location else property_obj.detected_latitude,
                location.longitude if location else property_obj.detected_longitude,
                location.provider if location else ("detected_coordinates" if property_obj.detected_latitude is not None else ""),
                source_names(property_obj),
                agency_names(property_obj),
                listing.url if listing else "",
                "" if listing else "Sin Listing asociado en la base",
                property_obj.get_status_display(),
                property_obj.get_property_type_display(),
                property_obj.operation,
                property_obj.currency,
                excel_number(property_obj.price),
                property_obj.rooms,
                property_obj.bedrooms,
                excel_number(property_obj.bathrooms),
                excel_number(property_obj.covered_area),
                excel_number(property_obj.total_area),
                property_obj.last_seen_at.isoformat() if property_obj.last_seen_at else "",
                residual_reason(property_obj),
                json.dumps(evidence, ensure_ascii=False, default=str),
                "",
                "",
            ]
        )
    ws.auto_filter.ref = ws.dimensions
    if ws.max_row > 1:
        table = Table(displayName="PendientesMapasTable", ref=ws.dimensions)
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        ws.add_table(table)
    ws.freeze_panes = "A2"
    header_fill = PatternFill("solid", fgColor="1F4E78")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    price_col = headers.index("precio") + 1
    decision_col = headers.index("decision_manual") + 1
    for row_idx in range(2, ws.max_row + 1):
        ws.cell(row=row_idx, column=price_col).number_format = "#,##0.00"
    if ws.max_row > 1:
        decision_letter = get_column_letter(decision_col)
        validation = DataValidation(
            type="list",
            formula1='"APROBADO,RECHAZADO"',
            allow_blank=True,
            showErrorMessage=True,
            errorTitle="Valor no valido",
            error="Usa APROBADO o RECHAZADO.",
        )
        ws.add_data_validation(validation)
        validation.add(f"{decision_letter}2:{decision_letter}{ws.max_row}")
    for column in range(1, ws.max_column + 1):
        letter = get_column_letter(column)
        max_len = max(len(str(ws.cell(row=row, column=column).value or "")) for row in range(1, min(ws.max_row, 80) + 1))
        ws.column_dimensions[letter].width = min(max(max_len + 2, 12), 52)

    summary = wb.create_sheet("Resumen")
    counts = current_counts()
    for row in [
        ("fecha", timezone.now().isoformat()),
        ("pendientes_total", len(pending)),
        ("sin_zona_operativa", counts["sin_zona_operativa"]),
        ("sin_direccion_util", counts["sin_direccion_util"]),
        ("miglierini_pendientes", counts["miglierini_pendientes"]),
        ("archivo_origen", str(INPUT_XLSX)),
    ]:
        summary.append(row)
    for cell in summary[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
    summary.freeze_panes = "A2"
    wb.save(path)
    return {"path": str(path), "pending_rows": len(pending), "columns": headers}


def validate_excel(path):
    wb = load_workbook(path, read_only=False, data_only=False)
    ws = wb["Pendientes"]
    headers = [cell.value for cell in ws[1]]
    price_col = headers.index("precio") + 1
    decision_col = headers.index("decision_manual") + 1
    notes_col = headers.index("notas_manual") + 1
    price_types = Counter()
    nonblank_decisions = 0
    nonblank_notes = 0
    for row_idx in range(2, ws.max_row + 1):
        price_types[type(ws.cell(row=row_idx, column=price_col).value).__name__] += 1
        if safe(ws.cell(row=row_idx, column=decision_col).value):
            nonblank_decisions += 1
        if safe(ws.cell(row=row_idx, column=notes_col).value):
            nonblank_notes += 1
    return {
        "sheets": wb.sheetnames,
        "rows": ws.max_row - 1,
        "auto_filter": ws.auto_filter.ref,
        "tables": list(ws.tables),
        "freeze_panes": ws.freeze_panes,
        "price_types": dict(price_types),
        "nonblank_decisions": nonblank_decisions,
        "nonblank_notes": nonblank_notes,
        "validations": [
            {"type": dv.type, "formula1": dv.formula1, "sqref": str(dv.sqref)}
            for dv in ws.data_validations.dataValidation
        ],
        "missing_columns": sorted(
            {"motivo_url", "decision_manual", "notas_manual", "precio"} - set(headers)
        ),
    }


def sample_state(ids):
    output = []
    for property_obj in (
        Property.objects.filter(pk__in=ids)
        .select_related("location", "location_intelligence")
        .prefetch_related("listings__source")
        .order_by("pk")
    ):
        location = location_or_none(property_obj)
        intel = intel_or_none(property_obj)
        output.append(
            {
                "id": property_obj.pk,
                "sources": sorted({listing.source.slug for listing in property_obj.listings.all()}),
                "address": property_obj.address,
                "locality": property_obj.locality,
                "neighborhood": property_obj.neighborhood,
                "zone": property_obj.inferred_zone,
                "intel_zone": intel.zone_name if intel else "",
                "zone_conflict": property_obj.zone_conflict,
                "zone_needs_review": property_obj.zone_needs_review,
                "location": None
                if not location
                else {
                    "lat": location.latitude,
                    "lon": location.longitude,
                    "provider": location.provider,
                    "manual": location.manually_corrected,
                },
                "manual_overrides": property_obj.manual_overrides or {},
            }
        )
    return output


def self_test():
    argenprop = '<div data-location-map data-latitude="-34,62549" data-longitude="-58,63528"></div>'
    wordpress = '<script>propertyMapData = {"lat":"-34.5704212","lang":"-58.6391637"};</script>'
    out1 = extract_map_coordinates(argenprop)
    out2 = extract_map_coordinates(wordpress)
    assert out1 and out1[0]["method"] == "data-latitude"
    assert round(out1[0]["latitude"], 5) == -34.62549
    assert round(out1[0]["longitude"], 5) == -58.63528
    assert out2 and out2[0]["method"] == "propertyMapData"
    assert round(out2[0]["latitude"], 7) == -34.5704212
    assert round(out2[0]["longitude"], 7) == -58.6391637
    assert not extract_map_coordinates('<div data-latitude="25.7308309" data-longitude="-80.444149"></div>')
    return {"self_test": "ok"}


def json_safe_summary(payload):
    def convert(value):
        if isinstance(value, defaultdict):
            return {key: convert(item) for key, item in value.items()}
        if isinstance(value, Counter):
            return dict(value)
        if isinstance(value, dict):
            return {key: convert(item) for key, item in value.items()}
        if isinstance(value, list):
            return [convert(item) for item in value]
        return value

    return convert(payload)


def write_json(path, payload):
    Path(path).write_text(json.dumps(json_safe_summary(payload), ensure_ascii=False, indent=2, default=str), encoding="utf-8")


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--apply", action="store_true")
    parser.add_argument("--generate-excel", action="store_true")
    parser.add_argument("--refresh-fetch-cache", action="store_true")
    parser.add_argument("--self-test", action="store_true")
    parser.add_argument("--json-out")
    args = parser.parse_args()

    if args.self_test:
        result = self_test()
        print(json.dumps(result, ensure_ascii=False, indent=2))
        return

    rows_by_id, decisions, notes, approved_ids = load_feedback_rows()
    before = current_counts()
    fetcher = Fetcher(refresh=args.refresh_fetch_cache)
    if args.apply and not BACKUP_PATH.exists():
        shutil.copy2(DB_PATH, BACKUP_PATH)

    with transaction.atomic():
        approved = apply_approved_feedback(rows_by_id, approved_ids, fetcher, args.apply)
        existing_coords = apply_existing_coordinate_sweep(args.apply)
        map_sweep = apply_map_source_sweep(fetcher, set(approved_ids), args.apply)
        if not args.apply:
            transaction.set_rollback(True)

    excel = None
    excel_validation = None
    if args.generate_excel:
        excel = generate_residual_excel(OUTPUT_XLSX)
        excel_validation = validate_excel(OUTPUT_XLSX)

    after = current_counts()
    result = {
        "mode": "apply" if args.apply else "dry_run",
        "input_excel": str(INPUT_XLSX),
        "output_excel": str(OUTPUT_XLSX) if args.generate_excel else "",
        "backup": str(BACKUP_PATH) if args.apply else "",
        "approved_expected_ids": sorted(APPROVED_IDS_EXPECTED),
        "approved_found_ids": approved_ids,
        "approved_missing_expected": sorted(APPROVED_IDS_EXPECTED - set(approved_ids)),
        "before": before,
        "after": after,
        "approved": approved,
        "existing_coordinates": existing_coords,
        "map_sweep": map_sweep,
        "notes_ids": sorted(notes),
        "sample": sample_state(approved_ids),
        "excel": excel,
        "excel_validation": excel_validation,
    }
    if args.json_out:
        write_json(args.json_out, result)
    print(json.dumps(json_safe_summary(result), ensure_ascii=False, indent=2, default=str))


if __name__ == "__main__":
    main()
