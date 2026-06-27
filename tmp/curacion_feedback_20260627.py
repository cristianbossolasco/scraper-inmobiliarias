import argparse
import json
import os
import re
import shutil
import sys
import time
from collections import Counter, defaultdict
from pathlib import Path

import requests
from bs4 import BeautifulSoup
from django.db.models import Q
from django.utils import timezone
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))
os.environ.setdefault("DJANGO_SETTINGS_MODULE", "config.settings")

import django

django.setup()

from properties.models import Listing, Property, PropertyLocation, PropertyLocationIntelligence
from properties.scrapers.argencasas import ArgencasasScraper
from properties.scrapers.local_sites import BecerraScraper
from properties.scrapers.pending_sources import GuarnieriScraper
from properties.scrapers.parsing import extract_map_coordinates
from properties.services.geocoding import Geocoder, best_address
from properties.services.normalization import (
    canonical_address_alias,
    normalize_address,
    normalize_whitespace,
)
from properties.services.territory_hierarchy import infer_territory_for_point


INPUT_XLSX = ROOT / "tmp" / "curacion_zonas_pendientes_2026-06-26_feedback.xlsx"
OUTPUT_XLSX = ROOT / "tmp" / "curacion_zonas_pendientes_2026-06-27_feedback.xlsx"
BACKUP_PATH = ROOT / "tmp" / "db.sqlite3.backup_curacion_feedback_20260627"
DB_PATH = ROOT / "db.sqlite3"
SCRIPT_TAG = "curacion_feedback_20260627"

APPROVED_IDS = {
    1131,
    1881,
    1942,
    2099,
    2101,
    2254,
    2311,
    2492,
    2596,
    2875,
    2879,
    2880,
    2883,
    2886,
    2898,
    2901,
    3181,
    3284,
    3289,
    4454,
    5829,
}
NO_ACTION_IDS = {2875, 2879, 2883, 2886, 3301, 3330, 3349, 3402, 3503}
CONFIRM_REMOVE_IDS = {1942}
SAMPLE_IDS = sorted(APPROVED_IDS | NO_ACTION_IDS)

DIRECT_FIXES = {
    1131: {"address": "Gral. Toribio de Luzuriaga 1700", "locality": "Hurlingham", "geocode": True},
    1881: {
        "address": "Tte. Gral. Pablo Ricchieri (el mirador) 1500",
        "locality": "Hurlingham",
        "latitude": -34.587651022465,
        "longitude": -58.636475811363,
        "provider": "argencasas_map",
    },
    2099: {"address": "Esteban De Luca 100", "locality": "Hurlingham", "geocode": True},
    2101: {"address": "Dip. Hector Finochietto 1900", "locality": "Hurlingham", "geocode": True},
    2254: {"address": "Schubert 2400", "locality": "Hurlingham", "geocode": True},
    2311: {"address": "Esteban de Luca 100", "locality": "Hurlingham", "geocode": True},
    2492: {"address": "Eva Peron 100", "locality": "Hurlingham", "geocode": True},
    2596: {"address": "Gral. Martin Guemes 1400", "locality": "Hurlingham", "geocode": True},
    2880: {"address": "Av. Rosas Castillo 2900", "locality": "Hurlingham", "geocode": True},
    2898: {
        "address": "Jose de Andonaegui 1600",
        "locality": "Hurlingham",
        "latitude": -34.601427,
        "longitude": -58.6495483,
        "provider": "becerra_map",
    },
    2901: {
        "address": "Juan Diaz de Solis 1500",
        "locality": "Hurlingham",
        "latitude": -34.5888949406,
        "longitude": -58.639576753,
        "provider": "becerra_map",
    },
    3181: {"address": "Virriato Unia 2412", "locality": "Hurlingham", "geocode": True},
    3284: {"address": "Gutenberg 2100", "locality": "William C. Morris", "geocode": True},
    3289: {"address": "Gral. Alfredo Rodriguez 1635", "locality": "Hurlingham", "geocode": True},
    4454: {
        "latitude": -34.590452874017,
        "longitude": -58.644783496857,
        "provider": "guarnieri_map",
    },
    5829: {"address": "Schumann 1367", "locality": "Hurlingham", "geocode": True},
}

HEADERS = [
    "id_propiedad",
    "titulo",
    "descripcion_resumen",
    "domicilio_actual",
    "direccion_detectada",
    "direccion_sugerida",
    "direccion_normalizada",
    "localidad_actual",
    "localidad_detectada",
    "barrio_actual",
    "barrio_detectado",
    "zona_actual",
    "zona_sugerida",
    "zona_inteligencia",
    "confianza",
    "motivo",
    "estado",
    "latitud",
    "longitud",
    "fuente_coordenadas",
    "fuente",
    "inmobiliaria",
    "url_publicacion",
    "motivo_url",
    "estado_publicacion",
    "tipo_propiedad",
    "operacion",
    "moneda",
    "precio",
    "ambientes",
    "dormitorios",
    "banos",
    "superficie_cubierta",
    "superficie_total",
    "fecha_ultima_vista",
    "motivo_pendiente",
    "evidencia",
    "decision_manual",
    "notas_manual",
]

SOURCE_DELAYS = {"argencasas": 2, "becerra": 3, "guarnieri": 3}
FETCH_STATE = {}


def safe(value):
    return normalize_whitespace(str(value)) if value not in (None, "") else ""


def no_zone_query():
    return (
        (Q(inferred_zone__isnull=True) | Q(inferred_zone=""))
        & (
            Q(location_intelligence__isnull=True)
            | Q(location_intelligence__zone_name__isnull=True)
            | Q(location_intelligence__zone_name="")
        )
    )


def location_or_none(property_obj):
    return getattr(property_obj, "location", None)


def intel_or_none(property_obj):
    return getattr(property_obj, "location_intelligence", None)


def curation_no_action(property_obj):
    evidence = property_obj.zone_inference_evidence or {}
    return bool((evidence.get("curation_no_action") or {}).get(SCRIPT_TAG))


def residual_reason(property_obj):
    reasons = []
    intel = intel_or_none(property_obj)
    if not property_obj.inferred_zone and not (intel.zone_name if intel else ""):
        reasons.append("sin zona")
    if not best_address(property_obj):
        reasons.append("sin direccion util")
    return ", ".join(reasons)


def current_counts():
    props = list(
        Property.objects.select_related("location", "location_intelligence").prefetch_related(
            "listings__source"
        )
    )
    by_source = defaultdict(lambda: Counter())
    for property_obj in props:
        if property_obj.is_hidden or property_obj.status == Property.Status.REMOVED or curation_no_action(property_obj):
            continue
        reason = residual_reason(property_obj)
        if not reason:
            continue
        sources = {listing.source.slug for listing in property_obj.listings.all()} or {"sin_listing"}
        for source in sources:
            by_source[source]["pendientes"] += 1
            if "sin zona" in reason:
                by_source[source]["sin_zona"] += 1
            if "sin direccion" in reason:
                by_source[source]["sin_direccion"] += 1
    no_zone = [
        prop
        for prop in props
        if not prop.is_hidden
        and prop.status != Property.Status.REMOVED
        and not curation_no_action(prop)
        and residual_reason(prop)
        and "sin zona" in residual_reason(prop)
    ]
    no_address = [
        prop
        for prop in props
        if not prop.is_hidden
        and prop.status != Property.Status.REMOVED
        and not curation_no_action(prop)
        and not best_address(prop)
    ]
    return {
        "sin_zona_operativa": len(no_zone),
        "sin_direccion_util": len(no_address),
        "pendientes_por_fuente": dict(sorted((source, dict(counts)) for source, counts in by_source.items())),
    }


def load_feedback():
    wb = load_workbook(INPUT_XLSX, read_only=True, data_only=False)
    ws = wb["Pendientes"]
    headers = [cell.value for cell in ws[1]]
    rows = {}
    decisions = Counter()
    notes = {}
    for values in ws.iter_rows(min_row=2, values_only=True):
        row = dict(zip(headers, values))
        try:
            property_id = int(row.get("id_propiedad") or 0)
        except (TypeError, ValueError):
            continue
        rows[property_id] = row
        decision = safe(row.get("decision_manual")).upper()
        if decision:
            decisions[decision] += 1
        note = safe(row.get("notas_manual"))
        if note:
            notes[property_id] = note
    return rows, decisions, notes


def add_summary(summary, key, item):
    summary.setdefault(key, []).append(item)


def first_listing(property_obj):
    return property_obj.listings.select_related("source", "agency").order_by("-active", "-last_seen_at").first()


def first_listing_url(property_obj):
    listing = first_listing(property_obj)
    return listing.url if listing else ""


def mark_manual_property_field(property_obj, field, value, now, update_fields, apply):
    value = safe(value)
    if not value:
        return
    if getattr(property_obj, field) != value:
        setattr(property_obj, field, value)
        update_fields.add(field)
    if field == "address":
        normalized = normalize_address(value)
        if property_obj.normalized_address != normalized:
            property_obj.normalized_address = normalized
            update_fields.add("normalized_address")
        current = location_or_none(property_obj)
        if current and not current.manually_corrected and apply:
            current.delete()
    overrides = dict(property_obj.manual_overrides or {})
    overrides[field] = now.isoformat()
    property_obj.manual_overrides = overrides
    property_obj.data_manually_corrected_at = now
    update_fields.update({"manual_overrides", "data_manually_corrected_at"})


def save_property(property_obj, update_fields, apply):
    if update_fields and apply:
        property_obj.save(update_fields=sorted(update_fields))


def update_location(property_obj, latitude, longitude, provider, query, summary, apply):
    current = location_or_none(property_obj)
    if current and current.manually_corrected:
        add_summary(summary, "manual_location_preserved", property_obj.pk)
        return False
    defaults = {
        "latitude": float(latitude),
        "longitude": float(longitude),
        "precision": PropertyLocation.Precision.EXACT,
        "query": query or "",
        "provider": provider,
        "confidence": 1.0,
        "outside_target": False,
        "manually_corrected": False,
    }
    if apply:
        PropertyLocation.objects.update_or_create(property=property_obj, defaults=defaults)
        evidence = dict(property_obj.location_evidence or {})
        evidence[SCRIPT_TAG] = {
            "provider": provider,
            "latitude": float(latitude),
            "longitude": float(longitude),
            "query": query or "",
        }
        property_obj.detected_latitude = float(latitude)
        property_obj.detected_longitude = float(longitude)
        property_obj.location_source = Property.LocationSource.MAP
        property_obj.location_confidence = Property.LocationConfidence.HIGH
        property_obj.location_evidence = evidence
        property_obj.save(
            update_fields=[
                "detected_latitude",
                "detected_longitude",
                "location_source",
                "location_confidence",
                "location_evidence",
            ]
        )
    return True


def apply_zone_from_point(property_obj, latitude, longitude, provider, reason, summary, apply):
    result = infer_territory_for_point(latitude, longitude)
    zone = safe(result.zone if result else "")
    if not zone:
        add_summary(summary, "zone_skipped", {"id": property_obj.pk, "reason": "sin_match_poligono"})
        return ""
    now = timezone.now()
    evidence = dict(result.evidence if result and result.evidence else {})
    evidence.update({"provider": provider, "curacion": SCRIPT_TAG, "reason": reason})
    if apply:
        property_obj.inferred_partido = result.partido
        property_obj.inferred_locality = result.locality
        property_obj.inferred_zone = zone
        property_obj.inferred_neighborhood = zone
        property_obj.territory_confidence = result.confidence
        property_obj.territory_source_method = result.source_method
        property_obj.territory_needs_review = bool(result.needs_review)
        property_obj.territory_evidence = evidence
        property_obj.territory_inferred_at = now
        property_obj.zone_conflict = False
        property_obj.zone_needs_review = bool(result.needs_review)
        property_obj.zone_inference_evidence = evidence
        property_obj.zone_inferred_at = now
        property_obj.save(
            update_fields=[
                "inferred_partido",
                "inferred_locality",
                "inferred_zone",
                "inferred_neighborhood",
                "territory_confidence",
                "territory_source_method",
                "territory_needs_review",
                "territory_evidence",
                "territory_inferred_at",
                "zone_conflict",
                "zone_needs_review",
                "zone_inference_evidence",
                "zone_inferred_at",
            ]
        )
        PropertyLocationIntelligence.objects.update_or_create(
            property=property_obj,
            defaults={
                "partido_name": result.partido,
                "locality_name": result.locality,
                "zone_name": zone,
                "match_method": PropertyLocationIntelligence.MatchMethod.COORDINATES,
                "confidence": result.confidence,
            },
        )
    add_summary(summary, "zone_applied", {"id": property_obj.pk, "zone": zone, "provider": provider})
    return zone


def mark_removed(property_obj, reason, summary, apply):
    listing_ids = list(property_obj.listings.filter(active=True).values_list("id", flat=True))
    if apply:
        property_obj.listings.filter(active=True).update(active=False, source_status="removed", missing_runs=2)
        marker = f"{SCRIPT_TAG}: {reason}"
        notes = property_obj.personal_notes or ""
        if marker not in notes:
            property_obj.personal_notes = "\n\n".join(part for part in [notes, marker] if part)
        property_obj.status = Property.Status.REMOVED
        property_obj.is_hidden = True
        property_obj.save(update_fields=["status", "is_hidden", "personal_notes"])
    add_summary(summary, "removed", {"id": property_obj.pk, "listing_ids": listing_ids, "reason": reason})


def mark_reviewed_no_action(property_obj, note, summary, apply):
    if apply:
        evidence = dict(property_obj.zone_inference_evidence or {})
        no_action = dict(evidence.get("curation_no_action") or {})
        no_action[SCRIPT_TAG] = {"note": note or "sin datos utiles publicados", "reviewed_at": timezone.now().isoformat()}
        evidence["curation_no_action"] = no_action
        property_obj.zone_inference_evidence = evidence
        property_obj.reviewed_at = timezone.now()
        property_obj.save(update_fields=["zone_inference_evidence", "reviewed_at"])
    add_summary(summary, "reviewed_no_action", {"id": property_obj.pk, "note": note})


def fetch_url_text(url, source_slug=""):
    delay = SOURCE_DELAYS.get(source_slug, 1)
    elapsed = time.monotonic() - FETCH_STATE.get(source_slug, 0)
    if elapsed < delay:
        time.sleep(delay - elapsed)
    session = requests.Session()
    session.headers.update({"User-Agent": "Mozilla/5.0 (compatible; RadarCuration/1.0)"})
    response = session.get(url, timeout=35)
    FETCH_STATE[source_slug] = time.monotonic()
    return response.status_code, response.text or ""


def parse_with_scraper(scraper_cls, url, markup):
    scraper = scraper_cls()
    soup = BeautifulSoup(markup or "", "lxml")
    scraper.soup = lambda _url: soup
    return scraper.parse(url)


def has_removed_marker(status_code, markup):
    return status_code == 404 and re.search(
        r"propiedad\s+ha\s+sido\s+retirada\s+del\s+sistema|publicaci[oó]n\s+retirada",
        markup or "",
        re.I,
    )


def confirm_argencasas_removed(property_obj, summary, apply):
    listing = first_listing(property_obj)
    if not listing or not listing.url:
        add_summary(summary, "remove_skipped", {"id": property_obj.pk, "reason": "sin_url"})
        return
    try:
        status_code, markup = fetch_url_text(listing.url, "argencasas")
    except Exception as exc:
        add_summary(summary, "remove_skipped", {"id": property_obj.pk, "reason": repr(exc)})
        return
    if has_removed_marker(status_code, markup):
        mark_removed(property_obj, "Argencasas 404 + propiedad retirada del sistema", summary, apply)
    else:
        add_summary(summary, "remove_skipped", {"id": property_obj.pk, "http_status": status_code})


def apply_geocode(property_obj, reason, summary, apply):
    add_summary(summary, "geocode_required", property_obj.pk)
    if not apply:
        return
    property_obj.refresh_from_db()
    geocoder = Geocoder()
    location = geocoder.geocode_property_from_cache(property_obj, force=True)
    if not location:
        try:
            location = geocoder.geocode_property(property_obj, force=True)
        except Exception as exc:
            add_summary(summary, "geocode_errors", {"id": property_obj.pk, "error": repr(exc)})
            return
    if location and not location.outside_target:
        add_summary(
            summary,
            "geocode_location",
            {
                "id": property_obj.pk,
                "lat": location.latitude,
                "lon": location.longitude,
                "provider": location.provider,
            },
        )
        apply_zone_from_point(property_obj, location.latitude, location.longitude, location.provider, reason, summary, apply)
    elif location:
        add_summary(summary, "geocode_outside_target", {"id": property_obj.pk, "lat": location.latitude, "lon": location.longitude})


def apply_direct_fixes(summary, apply):
    now = timezone.now()
    for property_id, fix in DIRECT_FIXES.items():
        property_obj = Property.objects.select_related("location", "location_intelligence").filter(pk=property_id).first()
        if not property_obj:
            add_summary(summary, "missing_property", property_id)
            continue
        update_fields = set()
        for field in ("address", "locality", "neighborhood"):
            if fix.get(field):
                mark_manual_property_field(property_obj, field, fix[field], now, update_fields, apply)
                add_summary(summary, f"{field}_updates", {"id": property_id, "to": fix[field]})
        save_property(property_obj, update_fields, apply)
        if fix.get("latitude") is not None and fix.get("longitude") is not None:
            if update_location(
                property_obj,
                fix["latitude"],
                fix["longitude"],
                fix.get("provider") or "curation_map",
                first_listing_url(property_obj),
                summary,
                apply,
            ):
                add_summary(summary, "location_updates", {"id": property_id, "provider": fix.get("provider")})
            apply_zone_from_point(
                property_obj,
                fix["latitude"],
                fix["longitude"],
                fix.get("provider") or "curation_map",
                "feedback aprobado + fuente mapa/geocoding",
                summary,
                apply,
            )
        elif fix.get("geocode"):
            apply_geocode(property_obj, "feedback aprobado + direccion curada", summary, apply)


def canonical_propagated_address(value):
    text = safe(value)
    if not text:
        return ""
    bustamante = re.search(r"\bBustamante\s+(\d{2,5})\b", text, re.I)
    if bustamante:
        return f"Eva Peron {bustamante.group(1)}"
    return canonical_address_alias(text)


def apply_bustamante_propagation(summary, apply):
    qs = (
        Property.objects.select_related("location")
        .filter(status=Property.Status.ACTIVE, is_hidden=False)
        .filter(Q(address__icontains="Bustamante") | Q(detected_address__icontains="Bustamante"))
        .filter(location__isnull=True)
        .exclude(pk__in=DIRECT_FIXES)
        .distinct()
        .order_by("id")
    )
    now = timezone.now()
    for property_obj in qs:
        manual = property_obj.manual_overrides or {}
        if "address" in manual:
            continue
        original = property_obj.address or property_obj.detected_address
        candidate = canonical_propagated_address(original)
        if not candidate or candidate == original or not re.search(r"\d{2,5}", candidate):
            continue
        update_fields = set()
        mark_manual_property_field(property_obj, "address", candidate, now, update_fields, apply)
        save_property(property_obj, update_fields, apply)
        add_summary(summary, "propagated_address_updates", {"id": property_obj.pk, "from": original, "to": candidate, "rule": "Bustamante -> Eva Peron"})
        apply_geocode(property_obj, "propagacion segura Bustamante -> Eva Peron", summary, apply)


def source_pending_queryset(source_slug):
    return (
        Property.objects.select_related("location", "location_intelligence")
        .prefetch_related("listings__source")
        .filter(listings__source__slug=source_slug, status=Property.Status.ACTIVE, is_hidden=False)
        .filter(no_zone_query() | Q(location__isnull=True) | Q(address="") | Q(address__isnull=True))
        .distinct()
        .order_by("id")
    )


def sweep_argencasas(summary, apply):
    for property_obj in source_pending_queryset("argencasas"):
        if property_obj.pk in DIRECT_FIXES or property_obj.pk in CONFIRM_REMOVE_IDS:
            continue
        listing = first_listing(property_obj)
        if not listing or not listing.url:
            continue
        try:
            status_code, markup = fetch_url_text(listing.url, "argencasas")
        except Exception as exc:
            add_summary(summary, "source_fetch_errors", {"id": property_obj.pk, "source": "argencasas", "error": repr(exc)})
            continue
        if has_removed_marker(status_code, markup):
            mark_removed(property_obj, "Argencasas 404 + propiedad retirada del sistema en barrido", summary, apply)
            continue
        coordinates = extract_map_coordinates(markup, require_target_bounds=True)
        if coordinates:
            coordinate = coordinates[0]
            if update_location(property_obj, coordinate["latitude"], coordinate["longitude"], "argencasas_map", listing.url, summary, apply):
                add_summary(summary, "sweep_map_locations", {"id": property_obj.pk, "source": "argencasas", "method": coordinate["method"]})
            apply_zone_from_point(
                property_obj,
                coordinate["latitude"],
                coordinate["longitude"],
                "argencasas_map",
                "barrido Argencasas con coordenada fuente",
                summary,
                apply,
            )


def sweep_becerra(summary, apply):
    for property_obj in source_pending_queryset("becerra"):
        if property_obj.pk in DIRECT_FIXES:
            continue
        listing = first_listing(property_obj)
        if not listing or not listing.url:
            continue
        try:
            status_code, markup = fetch_url_text(listing.url, "becerra")
        except Exception as exc:
            add_summary(summary, "source_fetch_errors", {"id": property_obj.pk, "source": "becerra", "error": repr(exc)})
            continue
        if status_code >= 400:
            add_summary(summary, "source_fetch_errors", {"id": property_obj.pk, "source": "becerra", "http_status": status_code})
            continue
        try:
            data = parse_with_scraper(BecerraScraper, listing.url, markup)
        except Exception as exc:
            add_summary(summary, "source_parse_errors", {"id": property_obj.pk, "source": "becerra", "error": repr(exc)})
            continue
        if data.get("latitude") is not None and data.get("longitude") is not None:
            if update_location(property_obj, data["latitude"], data["longitude"], "becerra_map", listing.url, summary, apply):
                add_summary(summary, "sweep_map_locations", {"id": property_obj.pk, "source": "becerra"})
            apply_zone_from_point(
                property_obj,
                data["latitude"],
                data["longitude"],
                "becerra_map",
                "barrido Becerra con coordenada fuente",
                summary,
                apply,
            )


def sweep_guarnieri(summary, apply):
    for property_obj in source_pending_queryset("guarnieri"):
        if property_obj.pk in DIRECT_FIXES:
            continue
        listing = first_listing(property_obj)
        if not listing or not listing.url:
            continue
        try:
            status_code, markup = fetch_url_text(listing.url, "guarnieri")
        except Exception as exc:
            add_summary(summary, "source_fetch_errors", {"id": property_obj.pk, "source": "guarnieri", "error": repr(exc)})
            continue
        if status_code >= 400:
            add_summary(summary, "source_fetch_errors", {"id": property_obj.pk, "source": "guarnieri", "http_status": status_code})
            continue
        coordinates = extract_map_coordinates(markup, require_target_bounds=True)
        if coordinates:
            coordinate = coordinates[0]
            if update_location(property_obj, coordinate["latitude"], coordinate["longitude"], "guarnieri_map", listing.url, summary, apply):
                add_summary(summary, "sweep_map_locations", {"id": property_obj.pk, "source": "guarnieri", "method": coordinate["method"]})
            apply_zone_from_point(
                property_obj,
                coordinate["latitude"],
                coordinate["longitude"],
                "guarnieri_map",
                "barrido Guarnieri con coordenada fuente",
                summary,
                apply,
            )


def source_names(property_obj):
    names = []
    for listing in property_obj.listings.select_related("source").all():
        if listing.source.name not in names:
            names.append(listing.source.name)
    return ", ".join(names)


def agency_names(property_obj):
    names = []
    for listing in property_obj.listings.select_related("agency").all():
        if listing.agency and listing.agency.name not in names:
            names.append(listing.agency.name)
    return ", ".join(names)


def status_label(status):
    return dict(Property.Status.choices).get(status, status)


def residual_queryset():
    rows = (
        Property.objects.select_related("location", "location_intelligence")
        .prefetch_related("listings__source", "listings__agency")
        .filter(no_zone_query() | Q(address="") | Q(address__isnull=True))
        .filter(is_hidden=False)
        .exclude(status=Property.Status.REMOVED)
        .distinct()
        .order_by("id")
    )
    return [property_obj for property_obj in rows if not curation_no_action(property_obj)]


def property_row(property_obj, reviewed_notes):
    location = location_or_none(property_obj)
    intel = intel_or_none(property_obj)
    listing = first_listing(property_obj)
    url = listing.url if listing else ""
    evidence = {
        "location_precision": location.precision if location else "",
        "location_query": location.query if location else "",
        "manual_overrides": property_obj.manual_overrides or {},
        "source_locality": property_obj.locality,
        "source_zone": property_obj.inferred_zone,
        "territory_evidence": property_obj.territory_evidence or {},
    }
    if property_obj.pk in reviewed_notes:
        evidence["feedback_20260627"] = reviewed_notes[property_obj.pk] or "revisado manualmente"
    return [
        property_obj.pk,
        property_obj.title,
        (property_obj.description or "")[:500],
        property_obj.address or "",
        property_obj.detected_address or "",
        "",
        property_obj.normalized_address or "",
        property_obj.locality or "",
        property_obj.detected_locality or "",
        property_obj.neighborhood or "",
        property_obj.detected_neighborhood or "",
        property_obj.inferred_zone or "",
        "",
        intel.zone_name if intel else "",
        "BAJA",
        residual_reason(property_obj),
        "DUDOSO",
        location.latitude if location else None,
        location.longitude if location else None,
        location.provider if location else "",
        source_names(property_obj),
        agency_names(property_obj),
        url,
        "" if url else "Sin Listing asociado en la base",
        status_label(property_obj.status),
        property_obj.get_property_type_display(),
        property_obj.operation,
        property_obj.currency or "",
        float(property_obj.price) if property_obj.price is not None else None,
        property_obj.rooms,
        property_obj.bedrooms,
        float(property_obj.bathrooms) if property_obj.bathrooms is not None else None,
        float(property_obj.covered_area) if property_obj.covered_area is not None else None,
        float(property_obj.total_area) if property_obj.total_area is not None else None,
        property_obj.last_seen_at.isoformat() if property_obj.last_seen_at else "",
        residual_reason(property_obj),
        json.dumps(evidence, ensure_ascii=False, default=str),
        "",
        "",
    ]


def generate_excel(reviewed_notes):
    wb = Workbook()
    ws = wb.active
    ws.title = "Pendientes"
    ws.append(HEADERS)
    header_fill = PatternFill("solid", fgColor="1F4E78")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center")
    rows = list(residual_queryset())
    for property_obj in rows:
        ws.append(property_row(property_obj, reviewed_notes))

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    table = Table(displayName="PendientesFeedback20260627", ref=ws.dimensions)
    table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
    ws.add_table(table)
    decision_col = HEADERS.index("decision_manual") + 1
    validation = DataValidation(type="list", formula1='"APROBADO,RECHAZADO"', allow_blank=True)
    ws.add_data_validation(validation)
    if ws.max_row >= 2:
        validation.add(f"{get_column_letter(decision_col)}2:{get_column_letter(decision_col)}{ws.max_row}")

    widths = {
        "A": 12,
        "B": 42,
        "C": 52,
        "D": 28,
        "F": 24,
        "M": 22,
        "N": 22,
        "W": 58,
        "AJ": 24,
        "AK": 70,
        "AL": 18,
        "AM": 36,
    }
    for column in range(1, len(HEADERS) + 1):
        letter = get_column_letter(column)
        ws.column_dimensions[letter].width = widths.get(letter, 16)
    price_col = HEADERS.index("precio") + 1
    for row in range(2, ws.max_row + 1):
        ws.cell(row, price_col).number_format = "#,##0"
        ws.cell(row, HEADERS.index("descripcion_resumen") + 1).alignment = Alignment(wrap_text=True, vertical="top")
        ws.cell(row, HEADERS.index("evidencia") + 1).alignment = Alignment(wrap_text=True, vertical="top")

    summary_ws = wb.create_sheet("Resumen")
    summary_ws.append(["metrica", "valor"])
    summary_ws.append(["pendientes", len(rows)])
    summary_ws.append(["generado_por", SCRIPT_TAG])
    wb.save(OUTPUT_XLSX)
    return {"path": str(OUTPUT_XLSX), "pending_rows": len(rows), "columns": HEADERS}


def validate_excel():
    wb = load_workbook(OUTPUT_XLSX, data_only=True)
    ws = wb["Pendientes"]
    headers = [cell.value for cell in ws[1]]
    price_col = headers.index("precio") + 1
    decision_col = headers.index("decision_manual") + 1
    notes_col = headers.index("notas_manual") + 1
    price_types = Counter(type(ws.cell(row, price_col).value).__name__ for row in range(2, ws.max_row + 1))
    return {
        "sheets": wb.sheetnames,
        "rows": ws.max_row - 1,
        "auto_filter": ws.auto_filter.ref,
        "tables": list(ws.tables.keys()),
        "freeze_panes": ws.freeze_panes,
        "price_types": dict(price_types),
        "nonblank_decisions": sum(1 for row in range(2, ws.max_row + 1) if ws.cell(row, decision_col).value),
        "nonblank_notes": sum(1 for row in range(2, ws.max_row + 1) if ws.cell(row, notes_col).value),
        "validations": [
            {"type": dv.type, "formula1": dv.formula1, "sqref": str(dv.sqref)}
            for dv in ws.data_validations.dataValidation
        ],
        "missing_columns": [column for column in HEADERS if column not in headers],
    }


def cleanup_tmp(summary, apply):
    targets = []
    targets.extend(ROOT.glob("tmp/*.err.log"))
    targets.extend(ROOT.glob("tmp/*.out.log"))
    targets.append(ROOT / "tmp" / "curacion_zonas_pendientes_2026-06-19_filtros.xlsx.inspect.ndjson")
    pycache = ROOT / "tmp" / "__pycache__"
    removed = []
    for target in targets:
        if not target.exists():
            continue
        removed.append(str(target))
        if apply:
            target.unlink()
    if pycache.exists():
        removed.append(str(pycache))
        if apply:
            shutil.rmtree(pycache)
    summary["tmp_cleanup"] = removed


def sample(ids):
    out = []
    for property_obj in (
        Property.objects.select_related("location", "location_intelligence")
        .prefetch_related("listings__source")
        .filter(pk__in=ids)
        .order_by("id")
    ):
        location = location_or_none(property_obj)
        intel = intel_or_none(property_obj)
        out.append(
            {
                "id": property_obj.pk,
                "address": property_obj.address,
                "locality": property_obj.locality,
                "zone": property_obj.inferred_zone,
                "intel_zone": intel.zone_name if intel else "",
                "status": property_obj.status,
                "is_hidden": property_obj.is_hidden,
                "reviewed_at": property_obj.reviewed_at,
                "location": None
                if not location
                else {
                    "lat": location.latitude,
                    "lon": location.longitude,
                    "provider": location.provider,
                    "manual": location.manually_corrected,
                },
                "manual_overrides": property_obj.manual_overrides or {},
                "zone_evidence": property_obj.zone_inference_evidence or {},
                "corrected": bool(property_obj.data_manually_corrected_at),
                "active_listings": list(property_obj.listings.filter(active=True).values_list("id", flat=True)),
            }
        )
    return out


def apply_feedback(apply):
    summary = {
        "mode": "apply" if apply else "dry_run",
        "input_excel": str(INPUT_XLSX),
        "output_excel": str(OUTPUT_XLSX) if apply else "",
    }
    rows, decisions, notes = load_feedback()
    summary["feedback_decisions"] = dict(decisions)
    summary["approved_found"] = sorted(
        pid for pid in APPROVED_IDS if safe(rows.get(pid, {}).get("decision_manual")).upper() == "APROBADO"
    )
    summary["unexpected_approved"] = sorted(
        pid
        for pid, row in rows.items()
        if safe(row.get("decision_manual")).upper() == "APROBADO" and pid not in APPROVED_IDS
    )
    summary["before"] = current_counts()

    if apply:
        shutil.copy2(DB_PATH, BACKUP_PATH)
        summary["backup"] = str(BACKUP_PATH)
    else:
        summary["backup"] = ""

    apply_direct_fixes(summary, apply)
    apply_bustamante_propagation(summary, apply)

    for property_id in CONFIRM_REMOVE_IDS:
        property_obj = Property.objects.prefetch_related("listings__source").filter(pk=property_id).first()
        if property_obj:
            confirm_argencasas_removed(property_obj, summary, apply)

    sweep_argencasas(summary, apply)
    sweep_becerra(summary, apply)
    sweep_guarnieri(summary, apply)

    for property_id in sorted(NO_ACTION_IDS):
        property_obj = Property.objects.filter(pk=property_id).first()
        if property_obj:
            mark_reviewed_no_action(
                property_obj,
                notes.get(property_id, "revisado manualmente, sin datos utiles publicados"),
                summary,
                apply,
            )

    summary["after"] = current_counts()
    return summary, notes


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--apply", action="store_true")
    parser.add_argument("--generate-excel", action="store_true")
    parser.add_argument("--cleanup", action="store_true")
    parser.add_argument("--json-out")
    args = parser.parse_args()

    summary, reviewed_notes = apply_feedback(args.apply)
    if args.generate_excel:
        summary["excel"] = generate_excel(reviewed_notes)
        summary["excel_validation"] = validate_excel()
    else:
        summary["excel"] = None
        summary["excel_validation"] = None
    if args.cleanup:
        cleanup_tmp(summary, args.apply)
    summary["sample"] = sample(SAMPLE_IDS)
    output = json.dumps(summary, ensure_ascii=False, indent=2, default=str)
    print(output)
    if args.json_out:
        Path(args.json_out).write_text(output, encoding="utf-8")


if __name__ == "__main__":
    main()
