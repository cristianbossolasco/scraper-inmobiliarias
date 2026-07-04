#!/usr/bin/env python3
"""Build Hurlingham crime time series and GeoJSON artifacts.

The public official sources used here are municipal/department-level for
Hurlingham, except SAT-HD radio-centroid fields for homicide records. The
zone GeoJSON therefore repeats municipal metrics on every zone and marks
the spatial precision as low. The SAT-HD centroid layer is auxiliary only.
"""
from __future__ import annotations

import argparse
import csv
import hashlib
import io
import json
import math
import re
import sys
import unicodedata
import zipfile
from collections import Counter, defaultdict
from datetime import date, datetime, timezone
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook


WGS84 = "EPSG:4326"
DEFAULT_WINDOW_START = 2017
DEFAULT_WINDOW_END = 2024
CSV_COLUMNS = [
    "source",
    "dataset",
    "source_key",
    "source_role",
    "geo_level",
    "municipality",
    "province",
    "period_year",
    "period_month",
    "crime_group",
    "crime_type",
    "measure",
    "value",
    "source_file",
]
MONTHS = {
    "enero": 1,
    "febrero": 2,
    "marzo": 3,
    "abril": 4,
    "mayo": 5,
    "junio": 6,
    "julio": 7,
    "agosto": 8,
    "septiembre": 9,
    "setiembre": 9,
    "octubre": 10,
    "noviembre": 11,
    "diciembre": 12,
}
MUNICIPAL_NOTE = (
    "Public official crime data found for Hurlingham is aggregated at "
    "partido/municipio level. Values are attached to every zone for spatial "
    "joins but are not neighborhood-specific."
)
SAT_HD_CENTROID_NOTE = (
    "SAT-HD latitud_radio/longitud_radio are census-radio centroids, not exact "
    "incident coordinates."
)


def utc_now() -> str:
    return datetime.now(timezone.utc).isoformat()


def norm_text(value: object) -> str:
    if value is None:
        return ""
    text = str(value).strip().lower()
    text = unicodedata.normalize("NFKD", text)
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    return re.sub(r"\s+", " ", text)


def norm_col(value: object) -> str:
    text = norm_text(value)
    text = re.sub(r"[^a-z0-9]+", "_", text).strip("_")
    return text


def clean_cell(value: object) -> object:
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, str):
        return value.strip()
    return value


def parse_number(value: object) -> float | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        if isinstance(value, float) and not math.isfinite(value):
            return None
        return float(value)
    text = str(value).strip()
    if not text or norm_text(text) in {"na", "nan", "none", "null", "s/d", "sd"}:
        return None
    text = text.replace(".", "").replace(",", ".") if "," in text else text
    try:
        parsed = float(text)
    except ValueError:
        return None
    return parsed if math.isfinite(parsed) else None


def parse_int(value: object) -> int | None:
    parsed = parse_number(value)
    if parsed is None:
        return None
    return int(parsed)


def output_number(value: float | int | None) -> int | float | None:
    if value is None:
        return None
    if float(value).is_integer():
        return int(value)
    return round(float(value), 4)


def parse_month(value: object) -> int | None:
    if value is None:
        return None
    parsed = parse_int(value)
    if parsed is not None:
        return parsed
    return MONTHS.get(norm_text(value))


def stable_json_dump(path: Path, payload: dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")


def normalize_header(header: Iterable[object]) -> list[str]:
    seen: Counter[str] = Counter()
    normalized = []
    for index, raw_col in enumerate(header):
        base = norm_col(raw_col) or f"unnamed_{index}"
        seen[base] += 1
        normalized.append(base if seen[base] == 1 else f"{base}_{seen[base]}")
    return normalized


def normalize_record(row: dict[str, object]) -> dict[str, object]:
    output: dict[str, object] = {}
    for index, (key, value) in enumerate(row.items()):
        col = norm_col(key) or f"unnamed_{index}"
        if col in output:
            col = f"{col}_{index}"
        output[col] = clean_cell(value)
    return output


def sniff_csv_dialect(data: bytes) -> csv.Dialect:
    sample = data[:8192].decode("utf-8", errors="ignore")
    try:
        return csv.Sniffer().sniff(sample, delimiters=",;\t|")
    except csv.Error:
        delimiter = ";" if data[:8192].count(b";") > data[:8192].count(b",") else ","
        class Fallback(csv.excel):
            pass

        Fallback.delimiter = delimiter
        return Fallback


def iter_csv_records(raw: Any, dialect: csv.Dialect) -> Iterable[dict[str, object]]:
    text = io.TextIOWrapper(raw, encoding="utf-8-sig", errors="replace", newline="")
    reader = csv.reader(text, dialect=dialect)
    try:
        header = next(reader)
    except StopIteration:
        return
    columns = normalize_header(header)
    for row in reader:
        if not row:
            continue
        yield {
            columns[index]: clean_cell(value)
            for index, value in enumerate(row)
            if index < len(columns)
        }


def read_xlsx_records_from_bytes(data: bytes) -> list[dict[str, object]]:
    workbook = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    sheet = workbook.active
    rows = sheet.iter_rows(values_only=True)
    try:
        header = next(rows)
    except StopIteration:
        return []
    columns = normalize_header(header)
    records = []
    for row in rows:
        if not any(value is not None and str(value).strip() for value in row):
            continue
        records.append({columns[index]: clean_cell(value) for index, value in enumerate(row) if index < len(columns)})
    return records


def iter_records(path: Path) -> Iterable[dict[str, object]]:
    suffix = path.suffix.lower()
    if suffix == ".zip":
        with zipfile.ZipFile(path) as archive:
            names = sorted(name for name in archive.namelist() if not name.endswith("/"))
            for name in names:
                lower = name.lower()
                if not lower.endswith((".csv", ".xlsx", ".xls")):
                    continue
                if lower.endswith(".csv"):
                    with archive.open(name) as sample_raw:
                        dialect = sniff_csv_dialect(sample_raw.read(8192))
                    with archive.open(name) as raw:
                        yield from iter_csv_records(raw, dialect)
                else:
                    yield from read_xlsx_records_from_bytes(archive.read(name))
        return
    if suffix == ".csv":
        with path.open("rb") as sample_raw:
            dialect = sniff_csv_dialect(sample_raw.read(8192))
        with path.open("rb") as raw:
            yield from iter_csv_records(raw, dialect)
        return
    if suffix in {".xlsx", ".xls"}:
        yield from read_xlsx_records_from_bytes(path.read_bytes())
        return
    raise ValueError(f"Unsupported source file extension: {path}")


def require_columns(row: dict[str, object], required: set[str], label: str) -> None:
    missing = required - set(row.keys())
    if missing:
        raise ValueError(f"{label} schema changed; missing columns: {sorted(missing)}")


def is_hurlingham_row(row: dict[str, object]) -> bool:
    province_values = [row.get(col) for col in ("provincia_nombre", "provincia", "jurisdiccion")]
    province_texts = [norm_text(value) for value in province_values if value not in (None, "")]
    if province_texts and not any(text == "buenos aires" for text in province_texts):
        return False

    geo_cols = (
        "departamento_nombre",
        "departamento",
        "municipio",
        "municipio_nombre",
        "partido",
        "localidad_nombre",
    )
    return any("hurlingham" in norm_text(row.get(col)) for col in geo_cols)


def classify_crime_group(value: object) -> str:
    text = norm_text(value)
    excluding_vehicle = "excluir de automotores" in text or "excluir de automotores y motocicletas" in text
    vehicle = (
        ("automotor" in text or "motocicleta" in text or "vehiculo" in text or "vehiculos" in text)
        and not excluding_vehicle
    )
    if "homicidio" in text:
        return "homicidio"
    if "lesion" in text:
        return "lesiones"
    if "violacion" in text or "sexual" in text:
        return "integridad_sexual"
    if "hurto" in text and vehicle:
        return "hurto_vehiculo"
    if "robo" in text and vehicle:
        return "robo_vehiculo"
    if "hurto" in text:
        return "hurto"
    if "robo" in text:
        return "robo"
    if "amenaza" in text:
        return "amenazas"
    if "secuestro" in text:
        return "secuestro"
    if "extorsion" in text:
        return "extorsion"
    return "otros"


def source_file_for(source: dict[str, Any], raw_dir: Path) -> Path:
    return raw_dir / str(source["target"])


def add_standard_row(
    rows: list[dict[str, Any]],
    source: dict[str, Any],
    source_path: Path,
    *,
    year: object,
    month: object,
    crime_type: object,
    measure: str,
    value: object,
    geo_level: str = "municipio",
) -> None:
    parsed_year = parse_int(year)
    parsed_value = parse_number(value)
    if parsed_year is None or parsed_value is None:
        return
    rows.append(
        {
            "source": source["name"],
            "dataset": source["name"],
            "source_key": source["key"],
            "source_role": source["role"],
            "geo_level": geo_level,
            "municipality": "Hurlingham",
            "province": "Buenos Aires",
            "period_year": parsed_year,
            "period_month": parse_month(month),
            "crime_group": classify_crime_group(crime_type),
            "crime_type": str(crime_type or "").strip() or "sin_tipo",
            "measure": measure,
            "value": output_number(parsed_value),
            "source_file": str(source_path),
        }
    )


def standardize_snic_monthly(
    source: dict[str, Any],
    source_path: Path,
    records: Iterable[dict[str, object]],
    code_name_map: dict[str, str],
) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    required = {"anio", "mes", "departamento_nombre", "codigo_delito_snic_id", "codigo_delito_snic_nombre"}
    seen = False
    for row in records:
        if not seen:
            require_columns(row, required, "SNIC monthly")
            seen = True
        if not is_hurlingham_row(row):
            continue
        code = str(row.get("codigo_delito_snic_id") or "").strip()
        crime_type = row.get("codigo_delito_snic_nombre") or f"SNIC code {code}"
        if code:
            code_name_map[code] = str(crime_type)
        for col in ("cantidad_hechos", "cantidad_victimas", "cantidad_victimas_masc", "cantidad_victimas_fem", "cantidad_victimas_sd"):
            if col in row:
                add_standard_row(
                    rows,
                    source,
                    source_path,
                    year=row.get("anio"),
                    month=row.get("mes"),
                    crime_type=crime_type,
                    measure=col,
                    value=row.get(col),
                )
    if not seen:
        raise ValueError(f"No records found in {source_path}")
    return rows


def standardize_snic_annual(
    source: dict[str, Any],
    source_path: Path,
    records: Iterable[dict[str, object]],
    code_name_map: dict[str, str],
) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    required = {"anio", "departamento_nombre", "codigo_delito_snic_id"}
    seen = False
    for row in records:
        if not seen:
            require_columns(row, required, "SNIC annual")
            seen = True
        if not is_hurlingham_row(row):
            continue
        code = str(row.get("codigo_delito_snic_id") or "").strip()
        crime_type = code_name_map.get(code, f"SNIC code {code}")
        for col in ("cantidad_hechos", "cantidad_victimas", "tasa_hechos", "tasa_victimas"):
            if col in row:
                add_standard_row(
                    rows,
                    source,
                    source_path,
                    year=row.get("anio"),
                    month=None,
                    crime_type=crime_type,
                    measure=col,
                    value=row.get(col),
                )
    if not seen:
        raise ValueError(f"No records found in {source_path}")
    return rows


def standardize_sat_property(
    source: dict[str, Any],
    source_path: Path,
    records: Iterable[dict[str, object]],
) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    required = {"anio", "mes", "departamento_nombre", "nombre_delito_sat_prop", "cantidad_hechos"}
    seen = False
    measures = [
        "cantidad_hechos",
        "cantidad_hechos_lugar_via_publ",
        "cantidad_hechos_lugar_establec",
        "cantidad_hechos_lugar_dom_part",
        "cantidad_hechos_lugar_sd",
        "cantidad_hechos_arma_de_fuego",
        "cantidad_hechos_arma_otra",
        "cantidad_hechos_arma_sin_arma",
        "cantidad_hechos_arma_sd",
        "cantidad_hechos_origen_denuncia",
        "cantidad_hechos_origen_intervenc",
        "cantidad_hechos_origen_orden_jud",
        "cantidad_hechos_origen_otro",
    ]
    for row in records:
        if not seen:
            require_columns(row, required, "SAT property")
            seen = True
        if not is_hurlingham_row(row):
            continue
        crime_type = row.get("nombre_delito_sat_prop")
        for measure in measures:
            if measure in row:
                add_standard_row(
                    rows,
                    source,
                    source_path,
                    year=row.get("anio"),
                    month=row.get("mes"),
                    crime_type=crime_type,
                    measure=measure,
                    value=row.get(measure),
                )
    if not seen:
        raise ValueError(f"No records found in {source_path}")
    return rows


def standardize_pba_municipal(
    source: dict[str, Any],
    source_path: Path,
    records: Iterable[dict[str, object]],
) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    required = {"ano", "municipio", "tipo", "valor"}
    seen = False
    for row in records:
        if not seen:
            require_columns(row, required, "PBA municipal")
            seen = True
        if not is_hurlingham_row(row):
            continue
        add_standard_row(
            rows,
            source,
            source_path,
            year=row.get("ano"),
            month=None,
            crime_type=row.get("tipo"),
            measure="cantidad_victimas" if "victimas" in norm_text(source.get("name")) else "cantidad_hechos",
            value=row.get("valor"),
        )
    if not seen:
        raise ValueError(f"No records found in {source_path}")
    return rows


def dedupe_sat_hd_events(records: Iterable[dict[str, object]]) -> list[dict[str, object]]:
    events: dict[str, dict[str, object]] = {}
    required = {"id_hecho", "tipo_persona", "departamento_nombre", "anio", "mes", "cant_vic"}
    seen = False
    for row in records:
        if not seen:
            require_columns(row, required, "SAT-HD")
            seen = True
        if not is_hurlingham_row(row):
            continue
        if not norm_text(row.get("tipo_persona")).startswith("victima"):
            continue
        event_id = str(row.get("id_hecho") or "").strip()
        if not event_id or event_id in events:
            continue
        events[event_id] = row
    if not seen:
        raise ValueError("No records found in SAT-HD source")
    return [events[key] for key in sorted(events, key=lambda item: (parse_int(item) or 0, item))]


def standardize_sat_hd(
    source: dict[str, Any],
    source_path: Path,
    records: Iterable[dict[str, object]],
) -> tuple[list[dict[str, Any]], list[dict[str, object]]]:
    events = dedupe_sat_hd_events(records)
    rows: list[dict[str, Any]] = []
    for event in events:
        for measure, value in (("cantidad_hechos_dedup", 1), ("cantidad_victimas", event.get("cant_vic"))):
            add_standard_row(
                rows,
                source,
                source_path,
                year=event.get("anio"),
                month=event.get("mes"),
                crime_type="Homicidios dolosos SAT-HD",
                measure=measure,
                value=value,
            )
    return rows, events


def point_in_ring(latitude: float, longitude: float, ring: list[list[float]]) -> bool:
    inside = False
    j = len(ring) - 1
    for i, coord in enumerate(ring):
        lng_i, lat_i = coord[:2]
        lng_j, lat_j = ring[j][:2]
        intersects = ((lat_i > latitude) != (lat_j > latitude)) and (
            longitude < (lng_j - lng_i) * (latitude - lat_i) / ((lat_j - lat_i) or 1e-12) + lng_i
        )
        if intersects:
            inside = not inside
        j = i
    return inside


def point_in_geometry(latitude: float, longitude: float, geometry: dict[str, Any]) -> bool:
    geom_type = geometry.get("type")
    coords = geometry.get("coordinates") or []
    polygons = [coords] if geom_type == "Polygon" else coords if geom_type == "MultiPolygon" else []
    for polygon in polygons:
        if not polygon or not point_in_ring(latitude, longitude, polygon[0]):
            continue
        holes = polygon[1:]
        if any(point_in_ring(latitude, longitude, hole) for hole in holes):
            continue
        return True
    return False


def assign_zone(latitude: float | None, longitude: float | None, zone_features: list[dict[str, Any]]) -> str | None:
    if latitude is None or longitude is None:
        return None
    for feature in zone_features:
        if point_in_geometry(latitude, longitude, feature.get("geometry") or {}):
            return str((feature.get("properties") or {}).get("zone_name") or "")
    return None


def load_zones(path: Path) -> dict[str, Any]:
    payload = json.loads(path.read_text(encoding="utf-8"))
    if payload.get("type") != "FeatureCollection":
        raise ValueError(f"Zones file is not a FeatureCollection: {path}")
    features = payload.get("features") or []
    if not features:
        raise ValueError(f"Expected at least one zone feature, got {len(features)}")
    bad = [
        (feature.get("properties") or {}).get("zone_name")
        for feature in features
        if (feature.get("geometry") or {}).get("type") not in {"Polygon", "MultiPolygon"}
    ]
    if bad:
        raise ValueError(f"Zone geometries must be Polygon/MultiPolygon: {bad[:5]}")
    return payload


def metric_rows(rows: list[dict[str, Any]], *, start_year: int, end_year: int) -> list[dict[str, Any]]:
    output = []
    for row in rows:
        year = parse_int(row.get("period_year"))
        if year is None or year < start_year or year > end_year:
            continue
        output.append(row)
    return output


def sum_values(
    rows: list[dict[str, Any]],
    *,
    source_key: str | None = None,
    source_role: str | None = None,
    measure: str | None = None,
    crime_group: str | set[str] | None = None,
) -> float:
    total = 0.0
    groups = {crime_group} if isinstance(crime_group, str) else crime_group
    for row in rows:
        if source_key and row.get("source_key") != source_key:
            continue
        if source_role and row.get("source_role") != source_role:
            continue
        if measure and row.get("measure") != measure:
            continue
        if groups is not None and row.get("crime_group") not in groups:
            continue
        value = parse_number(row.get("value"))
        if value is not None:
            total += value
    return total


def source_rows(rows: list[dict[str, Any]], source_key: str) -> list[dict[str, Any]]:
    return [row for row in rows if row.get("source_key") == source_key]


def build_metrics(rows: list[dict[str, Any]], *, start_year: int, end_year: int, generated_at: str) -> dict[str, Any]:
    window_rows = metric_rows(rows, start_year=start_year, end_year=end_year)
    all_years = [parse_int(row.get("period_year")) for row in rows]
    all_years = [year for year in all_years if year is not None]
    sources = sorted({str(row["source"]) for row in window_rows})

    sat_property = source_rows(window_rows, "sat_propiedad")
    snic_monthly = source_rows(window_rows, "snic_departamentos_mensual")
    sat_hd = source_rows(window_rows, "sat_homicidios")

    robbery = sum_values(sat_property, measure="cantidad_hechos", crime_group="robo")
    theft = sum_values(sat_property, measure="cantidad_hechos", crime_group="hurto")
    vehicle_robbery = sum_values(sat_property, measure="cantidad_hechos", crime_group="robo_vehiculo")
    vehicle_theft = sum_values(sat_property, measure="cantidad_hechos", crime_group="hurto_vehiculo")

    metrics = {
        "crime_data_scope": "municipio",
        "crime_spatial_precision": "low",
        "crime_municipality": "Hurlingham",
        "crime_municipality_localities": ["Hurlingham", "Villa Tesei", "William C. Morris"],
        "crime_metric_window_start_year": start_year,
        "crime_metric_window_end_year": end_year,
        "crime_timeseries_start_year": min(all_years) if all_years else None,
        "crime_timeseries_end_year": max(all_years) if all_years else None,
        "reported_crimes_total": output_number(
            sum_values(snic_monthly, measure="cantidad_hechos")
        ),
        "reported_property_crime_count": output_number(
            sum_values(sat_property, measure="cantidad_hechos")
        ),
        "reported_robbery_count": output_number(robbery),
        "reported_theft_count": output_number(theft),
        "reported_vehicle_robbery_count": output_number(vehicle_robbery),
        "reported_vehicle_theft_count": output_number(vehicle_theft),
        "reported_vehicle_crime_count": output_number(vehicle_robbery + vehicle_theft),
        "reported_homicide_count": output_number(
            sum_values(sat_hd, measure="cantidad_hechos_dedup", crime_group="homicidio")
        ),
        "reported_homicide_victim_count": output_number(
            sum_values(sat_hd, measure="cantidad_victimas", crime_group="homicidio")
        ),
        "reported_injury_count": output_number(
            sum_values(snic_monthly, measure="cantidad_hechos", crime_group="lesiones")
        ),
        "reported_sexual_integrity_count": output_number(
            sum_values(snic_monthly, measure="cantidad_hechos", crime_group="integridad_sexual")
        ),
        "reported_crime_sources": sources,
        "reported_crime_primary_sources": [
            "SNIC departamentos mensual por sexo",
            "SAT Propiedad 2017-2024",
            "SAT Homicidios dolosos",
        ],
        "reported_crime_data_note": MUNICIPAL_NOTE,
        "reported_crime_methodology": (
            "General totals use SNIC monthly department data. Property metrics use SAT Propiedad. "
            "Homicide counts use deduplicated SAT-HD victim rows by id_hecho."
        ),
        "crime_generated_at": generated_at,
    }
    return metrics


def build_zone_geojson(zones: dict[str, Any], metrics: dict[str, Any]) -> dict[str, Any]:
    features = []
    for feature in zones["features"]:
        props = dict(feature.get("properties") or {})
        props.update(metrics)
        features.append(
            {
                "type": "Feature",
                "properties": props,
                "geometry": feature.get("geometry"),
            }
        )
    return {
        "type": "FeatureCollection",
        "name": "crime_zones_hurlingham",
        "metadata": {
            "generated_at": metrics["crime_generated_at"],
            "crs": WGS84,
            "scope": "Partido de Hurlingham, Buenos Aires, Argentina",
            "spatial_precision": "low",
            "notes": MUNICIPAL_NOTE,
        },
        "features": features,
    }


def build_homicide_points(
    events: list[dict[str, object]],
    source: dict[str, Any],
    zone_features: list[dict[str, Any]],
    generated_at: str,
) -> dict[str, Any]:
    features = []
    no_coordinate_count = 0
    no_zone_count = 0
    for event in events:
        lat = parse_number(event.get("latitud_radio"))
        lon = parse_number(event.get("longitud_radio"))
        if lat is None or lon is None:
            no_coordinate_count += 1
            continue
        zone_name = assign_zone(lat, lon, zone_features)
        if not zone_name:
            no_zone_count += 1
        event_id = str(event.get("id_hecho") or "").strip()
        victims = output_number(parse_number(event.get("cant_vic")) or 0)
        props = {
            "id": f"sat_hd_hurlingham_{event_id}",
            "source": source["name"],
            "source_key": source["key"],
            "municipality": "Hurlingham",
            "province": "Buenos Aires",
            "crime_group": "homicidio",
            "crime_type": "Homicidios dolosos SAT-HD",
            "period_year": parse_int(event.get("anio")),
            "period_month": parse_month(event.get("mes")),
            "id_hecho": event_id,
            "victims_count": victims,
            "radio_censal": event.get("radio_censal") or None,
            "tipo_lugar": event.get("tipo_lugar") or None,
            "clase_arma": event.get("clase_arma") or None,
            "assigned_zone_name": zone_name,
            "is_exact_location": False,
            "spatial_precision": "radio_censal_centroid",
            "source_note": SAT_HD_CENTROID_NOTE,
            "generated_at": generated_at,
        }
        features.append(
            {
                "type": "Feature",
                "id": props["id"],
                "properties": props,
                "geometry": {"type": "Point", "coordinates": [round(lon, 7), round(lat, 7)]},
            }
        )
    features.sort(
        key=lambda feature: (
            feature["properties"].get("period_year") or 0,
            feature["properties"].get("period_month") or 0,
            str(feature["properties"].get("id_hecho") or ""),
        )
    )
    return {
        "type": "FeatureCollection",
        "name": "crime_homicide_radio_points_hurlingham",
        "metadata": {
            "generated_at": generated_at,
            "crs": WGS84,
            "source": source["name"],
            "scope": "Partido de Hurlingham, Buenos Aires, Argentina",
            "spatial_precision": "radio_censal_centroid",
            "is_exact_location": False,
            "notes": SAT_HD_CENTROID_NOTE,
            "event_count_with_coordinates": len(features),
            "event_count_without_coordinates": no_coordinate_count,
            "event_count_without_zone_match": no_zone_count,
        },
        "features": features,
    }


def write_timeseries_csv(path: Path, rows: list[dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    rows = sorted(
        rows,
        key=lambda row: (
            str(row.get("source_key") or ""),
            parse_int(row.get("period_year")) or 0,
            parse_int(row.get("period_month")) or 0,
            str(row.get("crime_type") or ""),
            str(row.get("measure") or ""),
        ),
    )
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=CSV_COLUMNS)
        writer.writeheader()
        for row in rows:
            writer.writerow({column: "" if row.get(column) is None else row.get(column) for column in CSV_COLUMNS})


def file_sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def load_manifest(raw_dir: Path) -> dict[str, Any]:
    manifest_path = raw_dir / "download_manifest.json"
    if not manifest_path.exists():
        return {"sources": []}
    return json.loads(manifest_path.read_text(encoding="utf-8"))


def validate_outputs(
    zone_geojson: dict[str, Any],
    homicide_points: dict[str, Any],
    rows: list[dict[str, Any]],
    metrics: dict[str, Any],
) -> dict[str, Any]:
    if zone_geojson.get("type") != "FeatureCollection":
        raise ValueError("crime_zones_hurlingham.geojson must be a FeatureCollection")
    features = zone_geojson.get("features") or []
    if not features:
        raise ValueError(f"Expected at least one crime zone feature, got {len(features)}")
    for feature in features:
        geometry_type = (feature.get("geometry") or {}).get("type")
        props = feature.get("properties") or {}
        if geometry_type not in {"Polygon", "MultiPolygon"}:
            raise ValueError(f"Invalid geometry type in crime zones: {geometry_type}")
        if props.get("crime_data_scope") != "municipio" or props.get("crime_spatial_precision") != "low":
            raise ValueError("Crime zone feature does not carry municipal/low precision markers")
        for key, value in props.items():
            if key.startswith("reported_") and key.endswith(("_count", "_total")):
                parsed = parse_number(value)
                if parsed is not None and parsed < 0:
                    raise ValueError(f"Negative metric found: {key}={value}")
    if homicide_points.get("type") != "FeatureCollection":
        raise ValueError("crime_homicide_radio_points_hurlingham.geojson must be a FeatureCollection")
    for feature in homicide_points.get("features") or []:
        props = feature.get("properties") or {}
        if props.get("is_exact_location") is not False:
            raise ValueError("SAT-HD auxiliary point is missing is_exact_location=false")
    non_hurlingham = [
        row
        for row in rows
        if row.get("municipality") != "Hurlingham" or row.get("province") != "Buenos Aires"
    ]
    if non_hurlingham:
        raise ValueError(f"Standardized CSV contains non-Hurlingham rows: {non_hurlingham[:3]}")
    return {
        "crime_zone_features": len(features),
        "homicide_radio_point_features": len(homicide_points.get("features") or []),
        "standardized_rows": len(rows),
        "metric_window": [metrics["crime_metric_window_start_year"], metrics["crime_metric_window_end_year"]],
    }


def build_summary(
    *,
    config_path: Path,
    raw_dir: Path,
    manifest: dict[str, Any],
    sources: list[dict[str, Any]],
    source_counts: dict[str, int],
    metrics: dict[str, Any],
    validation: dict[str, Any],
    outputs: dict[str, Path],
    generated_at: str,
) -> dict[str, Any]:
    raw_files = {}
    for source in sources:
        path = raw_dir / str(source["target"])
        if path.exists():
            raw_files[source["key"]] = {
                "path": str(path),
                "size_bytes": path.stat().st_size,
                "sha256": file_sha256(path),
            }
    return {
        "generated_at": generated_at,
        "config": str(config_path),
        "raw_dir": str(raw_dir),
        "outputs": {key: str(path) for key, path in outputs.items()},
        "metrics": metrics,
        "validation": validation,
        "source_row_counts": dict(sorted(source_counts.items())),
        "raw_files": raw_files,
        "download_manifest": manifest,
        "methodology_notes": [
            MUNICIPAL_NOTE,
            SAT_HD_CENTROID_NOTE,
            "No single crime score is generated.",
            "No neighborhood crime distribution is inferred from municipal totals.",
        ],
    }


def standardize_sources(
    sources: list[dict[str, Any]],
    raw_dir: Path,
) -> tuple[list[dict[str, Any]], list[dict[str, object]], dict[str, int]]:
    all_rows: list[dict[str, Any]] = []
    homicide_events: list[dict[str, object]] = []
    source_counts: dict[str, int] = {}
    code_name_map: dict[str, str] = {}

    source_by_key = {source["key"]: source for source in sources}
    ordered_keys = [
        "snic_departamentos_mensual",
        "snic_departamentos_anual",
        "sat_propiedad",
        "sat_homicidios",
        "pba_propiedad_municipio",
        "pba_personas_municipio",
    ]
    ordered_sources = [source_by_key[key] for key in ordered_keys if key in source_by_key]
    ordered_sources.extend(source for source in sources if source["key"] not in ordered_keys)

    for source in ordered_sources:
        path = source_file_for(source, raw_dir)
        if not path.exists():
            if source.get("required", True):
                raise FileNotFoundError(f"Required raw source is missing: {path}")
            continue
        print(f"Reading {path}")
        records = iter_records(path)
        if source["key"] == "snic_departamentos_mensual":
            rows = standardize_snic_monthly(source, path, records, code_name_map)
        elif source["key"] == "snic_departamentos_anual":
            rows = standardize_snic_annual(source, path, records, code_name_map)
        elif source["key"] == "sat_propiedad":
            rows = standardize_sat_property(source, path, records)
        elif source["key"] == "sat_homicidios":
            rows, homicide_events = standardize_sat_hd(source, path, records)
        elif source["key"] in {"pba_propiedad_municipio", "pba_personas_municipio"}:
            rows = standardize_pba_municipal(source, path, records)
        else:
            raise ValueError(f"Unsupported source key: {source['key']}")
        source_counts[source["key"]] = len(rows)
        print(f"  -> standardized rows: {len(rows):,}")
        all_rows.extend(rows)

    if not all_rows:
        raise ValueError("No standardized Hurlingham crime rows were generated.")
    if not homicide_events:
        raise ValueError("No SAT-HD Hurlingham homicide events were generated.")
    return all_rows, homicide_events, source_counts


def main() -> int:
    parser = argparse.ArgumentParser(description="Build Hurlingham crime GeoJSON artifacts.")
    parser.add_argument("--config", default="config/crime_sources.json")
    parser.add_argument("--raw-dir", default=None)
    parser.add_argument("--zones", default="data/geo/Zonas_Hurlingham_polygons.geojson")
    parser.add_argument("--out", default="data/geo/crime_zones_hurlingham.geojson")
    parser.add_argument("--csv", default="data/geo/crime_hurlingham_municipality_timeseries.csv")
    parser.add_argument("--homicide-points", default="data/geo/crime_homicide_radio_points_hurlingham.geojson")
    parser.add_argument("--summary", default="data/geo/crime_summary_hurlingham.json")
    parser.add_argument("--start-year", type=int, default=DEFAULT_WINDOW_START)
    parser.add_argument("--end-year", type=int, default=DEFAULT_WINDOW_END)
    parser.add_argument("--generated-at", default=None)
    args = parser.parse_args()

    try:
        config_path = Path(args.config)
        config = json.loads(config_path.read_text(encoding="utf-8"))
        raw_dir = Path(args.raw_dir or config.get("raw_dir") or "data/raw/crime")
        sources = list(config.get("sources") or [])
        generated_at = args.generated_at or utc_now()

        zones = load_zones(Path(args.zones))
        rows, homicide_events, source_counts = standardize_sources(sources, raw_dir)
        metrics = build_metrics(rows, start_year=args.start_year, end_year=args.end_year, generated_at=generated_at)
        zone_geojson = build_zone_geojson(zones, metrics)
        sat_hd_source = next(source for source in sources if source["key"] == "sat_homicidios")
        homicide_points = build_homicide_points(homicide_events, sat_hd_source, zones["features"], generated_at)

        outputs = {
            "timeseries_csv": Path(args.csv),
            "crime_zones_geojson": Path(args.out),
            "homicide_radio_points_geojson": Path(args.homicide_points),
            "summary_json": Path(args.summary),
        }
        write_timeseries_csv(outputs["timeseries_csv"], rows)
        stable_json_dump(outputs["crime_zones_geojson"], zone_geojson)
        stable_json_dump(outputs["homicide_radio_points_geojson"], homicide_points)
        validation = validate_outputs(zone_geojson, homicide_points, rows, metrics)
        summary = build_summary(
            config_path=config_path,
            raw_dir=raw_dir,
            manifest=load_manifest(raw_dir),
            sources=sources,
            source_counts=source_counts,
            metrics=metrics,
            validation=validation,
            outputs=outputs,
            generated_at=generated_at,
        )
        stable_json_dump(outputs["summary_json"], summary)

        print(json.dumps({"outputs": {k: str(v) for k, v in outputs.items()}, "validation": validation}, indent=2))
        return 0
    except Exception as exc:  # noqa: BLE001 - CLI should fail with a direct, useful message.
        print(f"ERROR {exc}", file=sys.stderr)
        return 2


if __name__ == "__main__":
    raise SystemExit(main())
